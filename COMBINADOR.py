import pyodbc
import openpyxl
from dataclasses import dataclass
from typing import Optional
import sys
import os 

#CONFIGURACION DE LA CONEXION

DB_LOCAL = {
    "server": r"localhost\SQLEXPRESS",
    "database": "MODULO_5",
    "driver": "ODBC Driver 17 for SQL Server",
}

DB_PROD = {
    "server":   "agpcol.database.windows.net",
    "database": "agpc-productivity",
    "driver":   "ODBC Driver 17 for SQL Server",
    "user":     "Consulta",
    "password": "@GPgl4$$2021",
}
#bro cambia estas rutas por el excel que descargues y donde lo guardes 

RUTA_EXCEL_FORMULAS = (
    r"C:\Users\abotero\OneDrive - AGP GROUP"
    r"\Documentos\MODULO_5\formulas_activas.xlsx"
)

MAPA_MERCADO_HOJA = {
    "MÉXICO": "México",
    "MEXICO": "México",
    "ESTADOS UNIDOS": "EEUU",
}

HOJA_DEFAULT = "CAM"

def hoja_para_mercado(mercado:str) -> str:
    return MAPA_MERCADO_HOJA.get(mercado.strip().upper(),HOJA_DEFAULT)

PERFILES_HOJA = {

    "México": {
        "col_cod_pieza": "Cod Pieza",
        "col_formula":   "Formula",
        "col_acero":     "¿Tiene acero?",
        "col_activo":    "Activo",  # filtrar solo filas Activa
        "acero_default": "NO",                 # no aplica (col_acero existe)
        "colores": [
            "01 - Green Light",
            "05-Gray Light PC",
            "06-Gray Light Glass",
            "10 -Gray Medium PC",
            "13- Gray Dark Glass",
            "18-Gray Medium Glass",
            "19-Gray Light Automotive",
            "20-Gray Medium Automotive+PC color 35%.",   # nombre largo MX
            "21-Gray Dark Automotive+PC color 22%.",     # nombre largo MX
            "22-G2 Gray Medium Automotive",
            "23-G2 Dark Medium Automotive",
        ],
    },

    "CAM": {
        "col_cod_pieza": "Cod Pieza",
        "col_formula":   "Formula",
        "col_acero":     "¿Tiene acero?",
        "col_activo":    None,    # CAM no tiene columna Activo — todas activas
        "acero_default": "NO",
        "colores": [
            "01 - Green Light",
            "05-Gray Light PC",
            "06-Gray Light Glass",
            "10 -Gray Medium PC",
            "13- Gray Dark Glass",
            "18-Gray Medium Glass",
            "19-Gray Light Automotive",
            "20-Gray Medium Automotive",    # nombre corto en CAM
            "21-Gray Dark Automotive",      # nombre corto en CAM
            "22-G2 Gray Medium Automotive",
            "23-G2 Dark Medium Automotive",
        ],
    },

    "EEUU": {
        # EEUU no tiene Cod Pieza — se identifica por Producto GENESIS.
        # PENDIENTE confirmar con equipo: ¿cómo se filtra por pieza en EEUU?
        "col_cod_pieza": "Producto GENESIS",  # provisional — confirmar
        "col_formula":   "Formula",
        "col_acero":     None,       # EEUU no tiene columna de acero
        "col_activo":    None,       # EEUU no tiene columna Activo
        "acero_default": "NO",       # provisional — confirmar con equipo
        "colores": [
            "00 - White",            # color extra solo en EEUU
            "01 - Green Light",
            "05-Gray Light PC",
            "06-Gray Light Glass",
            "10 -Gray Medium PC",
            "13- Gray Dark Glass",
            "18-Gray Medium Glass",
            "19-Gray Light Automotive",
            "20-Gray Medium Automotive",
            "21-Gray Dark Automotive",
            "22-G2 Gray Medium Automotive",
            "23-G2 Dark Medium Automotive",
        ],
    },
}

def get_conn_str(prod: bool = False) -> str :
    if prod:
        c = DB_PROD
        return (
            f"DRIVER={{{c['driver']}}};"
            f"SERVER={c['server']};"
            f"DATABASE={c['database']};"
            f"UID={c['user']};"
            f"PWD={c['password']};"
        )
    c = DB_LOCAL
    return (
        f"DRIVER={{{c['driver']}}};"
        f"SERVER={c['server']};"
        f"DATABASE={c['database']};"
        "Trusted_Connection=yes;"
    )

#ESTRUCTURA DE DATOSSSSS

@dataclass
class FilaFormula:
    cod_pieza: str
    pieza: str
    formula: str
    acero_variante: str
    colores: list


@dataclass
class TuplaVariante:
  
    zfer_origen:    str
    mercado:        str
    cod_pieza:      str
    tipo_peca:      str
    formula:        str        
    color_codigo:   str        
    acero_variante: str        
    estado:         str = "ACTIVA"
    # Herramentales heredados del pedido base (no cambian entre tuplas)
    zpla:           str = ""
    hoja_ruta:      str = ""
    plano:          str = ""
    url_plano:      str = ""
    archivo_corte:  str = ""

    def __str__(self):
        return (
            f"[{self.mercado}] {self.cod_pieza} | "
            f"{self.formula} | {self.color_codigo} | "
            f"Acero:{self.acero_variante} → {self.estado}"
        )

    def _extraer_codigo_color(self, color_nombre: str) -> str:
        """Extrae el código del color del nombre. Ej: '13- Gray Dark Glass' -> '13'"""
        if not color_nombre:
            return ""
        return color_nombre.split("-")[0].strip()

    def es_pedido_base(self, formula_base, color_base, acero_base) -> bool:
        """True si esta tupla es exactamente el pedido base (ya existe en SAP)."""
        codigo_color = self._extraer_codigo_color(self.color_codigo)
        codigo_base = self._extraer_codigo_color(color_base)
        return (
            self.formula        == formula_base and
            codigo_color        == codigo_base   and
            self.acero_variante == acero_base
        )

#AQUI HAGO EL LECTOR DE EXCEL DE FORMULAS BRO 

def _normalizar_acero(valor) -> str:
    """
    Convierte el texto del Excel al código interno de acero.
    El Excel dice "NO", "SI (SN)", "SI (SP)".
    Nosotros usamos "NO", "SN", "SP".
    """
    v = str(valor or "").strip().upper()
    if "SN" in v:   return "SN"
    if "SP" in v:   return "SP"
    return "NO"


def leer_excel_formulas(
    ruta_excel: str,
    nombre_hoja: str,
    formula_filtro: str = None,
    cod_pieza_filtro: str = None,
) -> list:
    if not os.path.exists(ruta_excel):
        raise FileNotFoundError(
            f"Exel de formulas no encontrado papi en:\n  {ruta_excel}\n"
            f"Descargalo desde sahrepoint y actualiza RUTA_EXCEL_FORMULAS"
        )
    if nombre_hoja not in PERFILES_HOJA:
        raise ValueError(
            f"Hoja '{nombre_hoja}' no tiene perfil definido en PERFILES_HOJA.\n"
            f"Perfiles disponibles: {list(PERFILES_HOJA.keys())}"
        )
    perfil = PERFILES_HOJA[nombre_hoja]

    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)

    if nombre_hoja not in wb.sheetnames:
        raise ValueError(
            f"Hoja '{nombre_hoja}' no existe en el Excel.\n"
            f"Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[nombre_hoja]

    col_id = perfil["col_cod_pieza"]   # "Cod Pieza" en MX/CAM, "Producto GENESIS" en EEUU
    headers = {}
    header_row_idx = None

    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        for col_idx, val in enumerate(row):
            if val and col_id in str(val):
                header_row_idx = row_idx
                for c_i, h_val in enumerate(row):
                    if h_val is not None:
                        headers[str(h_val).strip()] = c_i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError(
            f"No se encontró la columna '{col_id}' en las primeras 20 filas "
            f"de la hoja '{nombre_hoja}'.\n¿Cambió el encabezado del Excel?"
        )

    # ── Validar colores del perfil contra el Excel real ───────────────────
    cols_color_ok      = [c for c in perfil["colores"] if c in headers]
    cols_color_missing = [c for c in perfil["colores"] if c not in headers]

    print(f"  [Excel/{nombre_hoja}] Header en fila {header_row_idx + 1}. "
          f"Colores encontrados: {len(cols_color_ok)}/{len(perfil['colores'])}")

    if cols_color_missing:
        print(f"  [Excel/{nombre_hoja}] Colores NO encontrados (revisar perfil): "
              f"{cols_color_missing}")

    # ── Leer filas de datos ───────────────────────────────────────────────
    resultado = []

    def cel(row_values, col_name):
        """Lee celda por nombre de columna. None si no existe o está fuera de rango."""
        idx = headers.get(col_name)
        if idx is None or idx >= len(row_values):
            return None
        return row_values[idx]

    def tiene_valor(v):
        """True si la celda tiene cualquier valor no vacío (chulo, código SAP, etc.)"""
        return v not in (None, "", 0, False) and str(v).strip() != ""

    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        if all(v is None for v in row):
            continue

        # ── Filtro por fórmula (prioritario) o por pieza ─────────────────────
        # Si se pasa fórmula, filtra por fórmula. Si no, filtra por pieza (para atrás compat).
        formula = str(cel(row, perfil["col_formula"]) or "").strip()
        
        if formula_filtro:
            if formula != formula_filtro:
                continue
        elif cod_pieza_filtro:
            cod_pieza_fila = str(cel(row, perfil["col_cod_pieza"]) or "").strip()
            if cod_pieza_filtro not in cod_pieza_fila and cod_pieza_fila not in cod_pieza_filtro:
                continue

        if not formula:
            continue

        # ── Filtro por activo ─────────────────────────────────────────────
        # Si la hoja tiene columna Activo, filtramos por "activa".
        # Si no tiene (CAM, USSR), todas las filas se consideran activas.
        if perfil["col_activo"] is not None:
            activo = str(cel(row, perfil["col_activo"]) or "").strip().lower()
            # Vacío = activa por defecto. Solo excluir si explícitamente inactiva.
            if activo not in ("activa", ""):
                continue

        # ── Acero ─────────────────────────────────────────────────────────
        # Si la hoja tiene columna de acero, la leemos y normalizamos.
        # Si no tiene (EEUU), usamos el valor por defecto del perfil.
        if perfil["col_acero"] is not None:
            acero_norm = _normalizar_acero(cel(row, perfil["col_acero"]))
        else:
            acero_norm = perfil["acero_default"]

        # ── Colores disponibles ───────────────────────────────────────────
        # Celda con cualquier valor no vacío = chulo = color aplica.
        colores_fila = [
            nombre_color
            for nombre_color in cols_color_ok
            if tiene_valor(cel(row, nombre_color))
        ]

        cod_pieza_fila = str(cel(row, perfil["col_cod_pieza"]) or "").strip() if perfil["col_cod_pieza"] else ""
        
        resultado.append(FilaFormula(
            cod_pieza      = cod_pieza_fila,
            pieza          = str(cel(row, "Pieza") or "").strip(),
            formula        = formula,
            acero_variante = acero_norm,
            colores        = colores_fila,
        ))

    filtro_info = formula_filtro or cod_pieza_filtro or "todas"
    wb.close()
    print(f"  [Excel/{nombre_hoja}] Filas activas para '{filtro_info}': {len(resultado)}")
    return resultado



SQL_PEDIDO_BASE = """
SELECT TOP 1
    Zfer,
    Mercado,
    TipoPeca,
    codPieza,
    Formula,
    Color,
    SteelPlus,
    ZPLA,
    HojaRuta,
    Plano,
    URLPlano,
    ArchivoCorte
FROM VW_AppEnvolvente_LandMacro
WHERE Zfer = ?
"""

SQL_BLOQUEOS_ACTIVOS = """
SELECT formula, acero_variante, color_codigo
FROM   dbo.M5_Bloqueos
WHERE  pedido_origen = ?
  AND  activo = 1
"""




class MotorExplosion:
    """
    Genera la lista completa de TuplaVariante para un ZFER base.

    Necesita dos conexiones separadas porque los datos viven en dos BDs:
      - conn_prod  → agpc-productivity (VW_AppEnvolvente_LandMacro)
      - conn_local → MODULO_5 local (M5_Bloqueos)
    """

    def __init__(
        self,
        conn_prod:  pyodbc.Connection,
        conn_local: pyodbc.Connection,
        ruta_excel: str = RUTA_EXCEL_FORMULAS,
    ):
        self.conn_prod  = conn_prod
        self.conn_local = conn_local
        self.ruta_excel = ruta_excel

    def _leer_pedido_base(self, zfer: str) -> Optional[dict]:
        cursor = self.conn_prod.cursor()
        cursor.execute(SQL_PEDIDO_BASE, (zfer,))
        fila = cursor.fetchone()
        cursor.close()

        if not fila:
            return None

        
        steel_raw = str(fila[6] or "").strip().upper()
        if "SP" in steel_raw or steel_raw in ("SIM", "YES"):
            acero_base = "SP"
        elif "SN" in steel_raw:
            acero_base = "SN"
        else:
            acero_base = "NO"

        return {
            "zfer":          str(fila[0]  or "").strip(),
            "mercado":       str(fila[1]  or "").strip(),
            "tipo_peca":     str(fila[2]  or "").strip(),
            "cod_pieza":     str(fila[3]  or "").strip(),
            "formula":       str(fila[4]  or "").strip(),
            "color":         str(fila[5]  or "").strip(),
            "acero":         acero_base,
            "zpla":          str(fila[7]  or "").strip(),
            "hoja_ruta":     str(fila[8]  or "").strip(),
            "plano":         str(fila[9]  or "").strip(),
            "url_plano":     str(fila[10] or "").strip(),
            "archivo_corte": str(fila[11] or "").strip(),
        }

    def _leer_bloqueos(self, zfer: str) -> set:
        cursor = self.conn_local.cursor()
        cursor.execute(SQL_BLOQUEOS_ACTIVOS, (zfer,))
        filas = cursor.fetchall()
        cursor.close()
        # Cada fila = (formula, acero_variante, color_codigo)
        return {(f[0], f[1], f[2] or "") for f in filas}

    def explotar(self, zfer: str) -> list:
       
        print(f"\n  [Motor] Explosión para ZFER: {zfer}")

        # 1. Pedido base desde BD de producción
        pedido = self._leer_pedido_base(zfer)
        if not pedido:
            raise ValueError(
                f"ZFER '{zfer}' no encontrado en VW_AppEnvolvente_LandMacro.\n"
                f"Verifica que fue procesado correctamente por M3 y M4."
            )

        print(f"  [Motor] Pedido base →  Pieza:{pedido['cod_pieza']}  "
              f"Fórmula:{pedido['formula']}  Color:{pedido['color']}  "
              f"Acero:{pedido['acero']}  Mercado:{pedido['mercado']}")

        # 2. Fórmulas activas desde Excel
        # La hoja del Excel depende del mercado del ZFER:
        #   MÉXICO → "Mexico" | ESTADOS UNIDOS → "EEUU" | resto → "CAM"
        hoja = hoja_para_mercado(pedido["mercado"])
        print(f"  [Motor] Mercado: {pedido['mercado']} → hoja Excel: '{hoja}'")
        filas_formula = leer_excel_formulas(
            ruta_excel       = self.ruta_excel,
            nombre_hoja      = hoja,
            cod_pieza_filtro = pedido["cod_pieza"],
        )
        if not filas_formula:
            raise ValueError(
                f"Sin fórmulas activas en Excel para pieza '{pedido['cod_pieza']}' "
                f"en hoja '{hoja}'.\n"
                f"Verifica el Excel, la hoja o el perfil en PERFILES_HOJA para esta hoja."
            )

        # 3. Bloqueos desde BD local
        bloqueos = self._leer_bloqueos(zfer)
        print(f"  [Motor] Bloqueos activos: {len(bloqueos)}")

        # 4. Generar tuplas
        # Una FilaFormula = una fórmula + un tipo de acero + N colores.
        # Por cada color de esa fila → una TuplaVariante.
        resultado = []
        stats = {"generadas": 0, "base": 0, "bloqueadas": 0, "duplicados": 0}
        seen = set()

        for fila in filas_formula:
            for color in fila.colores:
                stats["generadas"] += 1

                # El acero viene del Excel (fila.acero_variante), NO se itera
                # sobre ["NO","SN","SP"]. El Excel ya define qué acero tiene
                # cada fórmula — multiplicar x3 generaría combinaciones falsas.
                t = TuplaVariante(
                    zfer_origen    = zfer,
                    mercado        = pedido["mercado"],
                    cod_pieza      = pedido["cod_pieza"],
                    tipo_peca      = pedido["tipo_peca"],
                    formula        = fila.formula,
                    color_codigo   = color,
                    acero_variante = fila.acero_variante,
                    zpla           = pedido["zpla"],
                    hoja_ruta      = pedido["hoja_ruta"],
                    plano          = pedido["plano"],
                    url_plano      = pedido["url_plano"],
                    archivo_corte  = pedido["archivo_corte"],
                )

                if t.es_pedido_base(pedido["formula"], pedido["color"], pedido["acero"]):
                    stats["base"] += 1
                    continue

                if (fila.formula, fila.acero_variante, color) in bloqueos:
                    stats["bloqueadas"] += 1
                    continue

                unique_key = (t.formula, t.color_codigo, t.acero_variante)
                if unique_key in seen:
                    stats["duplicados"] += 1
                    continue

                resultado.append(t)
                seen.add(unique_key)

        print(f"  [Motor] Generadas:{stats['generadas']}  "
              f"Excl.base:{stats['base']}  Bloqueadas:{stats['bloqueadas']}  "
              f"Duplicados:{stats['duplicados']}  "
              f"ACTIVAS:{len(resultado)}")

        return resultado




def main():
    print("\n" + "="*60)
    print("  MÓDULO 5 — Motor de explosión v2")
    print("="*60)

    # Acepta ZFER desde variable de entorno (cuando se llama desde MODULO5.py)
    # o usa el valor por defecto para ejecución manual.
    ZFER_PRUEBA = os.environ.get("M5_ZFER_BASE", "700178939")

    try:
        print("\n  Conectando a producción (agpcol.database.windows.net)...")
        conn_prod = pyodbc.connect(get_conn_str(prod=True), autocommit=True)

        print("  Conectando a local (MODULO_5)...")
        conn_local = pyodbc.connect(get_conn_str(prod=False), autocommit=True)

        motor = MotorExplosion(conn_prod, conn_local)
        tuplas = motor.explotar(ZFER_PRUEBA)

        ruta_salida = r"C:\Users\abotero\OneDrive - AGP GROUP\Documentos\MODULO_5\combinaciones5.xlsx"
        
        wb_out = openpyxl.Workbook()
        ws = wb_out.active
        ws.title = "Combinaciones"
        ws.append(["#", "ZFER_Origen", "Mercado", "Cod_Pieza", "Tipo_Pieza", "Formula", "Color", "Acero", "ZPLA", "Hoja_Ruta", "Plano", "URL_Plano", "Archivo_Corte"])
        
        for i, t in enumerate(tuplas, 1):
            ws.append([i, t.zfer_origen, t.mercado, t.cod_pieza, t.tipo_peca, t.formula, t.color_codigo, t.acero_variante, t.zpla, t.hoja_ruta, t.plano, t.url_plano, t.archivo_corte])
        
        wb_out.save(ruta_salida)
        print(f"\n  [Excel] Guardado en: {ruta_salida}")
        
        print(f"\n  Primeras {min(15, len(tuplas))} tuplas:")
        print(f"  {'#':<4} {'Fórmula':<12} {'Acero':<6} {'Color'}")
        print(f"  {'-'*65}")
        for i, t in enumerate(tuplas[:15], 1):
            print(f"  {i:<4} {t.formula:<12} {t.acero_variante:<6} {t.color_codigo}")
        if len(tuplas) > 15:
            print(f"  ... y {len(tuplas)-15} más.")


        conn_prod.close()
        conn_local.close()

    except FileNotFoundError as e:
        print(f"\n  [ERROR] {e}")
        sys.exit(1)
    except pyodbc.Error as e:
        print(f"\n  [ERROR BD] {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"\n  [ERROR] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

