"""
SAP_AUTOMATIZADOR.py
Modulo 5 — Automatizador SAP: Cambio de Color masivo via ZMME0001

Flujo por combinacion (formula × acero × color):
  1. Leer MM02 del ZFER base → FRANJA → determina P_FRANJ (00 / 01)
  2. ZMME0001 → Homologador → Cambio de color → ejecutar
  3. Leer ZFER_NUEVO + ZFOR_NUEVO del grid de resultados
  4. Volver al form → reemplazar Material ZFER con ZFER_NUEVO
  5. Comp BOM vs ZPLA → leer posiciones de la comparacion
  6. Agregar N filas en tabla inferior (una por posicion)
  7. Comp BOM vs ZPLA de nuevo (consolida) → reporte
  8. Ejecutar BOM → reporte
  9. MM02(ZFER_NUEVO) + MM02(ZFOR_NUEVO) → actualizar PARTNUMBER AGP
 10. Registrar resultado en M5_LogEjecucion (BD local)
 11. Generar reporte Excel detallado

PENDIENTE / TODO marcados como "# TODO:"
"""

import win32com.client
import time
import pyodbc
import getpass
import datetime
import uuid
import os
from dataclasses import dataclass, field
from typing import Optional, List
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ── Tiempos de espera (ajustar si SAP es lento en tu red) ────────────────────
T_RAPIDO = 0.4   # entre clicks simples
T_MEDIO  = 1.2   # despues de navegacion / F4
T_LENTO  = 2.5   # despues de ejecutar transacciones pesadas


# ── Config BD local ───────────────────────────────────────────────────────────
DB_LOCAL = {
    "server":   r"localhost\SQLEXPRESS",
    "database": "MODULO_5",
    "driver":   "ODBC Driver 17 for SQL Server",
}

# ── Config BD produccion (misma que COMBINADOR) ───────────────────────────────
DB_PROD = {
    "server":   "agpcol.database.windows.net",
    "database": "agpc-productivity",
    "driver":   "ODBC Driver 17 for SQL Server",
    "user":     "Consulta",
    "password": "@GPgl4$$2021",
}

# ── Rutas ─────────────────────────────────────────────────────────────────────
BASE_DIR = r"C:\Users\abotero\OneDrive - AGP GROUP\Documentos\MODULO_5"


# ── Estructuras de datos ──────────────────────────────────────────────────────

@dataclass
class ClasificacionSAP:
    """Datos de clasificacion de un material en MM02 → tab PIEZA."""
    partnumber:   str  = ""   # PARTNUMBER AGP  ej: "1786_003_L19-31_01_003"
    color:        str  = ""   # COLOR            ej: "G2 GRAY MEDIUM AUTOMOTIVE"
    franja:       str  = ""   # FRANJA           ej: "Sin Franja" o "01"
    tiene_franja: bool = False
    version_agp:  str  = ""   # VERSION AGP      ej: "002"


@dataclass
class ResultadoCombinacion:
    """Resultado completo del procesamiento de una combinacion en SAP."""
    batch_id:       str
    zfer_base:      str
    formula:        str
    acero:          str
    color:          str
    cod_pieza:      str = ""
    tipo_pieza:     str = ""
    # Resultados SAP
    zfer_nuevo:     str = ""
    zfor_nuevo:     str = ""
    posiciones_bom: List[str] = field(default_factory=list)
    # Estado
    estado:         str = "PENDIENTE"   # PENDIENTE | OK | ERROR | SALTADO
    error:          str = ""
    # Timestamps
    fecha_inicio:   Optional[datetime.datetime] = None
    fecha_fin:      Optional[datetime.datetime] = None

    @property
    def duracion_seg(self) -> float:
        if self.fecha_inicio and self.fecha_fin:
            return round((self.fecha_fin - self.fecha_inicio).total_seconds(), 1)
        return 0.0


# ── Automatizador SAP ─────────────────────────────────────────────────────────

class AutomatizadorSAP:
    """
    Controla SAP GUI via win32com (SAP GUI Scripting API).
    Requiere SAP GUI abierto con sesion activa y Scripting habilitado.

    Habilitar scripting:
      SAP GUI → tuerca → Options → Accessibility & Scripting → Enable Scripting ✓
    """

    # IDs de controles SAP extraidos del VBS grabado
    _ID_TCODE_BOX   = "wnd[0]/tbar[0]/okcd"
    _ID_STATUSBAR   = "wnd[0]/sbar"

    # ── ZMME0001 ──────────────────────────────────────────────────────────────────
    _ID_MATER_LOW   = "wnd[0]/usr/ctxtP_MATER-LOW"
    _ID_CTX_CENTER  = "wnd[0]/usr/ctxtP_CENTER"
    _ID_RAD_HOMOLOG = "wnd[0]/usr/radRB5"
    _ID_RAD_COLOR   = "wnd[0]/usr/radRB3_A1"
    _ID_CTX_P_COLOR = "wnd[0]/usr/ctxtP_COLOR"
    _ID_CTX_P_FRANJ = "wnd[0]/usr/ctxtP_FRANJ"
    _ID_CTX_P_ZPLA  = "wnd[0]/usr/ctxtP_ZPLA"
    _ID_ZPLA_SHELL  = "wnd[1]/usr/cntlLO_CONTAINER0500/shellcont/shell"
    _ID_BTN_EXEC    = "wnd[0]/tbar[1]/btn[8]"
    _ID_BTN_BACK    = "wnd[0]/tbar[0]/btn[3]"
    _ID_GRID_RESULT = "wnd[0]/usr/cntlGRID1/shellcont/shell"
    _ID_BTN_COMP    = "wnd[0]/usr/btnBUTTON1"

    # Tabla inferior ZMME0001 (paso 4)
    _TBL_BASE        = (
        "wnd[0]/usr/tabsTABSTRIP_MAX/tabpPUSH1"
        "/ssub%_SUBSCREEN_MAX:ZMME0001:0200"
    )
    _ID_BTN_INSERT    = _TBL_BASE + "/btnT_LISTA_MATERIA_INSERT"
    _ID_TBL_LISTA     = _TBL_BASE + "/tblZMME0001T_LISTA_MATERIA"
    _ID_BTN_COPY_ITEM = _TBL_BASE + "/btnCOPY_ITEM"

    # ── ZPPR0020 ──────────────────────────────────────────────────────────────────
    _ID_ZPPR_USER   = "wnd[0]/usr/txtS_USER-LOW"
    _ID_ZPPR_CENTRO = "wnd[0]/usr/ctxtS_WERKS-LOW"
    _SAP_USER       = "PROGRAING"

    # ── MM02 ──────────────────────────────────────────────────────────────────────
    _ID_MM02_MATNR = "wnd[0]/usr/ctxtRMMG1-MATNR"
    _ID_MM02_TAB03 = "wnd[0]/usr/tabsTABSPR1/tabpSP03"
    _ID_MM02_TAB4  = (
        "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000"
        "/tabsTABSTRIP_CHAR/tabpTAB4"
    )
    _ID_MM02_TABLA = (
        "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000"
        "/tabsTABSTRIP_CHAR/tabpTAB4"
        "/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100"
        "/tblSAPLCTMSCHARS_S"
    )
    _FILA_PARTNUMBER = 0
    _FILA_COLOR      = 1
    _FILA_FRANJA     = 2

    # ── Init / conexion ───────────────────────────────────────────────────────

    def __init__(self):
        self.app:      object = None
        self.conn_sap: object = None
        self.session:  object = None
        self.batch_id: str    = str(uuid.uuid4())
        self.resultados: List[ResultadoCombinacion] = []
        self._usuario  = getpass.getuser()
        self.formula_base:         str        = ""
        self.items_solo_reporte:   list       = []   # items no procesados (cambio formula)
        self._ruta_json:           str        = os.path.join(
            BASE_DIR, f"progreso_{self.batch_id[:8]}.json"
        )

    def conectar(self) -> bool:
        """
        Conecta al SAP GUI ya abierto en el equipo.
        Retorna True si la conexion fue exitosa.
        """
        print("  Conectando a SAP GUI...")
        try:
            sap_gui_auto = win32com.client.GetObject("SAPGUI")
            self.app     = sap_gui_auto.GetScriptingEngine

            if self.app.Children.Count == 0:
                print("  [ERROR] SAP GUI abierto pero sin ninguna conexion activa.")
                print("  → Abre SAP e inicia sesion antes de ejecutar este script.")
                return False

            self.conn_sap = self.app.Children(0)
            self.session  = self.conn_sap.Children(0)
            self.session.findById("wnd[0]").maximize()

            titulo = self.session.findById("wnd[0]").text
            print(f"  [OK] SAP GUI conectado — ventana: {titulo}")
            return True

        except Exception as e:
            print(f"  [ERROR] No se pudo conectar a SAP GUI: {e}")
            print("  Verifica:")
            print("    1. SAP GUI este abierto y con sesion activa")
            print("    2. Scripting habilitado: tuerca → Options → Accessibility & Scripting")
            return False

    # ── Helpers internos ──────────────────────────────────────────────────────

    def _esperar(self, segundos: float = T_RAPIDO):
        time.sleep(segundos)

    def _navegar(self, tcode: str):
        """Navega a una transaccion usando /N (sin cerrar sesion)."""
        self.session.findById(self._ID_TCODE_BOX).text = f"/N{tcode}"
        self.session.findById("wnd[0]").sendVKey(0)
        self._esperar(T_MEDIO)

    def _estado_sap(self) -> str:
        """Lee el mensaje de la barra de estado inferior de SAP."""
        try:
            return self.session.findById(self._ID_STATUSBAR).text.strip()
        except Exception:
            return ""

    def _cerrar_popup_si_existe(self, boton: str = "wnd[1]/tbar[0]/btn[12]"):
        """Intenta cerrar wnd[1] si existe (Cancel / ESC)."""
        try:
            self.session.findById(boton)
            self.session.findById(boton).press()
            self._esperar(T_RAPIDO)
        except Exception:
            pass

    def _aceptar_dialogo(self):
        """Acepta wnd[1] con Enter (btn[0]) si existe."""
        try:
            self.session.findById("wnd[1]").sendVKey(0)
            self._esperar(T_RAPIDO)
        except Exception:
            pass

    # ── MM02 — Leer clasificacion ─────────────────────────────────────────────

    def leer_clasificacion_zfer(self, zfer: str) -> "ClasificacionSAP":
        """
        MM02 del ZFER → tab PIEZA → lee PARTNUMBER, COLOR, FRANJA.
        FRANJA se devuelve como código SAP directo: "00","01","02","03","NA" o "".
        """
        self._navegar("MM02")
        self.session.findById(self._ID_MM02_MATNR).text = zfer
        self.session.findById("wnd[0]").sendVKey(0)
        self._esperar(T_MEDIO)
        self._aceptar_dialogo()
        self._aceptar_dialogo()

        self.session.findById(self._ID_MM02_TAB03).select()
        self._esperar(T_RAPIDO)
        self.session.findById(self._ID_MM02_TAB4).select()
        self._esperar(T_RAPIDO)

        def _leer(fila: int) -> str:
            try:
                return self.session.findById(
                    f"{self._ID_MM02_TABLA}/ctxtRCTMS-MWERT[1,{fila}]"
                ).text.strip()
            except Exception:
                return ""

        partnumber = _leer(self._FILA_PARTNUMBER)
        color      = _leer(self._FILA_COLOR)
        franja_txt = _leer(self._FILA_FRANJA)

        # Determinar si tiene franja por el código
        # "00" = Sin Franja, "" = sin entrada → no tiene franja
        # "01"(Azul),"02"(Verde),"03"(Gris),"NA"(No Aplica) → tiene franja
        tiene_franja = franja_txt not in ("00", "", "SIN FRANJA", "SIN ENTRADA")

        # El P_FRANJ que va a ZMME0001 es el código literal ("00","01","02","03","NA")
        # Si está vacío, usar "00" por defecto
        p_franj_code = franja_txt if franja_txt else "00"

        self.session.findById(self._ID_MM02_TAB03).select()  # salir limpio
        print(f"    MM02 {zfer}: partnumber={partnumber} | franja={franja_txt} | p_franj={p_franj_code}")
        return ClasificacionSAP(
            partnumber   = partnumber,
            color        = color,
            franja       = franja_txt,
            tiene_franja = tiene_franja,
            version_agp  = "",
        )

    # ── ZMME0001 — Paso 2: ejecutar y leer ZFER nuevo ────────────────────────

    def zmme0001_ejecutar(
        self,
        zfer_base: str,
        p_color:   str,
        p_franj:   str,
    ) -> tuple:
        """
        Paso 2 completo: ZMME0001 → Homologar → Cambio de color → Ejecutar.
        Retorna (zfer_nuevo, zfor_nuevo, zpla_seleccionado).
        El campo material se llena directamente (ctxtP_MATER-LOW), sin popup.
        """
        # Cerrar cualquier diálogo pendiente antes de navegar
        self._cerrar_dialogs_abiertos()
        self._navegar("ZMME0001")
        self.session.findById("wnd[0]").maximize()
        self._esperar(T_RAPIDO)

        # 1. Homologar
        self.session.findById(self._ID_RAD_HOMOLOG).setFocus()
        self.session.findById(self._ID_RAD_HOMOLOG).select()
        self._esperar(T_RAPIDO)

        # 2. Material ZFER (directo, sin popup)
        self.session.findById(self._ID_MATER_LOW).text = zfer_base
        self._esperar(T_RAPIDO)

        # 3. Centro
        self.session.findById(self._ID_CTX_CENTER).text = "CO01"

        # 4. Cambio de color
        self.session.findById(self._ID_RAD_COLOR).setFocus()
        self.session.findById(self._ID_RAD_COLOR).select()
        self._esperar(T_RAPIDO)

        # 5. Color y Franja
        self.session.findById(self._ID_CTX_P_COLOR).text = p_color
        self.session.findById(self._ID_CTX_P_FRANJ).text = p_franj

        # 6. ZPLA Referencia — F4 → seleccionar fila 0
        self.session.findById(self._ID_CTX_P_ZPLA).setFocus()
        self.session.findById(self._ID_CTX_P_ZPLA).caretPosition = 0
        self.session.findById("wnd[0]").sendVKey(4)  # F4
        self._esperar(T_MEDIO)
        self.session.findById(self._ID_ZPLA_SHELL).selectedRows = "0"
        self.session.findById(self._ID_ZPLA_SHELL).doubleClickCurrentCell()
        self._esperar(T_RAPIDO)

        # Leer el ZPLA que se seleccionó
        try:
            zpla_seleccionado = self.session.findById(self._ID_CTX_P_ZPLA).text.strip()
        except Exception:
            zpla_seleccionado = ""

        # 7. Ejecutar (F8)
        self.session.findById(self._ID_BTN_EXEC).press()
        self._esperar(T_LENTO)

        # 8. Leer resultado del grid
        msg_sap = self._estado_sap()
        try:
            grid       = self.session.findById(self._ID_GRID_RESULT)
            zfer_nuevo = grid.GetCellValue(0, "ZFER").strip()
            zfor_nuevo = grid.GetCellValue(0, "ZFOR").strip()
        except Exception as e:
            raise RuntimeError(
                f"No se pudo leer resultado del grid ZMME0001. "
                f"SAP: '{msg_sap}'. Detalle: {e}"
            )

        if not zfer_nuevo:
            raise RuntimeError(f"ZFER_NUEVO vacío tras ejecutar ZMME0001. SAP: '{msg_sap}'")

        print(f"    ZMME0001 OK: ZFER_NUEVO={zfer_nuevo} | ZFOR={zfor_nuevo} | ZPLA={zpla_seleccionado}")

        # Volver a pantalla de selección (F3) para que los campos estén disponibles
        # al regresar de ZPPR0020 (el resultado queda en la pantalla del grid)
        self.session.findById("wnd[0]").sendVKey(3)  # F3 = Retroceder
        self._esperar(T_RAPIDO)

        return zfer_nuevo, zfor_nuevo, zpla_seleccionado

    # ── ZMME0001 — Paso 4: Comparar BOM + agregar filas + copy ───────────────

    def zmme0001_leer_posiciones_popup(self) -> list:
        """
        Presiona Comparar BOM (btnBUTTON1), lee las posiciones del popup wnd[1],
        cierra el popup y retorna lista de strings de posición (ej ["0458"]).
        """
        self.session.findById(self._ID_BTN_COMP).press()
        self._esperar(T_MEDIO)

        posiciones = []

        # Intentar leer el popup como GuiTableControl clásico
        try:
            tabla   = self.session.findById("wnd[1]/usr/tblZMME0001T_COMP")
            n_filas = tabla.RowCount
            for i in range(n_filas):
                try:
                    # Columna 0 normalmente tiene POSNR
                    pos = tabla.GetCell(i, 0).text.strip()
                    if pos:
                        posiciones.append(pos)
                except Exception:
                    pass
        except Exception:
            pass

        # Fallback: GuiGridView en el popup
        if not posiciones:
            try:
                grid    = self.session.findById("wnd[1]/usr/cntlGRID1/shellcont/shell")
                n_filas = grid.RowCount
                for i in range(n_filas):
                    try:
                        pos = grid.GetCellValue(i, "POSNR").strip()
                        if pos:
                            posiciones.append(pos)
                    except Exception:
                        pass
            except Exception:
                pass

        # Cerrar popup
        try:
            self.session.findById("wnd[1]").close()
        except Exception:
            try:
                self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
            except Exception:
                pass
        self._esperar(T_RAPIDO)

        # Deduplicar conservando orden
        vistos = set()
        return [p for p in posiciones if not (p in vistos or vistos.add(p))]

    def zmme0001_agregar_filas_bom(self, posiciones: list, zpla: str):
        """
        Para cada posicion de la lista:
          1. Presiona Insert
          2. Llena POSNR
          3. Consulta BD para CLASE (= Clave Destino)
          4. Llena CLASE_DESTINO
        """
        for idx, pos in enumerate(posiciones):
            # Insertar fila
            self.session.findById(self._ID_BTN_INSERT).press()
            self._esperar(T_RAPIDO)

            # POSNR
            self.session.findById(
                f"{self._ID_TBL_LISTA}/txtWA_LISTA-POSNR[0,{idx}]"
            ).text = pos
            self._esperar(T_RAPIDO)

            # Consultar clase destino
            clase = self._consultar_clase_destino(zpla, pos)
            if not clase:
                print(f"    [WARN] Sin clase destino en BD para ZPLA={zpla} POS={pos}")

            # CLASE_DESTINO
            if clase:
                self.session.findById(
                    f"{self._ID_TBL_LISTA}/ctxtWA_LISTA-CLASE_DESTINO[3,{idx}]"
                ).text = clase
                self._esperar(T_RAPIDO)

            print(f"    Fila {idx}: POS={pos} CLASE={clase or '(sin clase)'}")

    def zmme0001_segunda_comparar_y_copy(self) -> bool:
        """
        Segunda pasada de Comparar BOM y luego btnCOPY_ITEM.
        Retorna True si el popup muestra resultado positivo (no error).
        """
        self.session.findById(self._ID_BTN_COMP).press()
        self._esperar(T_MEDIO)

        # Leer status del popup antes de cerrar
        ok = True
        try:
            # Verificar si hay errores en el popup (grid de errores)
            try:
                grid_err = self.session.findById(
                    "wnd[1]/usr/cntlGRID1/shellcont/shell"
                )
                n_err = grid_err.RowCount
                if n_err > 0:
                    # Leer primera fila para ver si es error
                    try:
                        tipo = grid_err.GetCellValue(0, "TY").strip()
                        if tipo.upper() == "E":
                            msg  = grid_err.GetCellValue(0, "VARIABLE_MENSAJE").strip()
                            ok   = False
                            print(f"    [ERROR] Segunda comparacion: {msg}")
                    except Exception:
                        pass
            except Exception:
                pass

            # Cerrar popup
            try:
                self.session.findById("wnd[1]").close()
            except Exception:
                try:
                    self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
                except Exception:
                    pass
            self._esperar(T_RAPIDO)
        except Exception:
            pass

        if ok:
            # COPY_ITEM (equivale a confirmar/ejecutar BOM)
            try:
                self.session.findById(self._ID_BTN_COPY_ITEM).press()
                self._esperar(T_LENTO)
                msg_final = self._estado_sap()
                print(f"    COPY_ITEM: {msg_final}")
                # Si el statusbar contiene "error" → reportar como fallo
                if msg_final and "error" in msg_final.lower():
                    ok = False
            except Exception as e:
                print(f"    [WARN] COPY_ITEM: {e}")

        return ok

    # ── MM02 — Actualizar PARTNUMBER AGP ─────────────────────────────────────

    def _cerrar_dialogs_abiertos(self):
        """Cierra cualquier wnd[1]/wnd[2] abierto antes de navegar."""
        for wnd in ("wnd[2]", "wnd[1]"):
            try:
                self.session.findById(wnd)
                # Intenta cancelar (F12) primero, luego Enter
                try:
                    self.session.findById(wnd).sendVKey(12)
                except Exception:
                    try:
                        self.session.findById(wnd).sendVKey(0)
                    except Exception:
                        pass
                self._esperar(T_RAPIDO)
            except Exception:
                pass

    def mm02_actualizar_partnumber(self, material: str, nuevo_partnumber: str):
        """
        MM02 del material → Clasificación → TAB4 (PIEZA) → actualiza PARTNUMBER (fila 0) → guarda.
        """
        self._cerrar_dialogs_abiertos()
        self._navegar("MM02")
        self.session.findById(self._ID_MM02_MATNR).text = material
        self.session.findById("wnd[0]").sendVKey(0)
        self._esperar(T_MEDIO)

        # Aceptar diálogos de vistas (pueden aparecer o no según configuración MM02)
        for _ in range(2):
            try:
                self.session.findById("wnd[1]").sendVKey(0)
                self._esperar(T_RAPIDO)
            except Exception:
                break

        self.session.findById(self._ID_MM02_TAB03).select()
        self._esperar(T_RAPIDO)
        self.session.findById(self._ID_MM02_TAB4).select()
        self._esperar(T_RAPIDO)

        # Actualizar PARTNUMBER (fila 0)
        self.session.findById(
            f"{self._ID_MM02_TABLA}/ctxtRCTMS-MWERT[1,{self._FILA_PARTNUMBER}]"
        ).text = nuevo_partnumber
        self._esperar(T_RAPIDO)

        # Guardar (btn[0] dos veces como en VBS)
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()
        self._esperar(T_RAPIDO)
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()
        self._esperar(T_MEDIO)

        # Confirmar si aparece diálogo de guardar
        try:
            self.session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
            self._esperar(T_RAPIDO)
        except Exception:
            pass

        # Volver
        try:
            self.session.findById(self._ID_MM02_TAB03).select()  # salir limpio
        except Exception:
            pass

        print(f"      MM02 {material} PARTNUMBER → {nuevo_partnumber}")

     
    def leer_formula_base_bd(self, zfer: str) -> str:
        """
        Consulta la formula del ZFER base en la BD de produccion.
        Busca en dos tablas en orden:
          1. ZFER_Characteristics_Genesis  (SpecID = ZFER, columna FormulaCode)
          2. TCAL_CALENDARIO_COLOMBIA_DIRECT (ZFER = ZFER, columna Formula)
        Retorna "" si no conecta o no encuentra el ZFER en ninguna tabla.
        """
        conn_str = (
            f"DRIVER={{{DB_PROD['driver']}}};"
            f"SERVER={DB_PROD['server']};"
            f"DATABASE={DB_PROD['database']};"
            f"UID={DB_PROD['user']};"
            f"PWD={DB_PROD['password']};"
        )
        try:
            conn   = pyodbc.connect(conn_str, autocommit=True, timeout=10)
            cursor = conn.cursor()

            # 1. ZFER_Characteristics_Genesis
            cursor.execute(
                "SELECT TOP 1 FormulaCode FROM dbo.ZFER_Characteristics_Genesis "
                "WHERE SpecID = ?",
                (zfer,)
            )
            fila = cursor.fetchone()
            if fila and fila[0]:
                conn.close()
                print(f"    Formula encontrada en Genesis: {fila[0].strip()}")
                return str(fila[0]).strip()

            # 2. TCAL_CALENDARIO_COLOMBIA_DIRECT
            cursor.execute(
                "SELECT TOP 1 Formula FROM dbo.TCAL_CALENDARIO_COLOMBIA_DIRECT "
                "WHERE ZFER = ?",
                (zfer,)
            )
            fila = cursor.fetchone()
            conn.close()
            if fila and fila[0]:
                print(f"    Formula encontrada en Calendario: {fila[0].strip()}")
                return str(fila[0]).strip()

        except Exception as e:
            print(f"    [WARN] leer_formula_base_bd: {e}")
        return ""

    # ── ZPPR0020 — Paso 3: esperar fases en sesion nueva ─────────────────────

    def zppr0020_esperar_fases(
        self,
        zfer_nuevo: str,
        intervalo_seg: int = 30,
        max_espera_seg: int = 600,
    ) -> dict:
        """
        Paso 3: Abre una NUEVA sesion SAP para ZPPR0020 (deja ZMME0001 intacta
        en la sesion principal). Hace polling hasta que:
          - Alguna fase tenga "E" → error, cierra sesion auxiliar, retorna ok=False
          - Mas de 7 fases tienen "S" → OK, cierra sesion auxiliar, retorna ok=True
          - Se agote el tiempo → timeout, cierra sesion auxiliar, retorna ok=False
        """
        # Abrir nueva sesion (ZMME0001 se queda abierta en self.session)
        print("     Abriendo sesion auxiliar para ZPPR0020...")
        self.session.createSession()
        self._esperar(T_LENTO)

        # Obtener la nueva sesion (ultima creada)
        idx_nueva = self.conn_sap.Children.Count - 1
        ses2 = self.conn_sap.Children(idx_nueva)
        ses2.findById("wnd[0]").maximize()

        try:
            # Navegar a ZPPR0020 en la sesion auxiliar
            ses2.findById("wnd[0]/tbar[0]/okcd").text = "ZPPR0020"
            ses2.findById("wnd[0]").sendVKey(0)
            self._esperar(T_MEDIO)

            ses2.findById(self._ID_ZPPR_USER).text  = self._SAP_USER
            ses2.findById(self._ID_ZPPR_CENTRO).text = "CO01"
            ses2.findById(self._ID_BTN_EXEC).press()
            self._esperar(T_LENTO)

            iteraciones = max(1, max_espera_seg // intervalo_seg)
            sin_datos_contador = 0

            for intento in range(iteraciones):
                resultado = self._leer_zppr0020_grid(zfer_nuevo, ses2)

                if resultado.get("encontrado"):
                    fases = resultado.get("fases", {})
                    zpla  = resultado.get("zpla", "")

                    # Si hay "E" en cualquier fase → error inmediato
                    for nombre_fase, valor in fases.items():
                        if str(valor).strip().upper() == "E":
                            return {
                                "ok":         False,
                                "zpla":       zpla,
                                "fase_error": nombre_fase,
                                "detalle":    f"{nombre_fase} tiene estado 'E' (Error) en ZPPR0020.",
                                "fases":      fases,
                            }

                    # Contar cuantas fases tienen "S"
                    fases_con_s = [v for v in fases.values() if str(v).strip().upper() == "S"]
                    n_s = len(fases_con_s)

                    if n_s > 7:
                        print(f"    ZPPR0020 OK: {n_s} fases con S | ZPLA={zpla}")
                        return {
                            "ok":         True,
                            "zpla":       zpla,
                            "fase_error": "",
                            "detalle":    f"{n_s} fases completadas con S",
                            "fases":      fases,
                        }

                    print(f"    ZPPR0020: {n_s} fases S (esperando >7)... intento {intento+1}/{iteraciones}")

                else:
                    sin_datos_contador += 1
                    print(f"    ZPPR0020: sin datos intento {intento+1}/{iteraciones}")
                    if sin_datos_contador >= 10:
                        return {
                            "ok":         False,
                            "zpla":       "",
                            "fase_error": "SIN_DATOS",
                            "detalle":    f"ZPPR0020 no mostró datos del ZFER {zfer_nuevo} en {sin_datos_contador * intervalo_seg}s.",
                            "fases":      {},
                        }

                if intento < iteraciones - 1:
                    time.sleep(intervalo_seg)
                    ses2.findById("wnd[0]").sendVKey(9)   # F9 refresh
                    self._esperar(T_MEDIO)

            return {
                "ok":         False,
                "zpla":       "",
                "fase_error": "TIMEOUT",
                "detalle":    f"ZPPR0020 no completó en {max_espera_seg//60} minutos.",
                "fases":      {},
            }

        finally:
            # Cerrar sesion auxiliar con /i (delete session) — más limpio que /nex
            try:
                ses2.findById("wnd[0]/tbar[0]/okcd").text = "/i"
                ses2.findById("wnd[0]").sendVKey(0)
            except Exception:
                try:
                    ses2.findById("wnd[0]").close()
                except Exception:
                    pass
            self._esperar(T_MEDIO)

            # Re-adquirir self.session: al cerrar ses2 el COM reference puede quedar
            # inválido si SAP GUI re-indexó las sesiones
            try:
                self.session = self.conn_sap.Children(0)
                self.session.findById("wnd[0]").maximize()
            except Exception as e_reacq:
                print(f"     [WARN] Re-adquirir sesion principal: {e_reacq}")

            print("     Sesion auxiliar ZPPR0020 cerrada.")

    def _buscar_grid_recursivo(self, contenedor, profundidad: int = 0):
        """
        Busca recursivamente el primer GuiGridView o GuiTableControl.
        Recorre TODOS los hijos (ContainerType o no) hasta profundidad 8.
        """
        if profundidad > 8:
            return None
        try:
            n = contenedor.Children.Count
        except Exception:
            return None
        for i in range(n):
            try:
                hijo = contenedor.Children(i)
            except Exception:
                continue
            try:
                tipo = hijo.Type
            except Exception:
                tipo = ""
            if tipo in ("GuiGridView", "GuiTableControl"):
                return hijo
            # Recurrir en cualquier hijo que tenga Children (sea o no ContainerType)
            encontrado = self._buscar_grid_recursivo(hijo, profundidad + 1)
            if encontrado is not None:
                return encontrado
        return None

    def _debug_imprimir_elementos_sap(self, titulo: str = "wnd[0]"):
        """
        Imprime TODOS los elementos de la pantalla para debug.
        Útil cuando no se encuentra el grid para identificar los IDs reales.
        """
        print(f"\n    === DEBUG: Elementos en {titulo} ===")
        try:
            wnd = self.session.findById("wnd[0]")
            self._imprimir_hijos(wnd, 0, max_nivel=3)
            print(f"    === FIN DEBUG ===\n")
        except Exception as e:
            print(f"    [ERROR DEBUG] {e}")

    def _imprimir_hijos(self, obj, nivel: int, max_nivel: int = 3):
        """Imprime los hijos de un objeto recursivamente."""
        if nivel > max_nivel:
            return
        try:
            n = obj.Children.Count
        except Exception:
            return
        pre = "  " * (nivel + 1)
        for i in range(n):
            try:
                hijo = obj.Children(i)
                tipo = getattr(hijo, "Type", "?")
                id_obj = getattr(hijo, "Id", "?")
                # Solo mostrar elementos interesantes (grids, containers, etc.)
                if tipo in ("GuiGridView", "GuiTableControl", "GuiSplitterShell", 
                            "GuiContainerShell", "GuiShell", "GuiToolbar") or nivel == 0:
                    print(f"{pre}[{i}] {tipo}: {id_obj}")
                # Continuar explorando
                self._imprimir_hijos(hijo, nivel + 1, max_nivel)
            except Exception:
                pass

    def _leer_zppr0020_grid(self, zfer_nuevo: str, ses=None) -> dict:
        """
        Lee el grid ALV de ZPPR0020 y busca la fila con ZFER = zfer_nuevo.
        Usa ses si se provee (sesion auxiliar), si no usa self.session.
        Columnas confirmadas: MAT_ZFER, MAT_ZPLA, PHASE1..PHASEn
        """
        if ses is None:
            ses = self.session
        resultado = {"encontrado": False, "zpla": "", "fases": {}}

        _RUTAS = [
            "wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell",
            "wnd[0]/usr/cntlGRID1/shellcont/shell",
            "wnd[0]/usr/cntlGRID/shellcont/shell",
            "wnd[0]/shellcont/shell",
        ]
        grid = None
        for ruta in _RUTAS:
            try:
                obj = ses.findById(ruta)
                rc  = obj.RowCount
                if rc is not None:
                    grid = obj
                    break
            except Exception:
                pass

        if grid is None:
            try:
                wnd  = ses.findById("wnd[0]")
                grid = self._buscar_grid_recursivo(wnd)
            except Exception:
                pass

        if grid is None:
            print("    [WARN] ZPPR0020: grid no encontrado.")
            return resultado

        try:
            n_filas = grid.RowCount
            if not n_filas:
                return resultado

            try:
                cols = list(grid.ColumnOrder)
            except Exception:
                cols = []

            for i in range(n_filas):
                # Buscar ZFER en fila
                zfer_fila = ""
                for col in cols:
                    if "MAT_ZFER" in col.upper() or col.upper() in ("ZFER", "MATNR_ZFER"):
                        try:
                            zfer_fila = str(grid.GetCellValue(i, col) or "").strip()
                            if zfer_fila:
                                break
                        except Exception:
                            pass
                if zfer_fila != zfer_nuevo:
                    continue

                resultado["encontrado"] = True

                # Leer ZPLA
                for col in cols:
                    if "MAT_ZPLA" in col.upper() or col.upper() in ("ZPLA", "MATNR_ZPLA"):
                        try:
                            v = str(grid.GetCellValue(i, col) or "").strip()
                            if v:
                                resultado["zpla"] = v
                                break
                        except Exception:
                            pass

                # Leer fases (columnas PHASE1, PHASE2, ...)
                fases = {}
                n_fase = 1
                for col in cols:
                    if col.upper().startswith("PHASE"):
                        try:
                            v = str(grid.GetCellValue(i, col) or "").strip()
                            fases[f"Fase {n_fase}"] = v
                            n_fase += 1
                        except Exception:
                            pass
                resultado["fases"] = fases

                n_s = sum(1 for v in fases.values() if v.upper() == "S")
                n_e = sum(1 for v in fases.values() if v.upper() == "E")
                print(f"    ZPPR0020 fila {i}: ZPLA={resultado['zpla']} | S={n_s} | E={n_e}")
                break

        except Exception as e:
            print(f"    [WARN] _leer_zppr0020_grid: {e}")

        return resultado

    def _consultar_clase_destino(self, zpla: str, posicion: str) -> str:
        """
        Consulta ODATA_ZPLA_BOM en BD produccion para obtener la CLASE (= Clave Destino).
        posicion puede venir como "0458" o "458" — normaliza para comparar.
        """
        try:
            conn_str = (
                f"DRIVER={{{DB_PROD['driver']}}};"
                f"SERVER={DB_PROD['server']};"
                f"DATABASE={DB_PROD['database']};"
                f"UID={DB_PROD['user']};"
                f"PWD={DB_PROD['password']};"
            )
            conn   = pyodbc.connect(conn_str, autocommit=True, timeout=10)
            cursor = conn.cursor()
            # Normalizar posicion: quitar ceros a la izquierda para comparar como entero
            pos_int = int(posicion.lstrip("0") or "0")
            cursor.execute(
                """
                SELECT TOP 1 CLASE
                FROM   dbo.ODATA_ZPLA_BOM
                WHERE  MATERIAL = ?
                  AND  CAST(POSICION AS INT) = ?
                ORDER BY POSICION ASC
                """,
                (zpla, pos_int),
            )
            fila = cursor.fetchone()
            conn.close()
            if fila and fila[0]:
                return str(fila[0]).strip()
        except Exception as e:
            print(f"    [WARN] _consultar_clase_destino ZPLA={zpla} POS={posicion}: {e}")
        return ""

    # ── Procesamiento de una combinacion ──────────────────────────────────────

    def _extraer_formula_de_partnumber(self, partnumber: str) -> str:
        """
        Extrae el código de fórmula del PARTNUMBER AGP.
        Patrón: {cod}_{seq}_{formula}_{color}_{ver}
        Ej: "1407_000_L40-2_01_002"  →  "L40-2"
        """
        partes = partnumber.split("_")
        return partes[2] if len(partes) >= 3 else ""

    def _guardar_progreso_json(self):
        """
        Persiste el estado actual del batch en un JSON.
        Se llama después de cada item procesado (checkpoint).
        """
        import json

        ok_n   = sum(1 for r in self.resultados if r.estado == "OK")
        err_n  = sum(1 for r in self.resultados if r.estado == "ERROR")

        data = {
            "batch_id":    self.batch_id,
            "fecha_inicio": self.resultados[0].fecha_inicio.isoformat() if self.resultados else None,
            "fecha_fin":    None,
            "zfer_base":   self.resultados[0].zfer_base if self.resultados else "",
            "formula_base": self.formula_base,
            "operador":    self._usuario,
            "totales": {
                "a_procesar_sap": len(self.resultados) + len([r for r in self.resultados if r.estado == "PENDIENTE"]),
                "procesados":  len(self.resultados),
                "exitosos":    ok_n,
                "errores":     err_n,
                "solo_reporte": len(self.items_solo_reporte),
            },
            "items_sap": [
                {
                    "idx":           i + 1,
                    "formula":       r.formula,
                    "acero":         r.acero,
                    "color":         r.color,
                    "cod_pieza":     r.cod_pieza,
                    "tipo_pieza":    r.tipo_pieza,
                    "zfer_base":     r.zfer_base,
                    "zfer_nuevo":    r.zfer_nuevo,
                    "zfor_nuevo":    r.zfor_nuevo,
                    "posiciones_bom": r.posiciones_bom,
                    "estado":        r.estado,
                    "error":         r.error,
                    "fecha_inicio":  r.fecha_inicio.isoformat() if r.fecha_inicio else None,
                    "fecha_fin":     r.fecha_fin.isoformat()    if r.fecha_fin    else None,
                    "duracion_seg":  r.duracion_seg,
                }
                for i, r in enumerate(self.resultados)
            ],
            "items_solo_reporte": self.items_solo_reporte,
        }

        try:
            with open(self._ruta_json, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"    [WARN] No se pudo guardar JSON progreso: {e}")

    def _extraer_numero_color(self, color: str) -> str:
        """
        Extrae el código numérico del color para el campo P_COLOR de ZMME0001.
        Ej: "19-Gray Light Automotive" → "19"
            "21-Gray Dark Automotive"  → "21"
            "G2 Gray Medium"           → "" (sin número — revisar manualmente)
        """
        parte = color.strip().split("-")[0].split(" ")[0]
        return parte if parte.isdigit() else ""

    def _construir_nuevo_partnumber(
        self,
        partnumber_base: str,
        p_color: str,
    ) -> str:
        """
        Reemplaza el segmento de color (índice 3) del PARTNUMBER AGP.

        Patrón: {codigo}_{seq}_{formula_code}_{color_num}_{version}
        Ejemplo: "1407_000_L40-2_01_002"  con p_color="19"
              →  "1407_000_L40-2_19_002"
        """
        if not partnumber_base or not p_color:
            return partnumber_base
        partes = partnumber_base.split("_")
        if len(partes) >= 4:
            partes[3] = p_color
            return "_".join(partes)
        return partnumber_base

    def procesar_combinacion(
        self,
        zfer_base:      str,
        formula:        str,
        acero:          str,
        color:          str,
        cod_pieza:      str,
        tipo_pieza:     str,
        p_color:        str,
        p_franj:        str,
    ) -> "ResultadoCombinacion":
        """
        Procesa una combinación completa en SAP siguiendo los 5 pasos:
          1. (ya hecho antes) Leer FRANJA → p_franj recibido como parámetro
          2. ZMME0001 → ejecutar → obtener ZFER_NUEVO + ZPLA
          3. ZPPR0020 → esperar fases (polling 30s, max 10 min)
          4. Comparar BOM → leer posiciones → llenar filas con CLASE_DESTINO → Comparar 2 → COPY_ITEM
          5. MM02 → actualizar PARTNUMBER del ZFER_NUEVO (y ZFOR_NUEVO si existe)
        """
        res = ResultadoCombinacion(
            batch_id     = self.batch_id,
            zfer_base    = zfer_base,
            formula      = formula,
            acero        = acero,
            color        = color,
            cod_pieza    = cod_pieza,
            tipo_pieza   = tipo_pieza,
            fecha_inicio = datetime.datetime.now(),
        )

        try:
            print(f"  ▶  {formula} / {acero} / {color[:40]}")
            print(f"     p_color={p_color}  p_franj={p_franj}")

            # ── PASO 2: ZMME0001 → ejecutar ──────────────────────────────────────
            zfer_nuevo, zfor_nuevo, zpla = self.zmme0001_ejecutar(
                zfer_base, p_color, p_franj
            )
            res.zfer_nuevo = zfer_nuevo
            res.zfor_nuevo = zfor_nuevo

            # ── PASO 3: ZPPR0020 → esperar fases ─────────────────────────────────
            print(f"     Esperando ZPPR0020 para ZFER_NUEVO={zfer_nuevo}...")
            fase_result = self.zppr0020_esperar_fases(zfer_nuevo)

            if not fase_result["ok"]:
                raise RuntimeError(
                    f"ZPPR0020 falló — {fase_result['fase_error']}: {fase_result['detalle']}"
                )

            # Si ZPLA no vino de zmme0001, usar el de ZPPR0020
            if not zpla and fase_result.get("zpla"):
                zpla = fase_result["zpla"]
            print(f"     ZPPR0020 OK. ZPLA={zpla}")

            # ── PASO 4: ZMME0001 en pantalla de seleccion (volvimos con F3) ────────
            # Re-establecer campos y cambiar material a ZFER_NUEVO para Comparar BOM

            try:
                self.session.findById(self._ID_RAD_HOMOLOG).setFocus()
                self.session.findById(self._ID_RAD_HOMOLOG).select()
                self._esperar(T_RAPIDO)
                self.session.findById(self._ID_CTX_CENTER).text = "CO01"
                self.session.findById(self._ID_RAD_COLOR).setFocus()
                self.session.findById(self._ID_RAD_COLOR).select()
                self._esperar(T_RAPIDO)
                self.session.findById(self._ID_CTX_P_COLOR).text = p_color
                self.session.findById(self._ID_CTX_P_FRANJ).text = p_franj
                # ZPLA: re-seleccionar si está vacío
                try:
                    zpla_actual = self.session.findById(self._ID_CTX_P_ZPLA).text.strip()
                except Exception:
                    zpla_actual = ""
                if not zpla_actual:
                    self.session.findById(self._ID_CTX_P_ZPLA).setFocus()
                    self.session.findById("wnd[0]").sendVKey(4)   # F4
                    self._esperar(T_MEDIO)
                    self.session.findById(self._ID_ZPLA_SHELL).selectedRows = "0"
                    self.session.findById(self._ID_ZPLA_SHELL).doubleClickCurrentCell()
                    self._esperar(T_RAPIDO)
            except Exception as e_p4:
                print(f"     [WARN] Re-establecer campos paso 4: {e_p4}")

            # Cambiar material por ZFER_NUEVO
            self.session.findById(self._ID_MATER_LOW).text = zfer_nuevo
            self.session.findById(self._ID_MATER_LOW).caretPosition = len(zfer_nuevo)
            self._esperar(T_RAPIDO)

            # Primera Comparar → leer posiciones del popup
            posiciones = self.zmme0001_leer_posiciones_popup()
            res.posiciones_bom = posiciones
            print(f"     Posiciones BOM ({len(posiciones)}): {posiciones}")

            # Agregar filas con POSNR y CLASE_DESTINO de BD
            if posiciones and zpla:
                self.zmme0001_agregar_filas_bom(posiciones, zpla)
            elif posiciones and not zpla:
                print("     [WARN] Sin ZPLA para consultar clase destino")

            # Segunda Comparar + COPY_ITEM
            ok_bom = self.zmme0001_segunda_comparar_y_copy()
            if not ok_bom:
                raise RuntimeError("Segunda Comparar BOM devolvió error — revisar Clave Destino")

            # ── PASO 5: MM02 → actualizar PARTNUMBER ─────────────────────────────
            try:
                clasif_base = self.leer_clasificacion_zfer(zfer_base)
                nuevo_pn    = self._construir_nuevo_partnumber(
                    clasif_base.partnumber, p_color
                )
                if nuevo_pn and nuevo_pn != clasif_base.partnumber:
                    self.mm02_actualizar_partnumber(zfer_nuevo, nuevo_pn)
                    if zfor_nuevo:
                        self.mm02_actualizar_partnumber(zfor_nuevo, nuevo_pn)
                    print(f"     PARTNUMBER → {nuevo_pn}")
                else:
                    print("     PARTNUMBER: sin cambio necesario")
            except Exception as e_pn:
                print(f"     [WARN] Paso 5 PARTNUMBER: {e_pn}")

            res.estado    = "OK"
            res.fecha_fin = datetime.datetime.now()
            print(f"   ✓  OK  ({res.duracion_seg}s)")

        except Exception as e:
            res.estado    = "ERROR"
            res.error     = str(e)
            res.fecha_fin = datetime.datetime.now()
            print(f"   ✗  ERROR: {e}")

        return res

    # ── Procesamiento en lote ─────────────────────────────────────────────────

    def procesar_lote(self, combinaciones: list) -> list:
        """
        Procesa todas las combinaciones activas (no bloqueadas).
        combinaciones: lista de ItemColor del VISTAAAA.py

        Uso desde VISTAAAA.py:
            items_activos = [it for it in vista.items.values()
                             if not it.bloqueado and not it.pendiente]
            auto = AutomatizadorSAP()
            resultados = auto.procesar_lote(items_activos)
        """
        if not self.conectar():
            raise RuntimeError("No se pudo conectar a SAP GUI.")

        print(f"\n{'='*60}")
        print(f"  LOTE SAP — Cambio de Color")
        print(f"  Batch ID: {self.batch_id}")
        print(f"  Combinaciones: {len(combinaciones)}")
        print(f"{'='*60}\n")

        if not combinaciones:
            print("  Sin combinaciones activas para procesar.")
            return []

        # Leer clasificacion del ZFER base para determinar P_FRANJ
        zfer_base = combinaciones[0].zfer_origen
        print(f"  Leyendo MM02 del ZFER base {zfer_base}...")
        try:
            clasif_base = self.leer_clasificacion_zfer(zfer_base)
            p_franj     = clasif_base.franja if clasif_base.franja else "00"
        except Exception as e:
            print(f"  [WARN] No se pudo leer MM02 base: {e}")
            p_franj = "00"
        print(f"  P_FRANJ determinado: {p_franj}\n")

        self.resultados = []
        for item in combinaciones:
            # P_COLOR = número al inicio del nombre de color
            # ej: "19-Gray Light Automotive" → "19"
            p_color = self._extraer_numero_color(item.color)

            res = self.procesar_combinacion(
                zfer_base  = item.zfer_origen,
                formula    = item.formula,
                acero      = item.acero, #aca entra logica
                color      = item.color,
                cod_pieza  = getattr(item, "cod_pieza",  ""),
                tipo_pieza = getattr(item, "tipo_pieza", ""),
                p_color    = p_color,
                p_franj    = p_franj,
            )
            self.resultados.append(res)
            self._log_bd(res)

        # Generar reporte final
        ruta = self._generar_reporte()
        self._imprimir_resumen()

        return self.resultados

    # ── BD: log de ejecucion ──────────────────────────────────────────────────

    def _log_bd(self, res: ResultadoCombinacion):
        """Inserta/actualiza el resultado en M5_LogEjecucion."""
        try:
            s = (
                f"DRIVER={{{DB_LOCAL['driver']}}};"
                f"SERVER={DB_LOCAL['server']};"
                f"DATABASE={DB_LOCAL['database']};"
                "Trusted_Connection=yes;"
            )
            conn = pyodbc.connect(s, autocommit=True)
            conn.cursor().execute("""
                INSERT INTO dbo.M5_LogEjecucion
                    (batch_id, pedido_origen, tipo_pieza, formula,
                     color_codigo, acero_variante, estado,
                     detalle_error, fecha_inicio, fecha_fin)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                res.batch_id,
                res.zfer_base,
                (res.tipo_pieza or "N/A")[:20],
                res.formula,
                res.color,
                res.acero,
                res.estado,
                (f"ZFER_NUEVO={res.zfer_nuevo} | ZFOR={res.zfor_nuevo} | "
                 f"POS={','.join(res.posiciones_bom)}")
                if res.estado == "OK" else res.error[:4000],
                res.fecha_inicio,
                res.fecha_fin,
            ))
            conn.close()
        except pyodbc.Error as e:
            print(f"    [WARN] BD log: {e}")

    # ── Reporte Excel ─────────────────────────────────────────────────────────

    def _generar_reporte(self) -> str:
        """
        Genera reporte Excel detallado con 4 hojas:
          RESUMEN         — batch info, totales, timing
          PROCESADOS_SAP  — items ejecutados en SAP (OK y ERROR)
          SOLO_REPORTE    — items no procesados (fórmula diferente)
          ERRORES         — solo los errores, referencia rápida
        """
        import json

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = os.path.join(BASE_DIR, f"reporte_sap_{timestamp}.xlsx")

        # ── Paleta ────────────────────────────────────────────────────────────────
        NAVY      = "2B3A47"
        BLANCO    = "FFFFFF"
        AZUL_AGP  = "7DBFD4"
        VERDE_H   = "D5F5E3"
        VERDE_TXT = "1A7340"
        ROJO_H    = "FADBD8"
        ROJO_TXT  = "922B21"
        AMBAR_H   = "FEF3CD"
        AMBAR_TXT = "9A5B00"
        GRIS_H    = "F0F4F8"

        def _hdr_fill():
            return PatternFill("solid", fgColor=NAVY)

        def _hdr_font():
            return Font(color=BLANCO, bold=True, name="Segoe UI", size=10)

        def _borde():
            return Border(
                bottom=Side(style="thin", color="CCCCCC"),
                right= Side(style="thin", color="DDDDDD"),
            )

        def _escribir_fila_encabezado(ws, headers, anchos):
            for col, (h, w) in enumerate(zip(headers, anchos), 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill      = _hdr_fill()
                c.font      = _hdr_font()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border    = _borde()
                ws.column_dimensions[c.column_letter].width = w
            ws.row_dimensions[1].height = 24

        wb = openpyxl.Workbook()

        # ════════════════════════════════════════════════════════════════════════
        # HOJA 1 — PROCESADOS_SAP
        # ════════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "PROCESADOS_SAP"
        ws1.freeze_panes = "A2"

        hdrs1  = ["#", "ZFER Base", "Formula", "Acero", "Color",
                   "ZFER Nuevo", "ZFOR Nuevo", "Posiciones BOM",
                   "Estado", "Detalle / Error",
                   "Inicio", "Fin", "Duración (s)"]
        anch1  = [5, 14, 12, 9, 35, 14, 14, 22, 11, 48, 20, 20, 13]
        _escribir_fila_encabezado(ws1, hdrs1, anch1)

        ok_n = err_n = 0
        duracion_total = 0.0

        for r_idx, res in enumerate(self.resultados, 2):
            if res.estado   == "OK":
                fill = PatternFill("solid", fgColor=VERDE_H)
                ok_n += 1
            elif res.estado == "ERROR":
                fill = PatternFill("solid", fgColor=ROJO_H)
                err_n += 1
            else:
                fill = PatternFill("solid", fgColor=AMBAR_H)

            dur = res.duracion_seg
            if dur:
                duracion_total += dur

            fila = [
                r_idx - 1,
                res.zfer_base,
                res.formula,
                res.acero,
                res.color,
                res.zfer_nuevo,
                res.zfor_nuevo,
                ", ".join(res.posiciones_bom),
                res.estado,
                res.error if res.estado == "ERROR" else (
                    f"ZFER={res.zfer_nuevo} | ZFOR={res.zfor_nuevo} | "
                    f"POS={','.join(res.posiciones_bom)}"
                ),
                res.fecha_inicio.strftime("%Y-%m-%d %H:%M:%S") if res.fecha_inicio else "",
                res.fecha_fin.strftime("%Y-%m-%d %H:%M:%S")    if res.fecha_fin    else "",
                dur,
            ]
            for c_idx, val in enumerate(fila, 1):
                c = ws1.cell(row=r_idx, column=c_idx, value=val)
                c.fill      = fill
                c.border    = _borde()
                c.font      = Font(name="Segoe UI", size=9)
                c.alignment = Alignment(vertical="center")
            ws1.row_dimensions[r_idx].height = 18

        # ════════════════════════════════════════════════════════════════════════
        # HOJA 2 — SOLO_REPORTE (fórmula diferente — pendiente cambio formula)
        # ════════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("SOLO_REPORTE")
        ws2.freeze_panes = "A2"

        hdrs2 = ["#", "ZFER Base", "Formula", "Acero", "Color",
                  "Cod Pieza", "Tipo Pieza", "Motivo"]
        anch2 = [5, 14, 12, 9, 35, 14, 14, 55]
        _escribir_fila_encabezado(ws2, hdrs2, anch2)

        fill_sr = PatternFill("solid", fgColor=AMBAR_H)
        for i, it in enumerate(self.items_solo_reporte, 2):
            fila2 = [
                i - 1,
                it.get("zfer_base",  ""),
                it.get("formula",    ""),
                it.get("acero",      ""),
                it.get("color",      ""),
                it.get("cod_pieza",  ""),
                it.get("tipo_pieza", ""),
                it.get("motivo",     "Fórmula diferente — requiere cambio de fórmula previo"),
            ]
            for c_idx, val in enumerate(fila2, 1):
                c = ws2.cell(row=i, column=c_idx, value=val)
                c.fill      = fill_sr
                c.border    = _borde()
                c.font      = Font(name="Segoe UI", size=9)
                c.alignment = Alignment(vertical="center")
            ws2.row_dimensions[i].height = 18

        # ════════════════════════════════════════════════════════════════════════
        # HOJA 3 — ERRORES (referencia rápida técnico SAP)
        # ════════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("ERRORES")
        ws3.freeze_panes = "A2"

        hdrs3 = ["#", "Formula", "Acero", "Color", "ZFER Base", "Error completo",
                  "Inicio", "Duración (s)"]
        anch3 = [5, 12, 9, 35, 14, 70, 20, 13]
        _escribir_fila_encabezado(ws3, hdrs3, anch3)

        fill_err = PatternFill("solid", fgColor=ROJO_H)
        errores = [r for r in self.resultados if r.estado == "ERROR"]
        for i, res in enumerate(errores, 2):
            fila3 = [
                i - 1,
                res.formula,
                res.acero,
                res.color,
                res.zfer_base,
                res.error,
                res.fecha_inicio.strftime("%Y-%m-%d %H:%M:%S") if res.fecha_inicio else "",
                res.duracion_seg,
            ]
            for c_idx, val in enumerate(fila3, 1):
                c = ws3.cell(row=i, column=c_idx, value=val)
                c.fill      = fill_err
                c.border    = _borde()
                c.font      = Font(name="Segoe UI", size=9,
                                   color=ROJO_TXT if c_idx == 6 else "000000")
                c.alignment = Alignment(vertical="center", wrap_text=(c_idx == 6))
            ws3.row_dimensions[i].height = 28

        # ════════════════════════════════════════════════════════════════════════
        # HOJA 4 — RESUMEN
        # ════════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("RESUMEN")

        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ini_str = (self.resultados[0].fecha_inicio.strftime("%Y-%m-%d %H:%M:%S")
                   if self.resultados and self.resultados[0].fecha_inicio else "—")

        resumen_bloques = [
            ("INFORMACIÓN DEL LOTE", [
                ("Batch ID",            self.batch_id),
                ("Fecha ejecución",     ini_str),
                ("Fecha reporte",       now_str),
                ("Operador",            self._usuario),
                ("ZFER Base",           self.resultados[0].zfer_base if self.resultados else "—"),
                ("Fórmula base",        self.formula_base or "—"),
            ]),
            ("RESULTADOS SAP", [
                ("Items procesados en SAP",     len(self.resultados)),
                ("Exitosos",                    ok_n),
                ("Con error",                   err_n),
                ("Duración total SAP (s)",      round(duracion_total, 1)),
                ("Duración promedio/item (s)",
                 round(duracion_total / len(self.resultados), 1) if self.resultados else 0),
            ]),
            ("SOLO REPORTE (pendiente cambio fórmula)", [
                ("Items no procesados",         len(self.items_solo_reporte)),
                ("Motivo",                      "Fórmula distinta a la base — procesar con ZMME0001 cambio fórmula"),
            ]),
            ("ARCHIVOS", [
                ("Reporte Excel",       ruta),
                ("JSON progreso",       self._ruta_json),
            ]),
        ]

        row = 1
        fill_bloque  = PatternFill("solid", fgColor=NAVY)
        fill_ok_sum  = PatternFill("solid", fgColor=VERDE_H)
        fill_err_sum = PatternFill("solid", fgColor=ROJO_H)
        fill_sr_sum  = PatternFill("solid", fgColor=AMBAR_H)
        fill_info    = PatternFill("solid", fgColor=GRIS_H)

        for bloque_titulo, items_bloque in resumen_bloques:
            # Título de bloque
            c = ws4.cell(row=row, column=1, value=bloque_titulo)
            c.fill      = fill_bloque
            c.font      = Font(color=BLANCO, bold=True, name="Segoe UI", size=10)
            c.alignment = Alignment(vertical="center")
            ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws4.row_dimensions[row].height = 20
            row += 1

            for k, v in items_bloque:
                # Determinar fill por contexto
                if "Exitosos" in k:
                    fill_row = fill_ok_sum
                elif "error" in k.lower():
                    fill_row = fill_err_sum
                elif "no procesados" in k.lower():
                    fill_row = fill_sr_sum
                else:
                    fill_row = fill_info

                c1 = ws4.cell(row=row, column=1, value=k)
                c1.fill      = fill_row
                c1.font      = Font(bold=True, name="Segoe UI", size=10)
                c1.border    = _borde()
                c1.alignment = Alignment(vertical="center")

                c2 = ws4.cell(row=row, column=2, value=v)
                c2.fill      = fill_row
                c2.font      = Font(name="Segoe UI", size=10)
                c2.border    = _borde()
                c2.alignment = Alignment(vertical="center", wrap_text=True)

                ws4.row_dimensions[row].height = 18
                row += 1

            row += 1  # espacio entre bloques

        ws4.column_dimensions["A"].width = 38
        ws4.column_dimensions["B"].width = 65

        # Ordenar hojas: Resumen primero
        wb.move_sheet("RESUMEN", offset=-wb.index(wb["RESUMEN"]))

        wb.save(ruta)
        print(f"\n  Reporte: {ruta}")

        # Finalizar JSON con fecha_fin
        try:
            with open(self._ruta_json, "r", encoding="utf-8") as f:
                jdata = json.load(f)
            jdata["fecha_fin"] = datetime.datetime.now().isoformat()
            jdata["totales"]["procesados"] = len(self.resultados)
            jdata["totales"]["exitosos"]   = ok_n
            jdata["totales"]["errores"]    = err_n
            with open(self._ruta_json, "w", encoding="utf-8") as f:
                json.dump(jdata, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        return ruta

    def _imprimir_resumen(self):
        ok  = sum(1 for r in self.resultados if r.estado == "OK")
        err = sum(1 for r in self.resultados if r.estado == "ERROR")
        print(f"\n{'='*60}")
        print(f"  RESUMEN  —  Batch {self.batch_id[:8]}...")
        print(f"  Total   : {len(self.resultados)}")
        print(f"  Exitosas: {ok}")
        print(f"  Errores : {err}")
        print(f"{'='*60}\n")


# ── Punto de entrada para prueba directa ─────────────────────────────────────

def main():
    """
    Prueba de conexion y lectura MM02.
    Para el lote completo, llamar desde VISTAAAA.py:
        auto = AutomatizadorSAP()
        resultados = auto.procesar_lote(items_activos)
    """
    print("\n" + "="*60)
    print("  MODULO 5 — Automatizador SAP: Cambio de Color")
    print("="*60)

    auto = AutomatizadorSAP()
    if not auto.conectar():
        input("\n  Presiona Enter para salir...")
        return

    # Prueba: leer clasificacion de un ZFER
    zfer_test = input("\n  ZFER base para probar lectura MM02: ").strip()
    if zfer_test:
        clasif = auto.leer_clasificacion_zfer(zfer_test)
        print(f"\n  Resultado MM02:")
        print(f"    PARTNUMBER : {clasif.partnumber}")
        print(f"    COLOR      : {clasif.color}")
        print(f"    FRANJA     : {clasif.franja}")
        print(f"    P_FRANJ    : {'01' if clasif.tiene_franja else '00'}")

    input("\n  Presiona Enter para salir...")


if __name__ == "__main__":
    main()
