"""
app.py — MODULO 5 AGP Glass
Sistema de consulta y reporte de ZFERs (Colombia CO01)
"""
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, abort, session, flash
import pyodbc
import threading
import os
import mimetypes
from functools import lru_cache, wraps
from concurrent.futures import ThreadPoolExecutor

app = Flask(__name__)
app.secret_key = "AGP_M5_2025_xK9!mQ#zL"

# ── Usuarios autorizados ──────────────────────────────────────────────────────
_USUARIOS = {
    "atcol@agpglass.com":              "AdminIng2025_It",
    "fguerrero@agpglass.com":          "1022438939",
    "jguanumen@agpglass.com":          "1023005676",
    "alexander.acosta@agpglass.com":   "93437119",
    "g.delgado@agpglass.com":          "1031180571",
    "kmorales@agpglass.com":           "1233501014",
    "jpinzon@agpglass.com":            "1030596420",
    "lpelaez@agpglass.com":            "1000047853",
    "mbernal@agpglass.com":            "1000007660",
    "nirojas@agpglass.com":            "1030688452",
    "asuarez@agpglass.com":            "1030690990",
    "dgrimaldo@agpglass.com":          "1000236441",
    "nleon@agpglass.com":              "1137624222",
    "jramirezf@agpglass.com":          "1031420151",
    "pract1@agpglass.com":             "PRACT_ING1",
    "pract2@agpglass.com":             "PRACT_ING2",
    "pract3@agpglass.com":             "PRACT_ING3",
    "pract4@agpglass.com":             "PRACT_ING4",
    "spina@agpglass.com":              "1010236538",
    "lcruz@agp_usuaglass.com":              "1032937021",
    "jgalvis@agpglass.com":            "1032877183",
    "spimentel@agpglass.com":          "1000034924",
    "practingenieria@agpglass.com":    "1000971646",
    "jmahecha@agpglass.com":           "1019982163",
    "dforero@agpglass.com":            "1000256251",
    "cegarcia@agpglass.com":           "1001092159",
    "lfalla@agpglass.com":             "1022930033",
    "leo@agpglass.com":                "123",
    "prueba@agpglass.com":              "prueba123"
}

def _usuario_actual() -> str:
    return session.get("usuario", "")

@app.context_processor
def _inject_usuario():
    return {"usuario_actual": _usuario_actual()}

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("usuario"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper

# ── Estado en memoria de jobs SAP activos  {batch_id: ResultadoItem} ─────────
_sap_jobs: dict = {}

# ── Configuración BD ──────────────────────────────────────────────────────────
DB_SAP = {
    "server":   "agpcolsap.database.windows.net",
    "database": "DB_COL_SAP",
    "driver":   "ODBC Driver 17 for SQL Server",
    "user":     "Viewer",
    "password": "AgpconsCol2023",
}

def _conn_str():
    return (
        f"DRIVER={{{DB_SAP['driver']}}};"
        f"SERVER={DB_SAP['server']};"
        f"DATABASE={DB_SAP['database']};"
        f"UID={DB_SAP['user']};"
        f"PWD={DB_SAP['password']};"
        "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=20;"
    )

_CONN_CALENDARIO = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=agpcolcalendario.database.windows.net;"
    "DATABASE=CalendarioAGP;"
    "UID=Consulta;"
    "PWD=@GPgl4$$2021;"
    "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=20;"
)

# ── Pool de conexiones (evita reconexión TCP en cada query) ───────────────────
import queue as _queue

_conn_pool: "_queue.Queue[pyodbc.Connection]" = _queue.Queue(maxsize=12)

class _PooledConn:
    """Wrapper que devuelve la conexión al pool en lugar de cerrarla."""
    __slots__ = ("_c",)
    def __init__(self, c): self._c = c
    def cursor(self): return self._c.cursor()
    def execute(self, *a, **kw): return self._c.execute(*a, **kw)
    def close(self):
        try:
            _conn_pool.put_nowait(self._c)
        except _queue.Full:
            try: self._c.close()
            except Exception: pass
    # Soporte para "with get_conn() as cn:" que usan algunos endpoints
    def __enter__(self): return self
    def __exit__(self, *_): self.close()

def get_conn() -> "_PooledConn":
    """Obtiene una conexión del pool; crea una nueva si el pool está vacío."""
    try:
        c = _conn_pool.get_nowait()
        try:
            c.execute("SELECT 1")   # health-check rápido
        except Exception:
            try: c.close()
            except Exception: pass
            c = pyodbc.connect(_conn_str(), autocommit=True)
    except _queue.Empty:
        c = pyodbc.connect(_conn_str(), autocommit=True)
    return _PooledConn(c)


# ── Catálogos ─────────────────────────────────────────────────────────────────
PIEZAS = {
    "000": "Parabrisas",
    "001": "Lateral Delantero Izquierdo", "002": "Lateral Delantero Derecho",
    "003": "Lateral Trasero Izquierdo",   "004": "Lateral Trasero Derecho",
    "005": "Ventilete Trasero Izquierdo", "006": "Ventilete Trasero Derecho",
    "007": "Cabina Trasera Izquierda",    "008": "Cabina Trasera Derecha",
    "009": "Posterior",                   "010": "Techo Solar Delantero",
    "011": "Lateral Extendido Izquierdo", "012": "Lateral Extendido Derecho",
    "013": "Posterior Izquierdo",         "014": "Posterior Derecho",
    "015": "Claraboya Izquierda",         "016": "Claraboya Derecha",
    "017": "Mirilla",                     "018": "Probeta",
    "019": "Ventilete Delantero Izquierdo","020": "Ventilete Delantero Derecho",
    "021": "Cabina Delantera Izquierda",  "022": "Cabina Delantera Derecha",
    "023": "Cabina Superior Izquierda",   "024": "Cabina Superior Derecha",
    "025": "Techo Solar B",               "026": "Parabrisas Derecho",
    "027": "Parabrisas Izquierdo",        "028": "Lateral Secundario Derecho",
    "029": "Lateral Secundario Izquierdo","030": "Partición",
    "031": "Arquitectura",                "034": "Porthole 1",
    "035": "Porthole 2",                  "036": "Porthole 3",
    "037": "Porthole 4",                  "040": "Pummel",
    "085": "Posterior Secundario",        "087": "Techo Solar Céntrico",
    "088": "Techo Solar D",               "090": "Techo Solar Panorámico",
    "091": "Probeta 2",  "092": "Probeta 3", "093": "Probeta Especial",
    "094": "Probeta 4",  "095": "Kit Opaco", "096": "Probeta 5",
    "097": "Probeta 6",
    "110": "Techo Solar A — Paquete",     "125": "Techo Solar B — Paquete",
    "187": "Techo Solar C — Paquete",     "190": "Techo Solar Panorámico — Paquete",
}
# Pares simétricos izquierda ↔ derecha (solo aplica a piezas con ambos lados)
_PARES_SIMETRIA = {
    "001": "002", "002": "001",  # Lateral Delantero
    "003": "004", "004": "003",  # Lateral Trasero
    "005": "006", "006": "005",  # Ventilete Trasero
    "007": "008", "008": "007",  # Cabina Trasera
    "011": "012", "012": "011",  # Lateral Extendido
    "013": "014", "014": "013",  # Posterior Izq/Der
    "015": "016", "016": "015",  # Claraboya
    "019": "020", "020": "019",  # Ventilete Delantero
    "021": "022", "022": "021",  # Cabina Delantera
    "023": "024", "024": "023",  # Cabina Superior
    "026": "027", "027": "026",  # Parabrisas Der/Izq
    "028": "029", "029": "028",  # Lateral Secundario
}

for _i in range(1, 20):
    PIEZAS[f"{40+_i:03d}"] = f"Pieza Especial {_i}"
for _i in range(1, 11):
    PIEZAS[f"{59+_i:03d}"] = f"Vidrio Especial {_i}"
for _i, _n in enumerate([25, 26, 27, 28], 70):
    PIEZAS[f"{_i:03d}"] = f"Pieza Plana Especial {_n}"
for _i in range(80, 87):
    PIEZAS[f"{_i:03d}"] = "Vidrio Especial Laminado"

COLORES = {
    "NA": "No Aplica",       "00": "Blanco",
    "01": "Green Light",     "02": "Bronze Light",
    "03": "Azul",            "04": "Gray Light",
    "05": "Gray Light PC",   "06": "Gray Light Glass",
    "07": "Verde",           "08": "Bronze Medium",
    "09": "Gray Medium",     "10": "Gray Medium PC",
    "11": "Bronze Dark",     "12": "Gray Dark",
    "13": "Gray Dark Glass", "14": "Parsol Gray",
    "15": "Privacy",         "16": "Clear",
    "17": "Solar Green",     "18": "Gray Medium Glass",
    "19": "Gray Light Automotive",
    "20": "Gray Medium Automotive + PC", 
    "21": "Gray Dark Automotive + PC",
    "22": "G2 Gray Medium Automotive",
    "23": "G2 Gray Dark Automotive",
}

# Únicos colores habilitados para combinaciones (los demás no se muestran ni procesan)
_COLORES_ACTIVOS = {"00","01","05","06","10","13","18","19","20","21","22","23"}

FRANJAS = {
    "00": "Sin Franja", "01": "Franja Azul",
    "02": "Franja Verde","03": "Franja Gris",
    "NA": "No Aplica",
}

DIFERENCIALES = {
    "01": "SOLAR PLUS",
    "02": "LIGHT WEIGHT",
    "03": "MULTI HIT",
    "04": "SUN ADVANCED",
    "05": "EXTREME PROTECT",
    "06": "STEEL PLUS",
    "07": "TNT",
    "08": "TNT FLEX",
    "09": "SUN BAND",
    "10": "GUNPORT",
    "11": "VARIO PLUS",
    "12": "AGP DURA P",
    "13": "AGP DURA NPC",
    "14": "AGP DURA G",
    "15": "HIGH PERFORMANCE",
    "16": "FRAMES",
    "17": "CLAMP",
    "18": "METALLIC SUPPORT FOR MIRROR",
    "19": "HEATING - METALLIC COATING",
    "20": "HEATING - WIRED - HEATPLEX",
    "21": "ANTIREFLECTIVE",
    "22": "SILVER PASTE",
    "23": "N.A",
    "24": "ENCAPSULATED - FRAMES",
}

SUBPRODUCTOS = {
    "B1":"B33","B2":"iB33","B3":"STANDARD","B4":"AGP PREMIUM","B5":"AGP TITANIUM",
    "B6":"OEM","B7":"ARCHITECTURAL BRG","B8":"B33 ESPECIAL","B9":"iB33 ESPECIAL",
    "B10":"STANDAR 40mm","B11":"STANDAR 45mm","B12":"3KL","B13":"3KL DURA P",
    "B14":"Antitheft","B15":"Impenetra Plus","B16":"Impenetra Plus DURA P",
    "B17":"NBR15000 II-A","B18":"iB33X",
    "D1":"LAND","D2":"NAVY","D3":"ARCHITECTURAL DEFENSE",
    "EA1":"STANDARD LAMINATED GLASS ARG","EA2":"STANDARD TOUGHENED GLASS ARG",
    "EA3":"MONOLITIC GLASS","EAV1":"AVO ONLY",
    "EL1":"ULT LAMINATED GLASS","EL2":"ULT LAMINATED GLASS W/AVO",
    "EL3":"STANDARD LAMINATED GLASS","EL4":"STAND LAMINATED GLASS W/AVO",
    "EL5":"SUP ULT LAMINATED GLASS","EL6":"SUP ULT LAMINATED GLASS W/AVO",
    "ET1":"ULTRALITE TOUGHNED GLASS","ET2":"ULTRALITE TOUGHNED GLASS W/AVO",
    "ET3":"STANDARD TOUGHENED  GLASS","ET4":"STAND TOUGHENED  GLASS  W/AVO",
    "M1":"BR6 Stoof","M2":"BR7 Stoof Opción 1","M3":"BR7 Stoof Opción 2 (con Heating)",
    "M4":"BR7 Farmingtons","M5":"Light Weigh DURA + SRF 14mm",
    "M6":"Estándar 18mm + SRF 14mm","M7":"VPAM 3 + SRF 14mm",
    "M8":"N5 Plasan Combinado","M9":"Marine NB",
    "M10":"Light Weight 28 y 30mm y VPAM 3","M11":"Estándar 42mm y VPAM 3","M12":"BR7 Ang 27G",
    "P1":"BR5 North Glass","P2":"Estándar 45mm North Glass","P3":"Nivel 4 Plasán",
    "S1":"Samples R&D",
    "X1":"21mm MH (POS LW 19mm & SRF 18mm)","X2":"21mm MH (SRF 18mm)",
    "X3":"B33 17mm","X4":"B33 17mm DURA P","X5":"B33 23mm  DURA",
    "X6":"B33 23mm (SRF 14mm)","X7":"B33 23mm (SRF 18mm)","X8":"B33 23mm DURA P",
    "X9":"B33 30mm - DURA","X10":"B33 43mm","X11":"B33 43mm DURA P",
    "X12":"BMW OEM VPAM 6","X13":"BR5 Tinted Galron","X14":"BR7 Stoof",
    "X15":"Estándar 18mm","X16":"Estándar 18mm DURA P","X17":"Estándar 21mm",
    "X18":"Estándar 21mm DURA P","X19":"Estándar 32mm","X20":"Estándar 32mm DURA P",
    "X21":"Estándar 33mm","X22":"Estándar 39mm","X23":"Estándar 40mm",
    "X24":"Estándar 40mm con Acero E","X25":"Estándar 40mm DURA P",
    "X26":"Estándar 42mm","X27":"Estándar 45mm","X28":"Estándar 45mm DURA P",
    "X29":"Estándar 48mm","X30":"Estándar 56mm","X31":"Estándar 56mm DURA P",
    "X32":"Estándar 58mm","X33":"Estándar 58mm DURA P","X34":"Estándar 60mm",
    "X35":"Estándar 60mm DURA P","X36":"Estándar 73mm","X37":"Estándar 76mm",
    "X38":"Estándar 76mm DURA P","X39":"Estándar 79mm","X40":"Estándar 79mm DURA P",
    "X41":"Estándar 82mm DURA P","X42":"Estándar 88mm","X43":"Estándar 88mm DURA P",
    "X44":"Estándar 97mm","X45":"Estándar 97mm DURA P",
    "X46":"Estándar BR3 17mm","X47":"Estándar BR3 18mm",
    "X48":"Estándar BR3 18mm DURA P","X49":"Estándar BR3 20mm DURA P",
    "X50":"Estándar Stop Gun 13mm","X51":"Estándar Stop Gun 14mm",
    "X52":"Estándar VPAM 3 15mm","X53":"Estándar VPAM 3 15mm DURA","X54":"GL43-01",
    "X55":"Light Weight 110mm DURA P","X56":"Light Weight 115mm DURA P",
    "X57":"Light Weight 19mm","X58":"Light Weight 19mm DURA","X59":"Light Weight 28mm",
    "X60":"Light Weight 28mm VSAG12","X61":"Light Weight 30mm",
    "X62":"Light Weight 30mm DURA P","X63":"Light Weight 36mm",
    "X64":"Light Weight 36mm DURA P","X65":"Light Weight 50mm",
    "X66":"Light Weight 50mm DURA P","X67":"Light Weight 50mm NORTH GLASS",
    "X68":"Light Weight 52mm DURA P","X69":"Light Weight 62mm",
    "X70":"Light Weight 62mm DURA P","X71":"Light Weight 69mm DURA P",
    "X72":"LW 19mm (LT´s 21mm & SRF 18mm)","X73":"LW 19mm (SRF Laminado)",
    "X74":"LW 69mm (PBS 50mm)","X75":"Marine Estándar 29mm",
    "X76":"Marine VPAM CL9 FRIGATTE 66 mm","X77":"Matine Estándar 40mm",
    "X78":"Mix 21mm STD y LW 19mm","X79":"MultiHit 21mm","X80":"MultiHit 21mm DURA P",
    "X81":"MultiHit 32mm","X82":"MultiHit 42mm","X83":"MultiHit 42mm DURA P",
    "X84":"NP58-2","X85":"NPC 85mm","X86":"PE NIJ III 38mm Blinsecurity",
    "X87":"PE STANAG 1 65mm TATRA","X88":"PE STANAG 1 Rheinmetall",
    "X89":"PE STANAG 2 60mm DURA P NIMR","X90":"PE WBS Rheinmetall",
    "X91":"Stop Gun 13mm DURA P","X92":"VPAM 3 15mm DURA P",
    "X93":"Estándar BR3 20mm","X94":"Estándar 24mm","X95":"Estándar 31mm",
    "X96":"N5 WBS Plasan 36mm","X97":"Estándar 40mm (Outer Glass 6mm)",
    "X98":"N5 WBS Plasan 43mm","X99":"Estándar 44mm","X100":"Estándar 45mm USA",
    "X101":"Estándar 47mm","X102":"BR7 Ang 55G Stoof","X103":"VPAM 9 Ang 55G Stoof",
    "X104":"Estándar 70mm","X105":"Estándar 72mm","X106":"Light Weight 66mm DURA P",
    "X107":"Estándar 71mm DURA P","X108":"Estándar 22mm","X109":"Marine NB 124",
    "X110":"Marine NB 155","X111":"Marine NB 124-1","X112":"Marine NB 103-2",
    "X113":"WBS 3 + 2","X114":"UL10 61mm",
    "X115":"Light Weight 30mm VPAM 6 OuterGlass 6mm","X116":"Estándar 48mm DURA P",
    "X117":"Estándar 67mm DURA P","X118":"Estándar 67mm","X119":"iB33 PLUS (FIJAS LW)",
    "X120":"AGP HEAT","X121":"Light Weight 66mm","X122":"L28CG y L28SCG",
    "X123":"3KL DUPA  P","X124":"B33 ESPECIAL VOLVO","X125":"iB33 NG",
    "X128":"B33 GEN2","X130":"Sunroof VPAM CL2 16mm DURA NPC",
    "X131":"Sunroof VPAM CL3 18mm DURA NPC","X132":"Estándar VPAM 2 11mm",
    "X139":"GL25-1","X140":"B33 EXPORTACIÓN LW","X141":"Ultra-Lightweight 12mm",
    "X142":"Envostar","X143":"iB33 G6","X145":"iB33 G6 EXPORTAÇÃO",
    "X146":"Estándar 30mm","X147":"Estándar VPAM CL2 17mm DURA NPC",
    "X149":"Marine NB 20mm","X150":"B33 23mm","X151":"B33 28mm",
    "X152":"LAMINADO 11mm","X153":"Estándar 50mm PBS",
    "X154":"Light Weight 18mm ARGENTINA","X155":"Estándar 32mm GALRON",
    "X156":"Light Weight 19mm LATAM","X157":"WBS 28mm DURA P ENVOSTAR",
    "X158":"Estándar 28mm CAM","X159":"Light Weight 70mm","X160":"Estándar 45mm URO",
    "X161":"Estándar 83mm","X162":"Estándar 69mm TPS MÉXICO",
    "X163":"Estándar 80mm DURA P","X164":"Estándar 102 DURA P NMIR",
    "X165":"Estándar 74mm DURA P PBS TENCATE",
    "X166":"Estándar 38mm DURA P NORTHGLASS","X167":"Estándar 38mm DURA P",
    "X168":"Estándar 43mm","X169":"Estándar 44mm DURA P USA",
    "X170":"Estándar 42mm NORTHGLASS","X171":"Estándar 88mm URO",
    "X172":"Estándar 55mm PBS ANGULO","X173":"Estándar 66mm PBS ANGULO",
    "X174":"Estándar 70mm STOOF","X175":"Estándar 82mm",
    "X176":"Estándar 29mm DURA P","X177":"Estándar 45mm ALEMANIA",
    "X178":"Estándar 44mm DURA P NORTHGLASS","X179":"Estándar 145mm DURA P",
    "X180":"Estándar 26mm DURA P MARINE","X181":"Estándar 38mm NORTHGLASS",
    "X182":"Estándar 85mm PBS","X183":"Estándar 86mm RHEINMETAL",
    "X184":"Estándar 61mm DURA P TENCATE","X185":"WBS 33mm DURA P ENVOSTAR",
    "X186":"Estándar 47mm DURA P","X187":"Estándar 43mm DURA P",
    "X189":"WBS 35mm PLASAN","X190":"Estándar 32mm DURA P NORTHGLASS",
    "X191":"Estándar 32mm NO EUROPA","X192":"Estándar 70mm DURA P",
    "X193":"Estándar 30mm DURA P","X194":"Estándar 50mm DURA P JANKEL",
    "X195":"Stop Gun 17mm DURA P NORTHGLASS","X196":"Estándar 84mm DURA P",
    "X197":"Exclusivo USA 49mm DOS","X198":"Estándar 86mm",
    "X199":"Estándar 60mm SENTINEL","X200":"Estándar 65mm",
    "X201":"Estándar 97mm CAMBLI","X202":"Estándar 46mm DURA P NORTHGLASS",
    "X203":"Estándar 62mm","X204":"Estándar 81mm MEXICO","X205":"Estándar 51mm GREIT",
    "X206":"Estándar 54mm GREIT","X207":"Estándar 33mm PLASAN",
    "X208":"Estándar 74mm DURA P TENCTE","X209":"Multihit 114mm DURA P NIMR",
    "X210":"Multihit 110mm DURA P NIMR","X211":"Estándar 42mm LICITACIÓN",
    "X212":"Arquitectónico 26mm DURA P","X213":"Light Weight 71mm",
    "X214":"WBS 7mm PLASAN","X215":"WBS 5mm NIMR","X216":"Exclusivo USA 52mm",
    "X217":"Estándar 22mm DURA P BOON EDAM",
    "X218":"Doble cara impacto 27mm DURA P BOON EDAM",
    "X219":"Estándar 22mm DURA P NORTHGLASS","X220":"WBS 21mm DURA P PMMA",
    "X221":"WBS 26mm  DURA P PMMA","X222":"FGR 47mm DURA NPC NAVANTIA",
    "X223":"Estándar 65mm PBS TNTF","X224":"WBS  35mm MIRLL AEROSPACE DURA P",
    "X225":"Estándar 47mm DEFENTURE","X226":"Light Weight 65mm DURA P",
    "X227":"Estándar 83mm DURA P AEROSPACE","X228":"WBS 8mm TECNOGETAFE",
    "X231":"Estándar 95mm","X233":"Multihit 43mm","X237":"OSOP","X241":"Estándar 51mm",
}

PAISES = {
    "AE":"Emiratos Árabes Unidos","AF":"Afganistán","AR":"Argentina",
    "AT":"Austria","AU":"Australia","AX":"Islas de Åland","BE":"Bélgica",
    "BH":"Baréin","BO":"Bolivia","BR":"Brasil","BY":"Bielorrusia",
    "CA":"Canadá","CH":"Suiza","CL":"Chile","CN":"China","CO":"Colombia",
    "CR":"Costa Rica","CZ":"República Checa","DE":"Alemania","DK":"Dinamarca",
    "DM":"Dominica","DO":"República Dominicana","EC":"Ecuador","EG":"Egipto",
    "ES":"España","FI":"Finlandia","FR":"Francia","GB":"United Kingdom",
    "GR":"Grecia","GT":"Guatemala","HK":"Hong Kong","HN":"Honduras",
    "HR":"Croacia","HT":"Haití","ID":"Indonesia","IL":"Israel","IN":"India",
    "IQ":"Iraq","IT":"Italia","JE":"Jersey","JO":"Jordania","JP":"Japón",
    "KE":"Kenia","KR":"Corea del Sur","LB":"Líbano","MA":"Marruecos",
    "MX":"México","MY":"Malasia","NG":"Nigeria","NL":"Holanda","NO":"Noruega",
    "OM":"Omán","PA":"Panamá","PE":"Perú","PG":"Papúa Nueva Guinea",
    "PH":"Filipinas","PK":"Pakistán","PL":"Polonia","PR":"Puerto Rico",
    "PT":"Portugal","PY":"Paraguay","QA":"Qatar","RO":"Rumanía","RS":"Serbia",
    "SA":"Arabia Saudí","SE":"Suecia","SG":"Singapur","SK":"Eslovaquia",
    "SV":"El Salvador","TH":"Tailandia","TR":"Turquía","TW":"Taiwán",
    "US":"Estados Unidos","UY":"Uruguay","VE":"Venezuela","YE":"Yemen",
    "ZA":"Sudáfrica",
}

ATNAM_LABELS = {
    "Z_VEHICLE_MODEL":          "Modelo Vehículo",
    "Z_SUBPRODUCT":             "Subproducto",
    "Z_FORMULA_CODE":           "Fórmula",
    "Z_COLOR":                  "Color",
    "Z_PIECE_TYPE":             "Tipo de Pieza",
    "Z_SHADE_BAND":             "Franja",
    "Z_AGP_LEVEL":              "Nivel AGP",
    "Z_BEHAVIOR_DIFFERENTIALS": "Differentials",
    "Z_COMMERCIAL_THICKNESS":   "Espesor Comercial",
    "Z_AGP_VERSION":            "Versión AGP",
    "Z_AGP_PARTNUMBER":         "Partnumber AGP",
}

def _decode_route(route: str) -> str:
    """Intenta decodificar código de ruta SAP a nombre de país."""
    if not route:
        return "Sin ruta"
    r = route.strip().upper()
    if r in PAISES:
        return PAISES[r]
    # Formato "XX-YY": intentar prefijo y sufijo
    if "-" in r:
        partes = r.split("-")
        for p in reversed(partes):   # sufijo primero
            if p in PAISES:
                return PAISES[p]
    # Primeros 2 chars
    if len(r) >= 2 and r[:2] in PAISES:
        return PAISES[r[:2]]
    return route


# ── Planos ───────────────────────────────────────────────────────────────────
# Nombre real de la tabla en DB_COL_SAP — ajustar si es diferente
_TABLA_PLANOS = "ODATA_ZFER_RUTAS_JPG"

_plano_cache: dict = {}   # {material: (ruta, doc) | None}

def _q_plano(material: str):
    """Retorna (ruta_unc, documento) o None. Cacheado en _plano_cache."""
    if material in _plano_cache:
        return _plano_cache[material]
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute(
            f"SELECT TOP 1 PLANO, DOCUMENTO FROM dbo.{_TABLA_PLANOS} "
            "WHERE MATERIAL=? AND CENTRO='CO01' ORDER BY VERSION DESC, PROCESSDATE DESC",
            (material,)
        )
        row = conn.cursor().fetchone() if False else cur.fetchone()
        conn.close()
        result = (str(row[0]).strip(), str(row[1]).strip() if row[1] else "") if row and row[0] else None
        _plano_cache[material] = result
        return result
    except Exception:
        return None


def _q_planos_bulk(mats: list) -> None:
    """Un solo IN query para poblar _plano_cache con todos los materiales no cacheados."""
    uncached = [m for m in mats if m not in _plano_cache]
    if not uncached:
        return
    try:
        conn = get_conn()
        cur  = conn.cursor()
        ph   = ",".join(["?"] * len(uncached))
        cur.execute(
            f"SELECT MATERIAL, PLANO, DOCUMENTO FROM dbo.{_TABLA_PLANOS} "
            f"WHERE MATERIAL IN ({ph}) AND CENTRO='CO01'",
            uncached
        )
        
        seen = set()
        for mat, plano, doc in cur.fetchall():
            m = str(mat)
            if m not in seen:
                seen.add(m)
                _plano_cache[m] = (str(plano).strip(), str(doc).strip() if doc else "") if plano else None
        conn.close()
        for m in uncached:           # los que no están en la tabla → None
            if m not in _plano_cache:
                _plano_cache[m] = None
    except Exception:
        pass

def _normalizar_unc(ruta: str) -> str:
    """Garantiza que la ruta empiece con \\\\ (UNC válido en Windows)."""
    ruta = ruta.strip().replace("/", "\\")
    if ruta.startswith("\\\\"):
        return ruta
    return "\\\\" + ruta.lstrip("\\")

@app.route("/api/plano/<material>")
def api_plano(material: str):
    info = _q_plano(material.strip())
    if not info:
        abort(404)
    ruta = _normalizar_unc(info[0])
    if not os.path.isfile(ruta):
        abort(404)
    mime = mimetypes.guess_type(ruta)[0] or "image/jpeg"
    return send_file(ruta, mimetype=mime, max_age=3600)


@app.route("/api/planos/batch")
def api_planos_batch():
    """Retorna {material: documento} — un solo IN query para todos los no cacheados."""
    mats = [m.strip() for m in request.args.get("mats", "").split(",") if m.strip()][:60]
    _q_planos_bulk(mats)           # 1 query para todos los que faltan en cache
    result = {}
    for m in mats:
        info = _q_plano(m)         # ahora todos son cache hits
        if info:
            result[m] = info[1]
    return jsonify(result)


# ── Queries ───────────────────────────────────────────────────────────────────

@lru_cache(maxsize=400)
def q_zfer_head(material: str):
    """Tabla 1: ODATA_ZFER_HEAD — info básica del ZFER."""
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT MATERIAL, CENTRO, TEXTO_BREVE_MATERIAL, STATUS,
                   ZFOR, GRUPO_ARTICULOS, CREADO_EL, ULTIMA_MOD, AREA
            FROM   dbo.ODATA_ZFER_HEAD
            WHERE  MATERIAL    = ?
              AND  CENTRO      = 'CO01'
              AND  UPPER(ISNULL(STATUS,'')) != 'ZZ'
        """, (material,))
        row = cur.fetchone()
        cols = [c[0] for c in cur.description]
        conn.close()
        return dict(zip(cols, row)) if row else None
    except Exception as e:
        return {"_error": str(e)}
    
@lru_cache(maxsize=400)
def q_atributos(material: str) -> dict:
    """Tabla 2: ODATA_ZFER_CLASS_001 — atributos de clasificación."""
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT ATNAM,
                   CASE WHEN ATNAM = 'Z_COMMERCIAL_THICKNESS' THEN CAST(ATFLV AS VARCHAR(50)) ELSE ATWRT END AS valor
            FROM   dbo.ODATA_ZFER_CLASS_001
            WHERE  MATERIAL = ?
              AND  CENTRO   = 'CO01'
              AND  ATNAM IN (
                'Z_AGP_LEVEL','Z_BEHAVIOR_DIFFERENTIALS','Z_VEHICLE_MODEL',
                'Z_AGP_PARTNUMBER','Z_SUBPRODUCT','Z_COLOR','Z_FORMULA_CODE',
                'Z_COMMERCIAL_THICKNESS','Z_AGP_VERSION','Z_PIECE_TYPE','Z_SHADE_BAND',
                'Z_GEOMETRY_TYPE'
              )
        """, (material,))
        rows = cur.fetchall()
        conn.close()
        return {r[0]: str(r[1]).strip() if r[1] is not None else "" for r in rows}
    except Exception as e:
        return {"_error": str(e)}


def q_entregas(material: str) -> list:
    """Tabla 3: ODATA_ZCDS_Entregas_Pos_CO — números de entrega (ntgew > 0)."""
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT DISTINCT entrega
            FROM   dbo.ODATA_ZCDS_Entregas_Pos_CO
            WHERE  matnr = ?
              AND  TRY_CAST(ntgew AS FLOAT) > 0
        """, (material,))
        rows = cur.fetchall()
        conn.close()
        return [str(r[0]) for r in rows if r[0] is not None]
    except Exception:
        return []

#AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
#necesitaba expresarme
#sigamos 
def _parsear_partnumber(pn: str) -> dict | None:
    """Parsea '1490_008_L23-26_12_000' → {vehiculo, version, formula, color, pieza}."""
    if not pn:
        return None
    parts = pn.strip().split("_")
    if len(parts) != 5:
        return None
    return {"vehiculo": parts[0], "version": parts[1], "formula": parts[2],
            "color": parts[3], "pieza": parts[4]}


def q_variantes_por_pn(vehiculo: str, version: str, formula: str, pieza: str) -> list:
    """
    Busca ZFERs activos (no ZZ) en CO01 cuyo PARTNUMBER comparte vehiculo+version+
    formula+pieza con cualquier color. Una sola query con JOINs.
    """
    try:
        conn = get_conn()
        cur  = conn.cursor()
        def _esc(s):
            return s.replace("!", "!!").replace("%", "!%").replace("_", "!_")
        pattern = "!_".join([_esc(vehiculo), _esc(version), _esc(formula), "%", _esc(pieza)])

        cur.execute("""
            SELECT
                c.MATERIAL,
                c.ATWRT                                    AS partnumber,
                MAX(CASE WHEN a.ATNAM='Z_COLOR'      THEN a.ATWRT END) AS color,
                MAX(CASE WHEN a.ATNAM='Z_SHADE_BAND' THEN a.ATWRT END) AS franja,
                h.STATUS,
                h.TEXTO_BREVE_MATERIAL
            FROM   dbo.ODATA_ZFER_CLASS_001 c
            JOIN   dbo.ODATA_ZFER_HEAD h
                ON h.MATERIAL = c.MATERIAL AND h.CENTRO = 'CO01'
            LEFT JOIN dbo.ODATA_ZFER_CLASS_001 a
                ON a.MATERIAL = c.MATERIAL AND a.CENTRO = 'CO01'
               AND a.ATNAM IN ('Z_COLOR', 'Z_SHADE_BAND')
            WHERE  c.CENTRO = 'CO01'
              AND  c.ATNAM  = 'Z_AGP_PARTNUMBER'
              AND  c.ATWRT  LIKE ? ESCAPE '!'
              AND  UPPER(ISNULL(h.STATUS,'')) != 'ZZ'
            GROUP BY c.MATERIAL, c.ATWRT, h.STATUS, h.TEXTO_BREVE_MATERIAL
            ORDER BY c.MATERIAL
        """, (pattern,))
        rows = cur.fetchall()
        conn.close()

        resultado = []
        for mat, pn, color_raw, franja_raw, status, texto in rows:
            cr = str(color_raw).strip() if color_raw else ""
            resultado.append({
                "material":     mat,
                "partnumber":   pn,
                "color_raw":    cr,
                "color_nombre": COLORES.get(cr, cr) if cr else "—",
                "franja_raw":   str(franja_raw).strip() if franja_raw else "",
                "status":       str(status).strip() if status else "",
                "texto":        str(texto).strip() if texto else "",
            })
        return resultado
    except Exception as e:
        return [{"_error": str(e)}]



def q_zplas_compatibles(formula_code: str, piece_type: str,
                        shade_band: str = "", differentials_base: str = "",
                        tiene_acero_base: bool | None = None) -> list:
    """
    Busca ZPLAs compatibles con la fórmula, tipo de pieza, franja y diferencial del ZFER base.

    tiene_acero_base (detectado por diferencial 06 del ZFER base):
      False → solo ZPLAs SIN diferencial 06 (sin acero)
      True  → solo ZPLAs CON diferencial 06 (con acero)
      None  → sin filtro por acero
    """
    if not formula_code or not piece_type:
        return []
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT
                c.MATERIAL,
                MAX(CASE WHEN c.ATNAM = 'Z_COLOR'                  THEN c.ATWRT ELSE NULL END) AS color,
                MAX(CASE WHEN c.ATNAM = 'Z_PIECE_TYPE'             THEN c.ATWRT ELSE NULL END) AS piece_types,
                MAX(CASE WHEN c.ATNAM = 'Z_SHADE_BAND'             THEN c.ATWRT ELSE NULL END) AS shade_band,
                MAX(CASE WHEN c.ATNAM = 'Z_BEHAVIOR_DIFFERENTIALS' THEN c.ATWRT ELSE NULL END) AS differentials,
                MAX(CASE WHEN c.ATNAM = 'Z_AGP_LEVEL'              THEN c.ATWRT ELSE NULL END) AS level
            FROM dbo.ODATA_ZPLA_CLASS_001 c
            JOIN dbo.ODATA_ZPLA_HEAD h
              ON h.MATERIAL = c.MATERIAL AND h.CENTRO = 'CO01'
            WHERE c.CENTRO   = 'CO01'
              AND c.TIPO_MAT = 'ZPLA'
              AND UPPER(ISNULL(h.STATUS, '')) != 'ZZ'
              AND c.MATERIAL IN (
                SELECT MATERIAL FROM dbo.ODATA_ZPLA_CLASS_001
                WHERE CENTRO = 'CO01' AND TIPO_MAT = 'ZPLA'
                  AND ATNAM  = 'Z_FORMULA_CODE' AND ATWRT = ?
              )
            GROUP BY c.MATERIAL
        """, (formula_code,))
        rows = cur.fetchall()
        conn.close()

        base_diffs = {d.strip() for d in differentials_base.split(",") if d.strip()}

        resultado = []
        for mat, color, piece_types_str, zpla_shade, differentials, level in rows:
            if not color:
                continue
            # Z_PIECE_TYPE multi-valor
            pieces = [p.strip() for p in (piece_types_str or "").split(",") if p.strip()]
            if piece_type not in pieces:
                continue
            # Franja
            if shade_band and shade_band not in ("00", ""):
                if (zpla_shade or "00") not in (shade_band, "00"):
                    continue
            # Diferencial base vs ZPLA
            zpla_diffs = {d.strip() for d in (differentials or "").split(",") if d.strip()}
            if base_diffs:
                if zpla_diffs and not base_diffs.intersection(zpla_diffs):
                    continue
            # Filtro acero: el diferencial 06 indica acero
            if tiene_acero_base is not None:
                zpla_tiene_acero = "06" in zpla_diffs
                if tiene_acero_base != zpla_tiene_acero:
                    continue
            resultado.append({
                "material":      mat,
                "color":         color.strip(),
                "color_nombre":  COLORES.get(color.strip(), color.strip()),
                "shade_band":    zpla_shade or "00",
                "differentials": differentials or "",
                "level":         level or "",
            })
        return sorted(resultado, key=lambda x: x["color"])
    except Exception as e:
        return [{"_error": str(e)}]

def q_formulas_por_pieza(piece_type: str, nivel: str, subproducto: str,
                          formula_base: str) -> list:
    """
    Busca fórmulas alternativas disponibles para el mismo tipo de pieza / nivel / subproducto.
    Retorna lista de dicts: {formula, colores: [{zpla, color, color_nombre, differentials}]}
    Excluye la fórmula base del ZFER.
    """
    if not piece_type:
        return []
    try:
        conn = get_conn()
        cur  = conn.cursor()
        pt_in    = piece_type.strip()
        niv_in   = nivel.strip()       if nivel       else ""
        sub_in   = subproducto.strip() if subproducto else ""
        fbase_in = formula_base.strip() if formula_base else ""

        cur.execute("""
            WITH
            ZPLA_HEAD AS (
                SELECT MATERIAL
                FROM   dbo.ODATA_ZPLA_HEAD
                WHERE  STATUS IS NULL
            ),
            ZPLA_001 AS (
                SELECT MATERIAL,
                       ATNAM,
                       LTRIM(RTRIM(ATWRT)) AS ATWRT
                FROM   dbo.ODATA_ZPLA_CLASS_001
                WHERE  ATNAM IN (
                    'Z_FORMULA_CODE', 'Z_COLOR', 'Z_PIECE_TYPE',
                    'Z_AGP_LEVEL', 'Z_SUBPRODUCT', 'Z_BEHAVIOR_DIFFERENTIALS'
                )
            )
            SELECT
                h.MATERIAL,
                MAX(CASE WHEN a.ATNAM = 'Z_FORMULA_CODE'           THEN a.ATWRT END) AS formula,
                MAX(CASE WHEN a.ATNAM = 'Z_COLOR'                  THEN a.ATWRT END) AS color,
                MAX(CASE WHEN a.ATNAM = 'Z_BEHAVIOR_DIFFERENTIALS' THEN a.ATWRT END) AS differentials,
                MAX(CASE WHEN a.ATNAM = 'Z_SUBPRODUCT'             THEN a.ATWRT END) AS subproducto
            FROM   ZPLA_HEAD h
            LEFT JOIN ZPLA_001 a ON a.MATERIAL = h.MATERIAL
            GROUP BY h.MATERIAL
            HAVING
                MAX(CASE WHEN a.ATNAM = 'Z_FORMULA_CODE' THEN a.ATWRT END) IS NOT NULL
                AND MAX(CASE WHEN a.ATNAM = 'Z_COLOR'    THEN a.ATWRT END) IS NOT NULL
                -- piece_type: campo CSV => CHARINDEX para buscar el valor dentro de la lista
                AND (? = '' OR CHARINDEX(?, ISNULL(MAX(CASE WHEN a.ATNAM = 'Z_PIECE_TYPE' THEN a.ATWRT END),'')) > 0)
                -- filtro nivel
                AND (? = '' OR MAX(CASE WHEN a.ATNAM = 'Z_AGP_LEVEL'  THEN a.ATWRT END) = ?)
                -- filtro subproducto
                AND (? = '' OR MAX(CASE WHEN a.ATNAM = 'Z_SUBPRODUCT' THEN a.ATWRT END) = ?)
                -- excluir la fórmula base del ZFER actual
                AND ISNULL(MAX(CASE WHEN a.ATNAM = 'Z_FORMULA_CODE' THEN a.ATWRT END), '') <> ?
            ORDER BY formula, color
        """, (pt_in, pt_in,
              niv_in, niv_in,
              sub_in, sub_in,
              fbase_in))
        rows = cur.fetchall()
        conn.close()

        print(f"  [q_formulas] piece={pt_in!r} nivel={niv_in!r} sub={sub_in!r} fbase={fbase_in!r} → {len(rows)} filas")
        if rows:
            r0 = rows[0]
            print(f"  [q_formulas] ejemplo fila0: mat={r0[0]} formula={r0[1]} color={r0[2]} differentials={r0[3]}")

        # Agrupar por fórmula → color (un color puede tener varios ZPLAs)
        formulas: dict = {}
        # subproducto_por_formula guarda el subproducto de cada fórmula
        subproducto_por_formula: dict = {}
        for row in rows:
            mat, formula, color, differentials, subprod = row[0], row[1], row[2], row[3], row[4]
            color_key = str(color).strip()
            zpla_str  = str(mat).strip()
            diff_str  = differentials or ""
            sub_str   = str(subprod or "").strip()
            # guardar subproducto por fórmula (primer valor no vacío)
            if formula not in subproducto_por_formula and sub_str:
                subproducto_por_formula[formula] = sub_str
            if formula not in formulas:
                formulas[formula] = {}
            if color_key not in formulas[formula]:
                formulas[formula][color_key] = {
                    "color":         color_key,
                    "color_nombre":  COLORES.get(color_key, color_key),
                    "differentials": diff_str,
                    "zpla_list":     [zpla_str],
                }
            else:
                if zpla_str not in formulas[formula][color_key]["zpla_list"]:
                    formulas[formula][color_key]["zpla_list"].append(zpla_str)
                # usa los differentials del zpla con acero si existe
                if "06" in diff_str.split(","):
                    formulas[formula][color_key]["differentials"] = diff_str

        result = []
        for f, color_dict in sorted(formulas.items()):
            colores = []
            for c in color_dict.values():
                zpla_list = c["zpla_list"]
                colores.append({
                    "zpla":          zpla_list[0],
                    "zpla_list":     zpla_list,
                    "zpla_count":    len(zpla_list),
                    "color":         c["color"],
                    "color_nombre":  c["color_nombre"],
                    "differentials": c["differentials"],
                })
            result.append({
                "formula":     f,
                "colores":     colores,
                "subproducto": subproducto_por_formula.get(f, ""),
            })
        return result
    except Exception as e:
        return [{"_error": str(e)}]


def q_mercados(entregas: list) -> list:
    """Tabla 4: ODATA_ZCDS_Entregas_Head_CO — conteo por route/mercado."""
    if not entregas:
        return []
    try:
        conn = get_conn()
        cur  = conn.cursor()
        ph   = ",".join(["?"] * len(entregas))
        cur.execute(f"""
            SELECT   route, COUNT(*) AS total
            FROM     dbo.ODATA_ZCDS_Entregas_Head_CO
            WHERE    entrega IN ({ph})
              AND    ISNULL(route,'') != ''
            GROUP BY route
            ORDER BY total DESC
        """, entregas)
        rows = cur.fetchall()
        conn.close()
        return [
            {"route": r[0], "pais": _decode_route(r[0]), "total": r[1]}
            for r in rows
        ]
    except Exception:
        return []

def q_explorar(vehiculo="", formula="", pieza="", color="", version="", nivel="",
               cod_vehiculo="", zfers_lista: list = None) -> list:
    """
    Busca ZFERs activos (no ZZ) en CO01 según filtros opcionales (LIKE parcial).
    Si se pasa zfers_lista, busca exactamente esos ZFERs y los enriquece con atributos.
    Retorna lista de dicts con los atributos clave de cada ZFER.
    Máximo 300 resultados.
    """
    def _esc(s):
        return s.replace("!", "!!").replace("%", "!%").replace("_", "!_")

    try:
        conn = get_conn()
        cur  = conn.cursor()

        if zfers_lista:
            # Búsqueda directa por lista de materiales
            ph = ",".join(["?"] * len(zfers_lista))
            cur.execute(f"""
                SELECT MATERIAL FROM dbo.ODATA_ZFER_HEAD
                WHERE  MATERIAL IN ({ph}) AND CENTRO = 'CO01'
                  AND  UPPER(ISNULL(STATUS,'')) != 'ZZ'
            """, zfers_lista)
            materiales = list(dict.fromkeys(str(r[0]) for r in cur.fetchall()))
        else:
            # Búsqueda por filtros con INTERSECT dinámico
            filtros = [
                ("Z_VEHICLE_MODEL", vehiculo.strip()),
                ("Z_FORMULA_CODE",  formula.strip()),
                ("Z_PIECE_TYPE",    pieza.strip()),
                ("Z_COLOR",         color.strip()),
                ("Z_AGP_VERSION",   version.strip()),
                ("Z_AGP_LEVEL",     nivel.strip()),
            ]
            activos = [(a, v) for a, v in filtros if v]

            # Un solo scan con OR + GROUP BY/HAVING en lugar de N INTERSECTs
            or_parts, params = [], []
            for atnam, val in activos:
                or_parts.append("(c.ATNAM=? AND c.ATWRT LIKE ? ESCAPE '!')")
                params.extend([atnam, f"%{_esc(val)}%"])
            # Código vehículo: prefijo del PARTNUMBER (ej: "1715" → "1715_...")
            if cod_vehiculo.strip():
                or_parts.append("(c.ATNAM='Z_AGP_PARTNUMBER' AND c.ATWRT LIKE ? ESCAPE '!')")
                params.append(f"{_esc(cod_vehiculo.strip())}!_%")

            if not or_parts:
                conn.close()
                return []
            
            n = len(activos) + (1 if cod_vehiculo.strip() else 0)
            cur.execute(f"""
                SELECT TOP 300 c.MATERIAL
                FROM dbo.ODATA_ZFER_CLASS_001 c
                JOIN dbo.ODATA_ZFER_HEAD h
                  ON h.MATERIAL = c.MATERIAL AND h.CENTRO = 'CO01'
                WHERE c.CENTRO = 'CO01'
                  AND UPPER(ISNULL(h.STATUS,'')) != 'ZZ'
                  AND ({" OR ".join(or_parts)})
                GROUP BY c.MATERIAL
                HAVING COUNT(DISTINCT c.ATNAM) >= {n}
                ORDER BY c.MATERIAL
            """, params)
            materiales = list(dict.fromkeys(str(r[0]) for r in cur.fetchall()))

        if not materiales:
            conn.close()
            return []

        ph = ",".join(["?"] * len(materiales))

        # Atributos clave para mostrar en tabla
        cur.execute(f"""
            SELECT MATERIAL, ATNAM, ATWRT
            FROM   dbo.ODATA_ZFER_CLASS_001
            WHERE  CENTRO = 'CO01' AND MATERIAL IN ({ph})
              AND  ATNAM IN (
                'Z_VEHICLE_MODEL','Z_FORMULA_CODE','Z_COLOR',
                'Z_PIECE_TYPE','Z_AGP_VERSION','Z_AGP_PARTNUMBER',
                'Z_SHADE_BAND','Z_BEHAVIOR_DIFFERENTIALS','Z_AGP_LEVEL'
              )
        """, materiales)
        pivot = {}
        for mat, atnam, atwrt in cur.fetchall():
            pivot.setdefault(str(mat), {})[atnam] = str(atwrt).strip() if atwrt is not None else ""

        # Cabecera (status, descripción, ZFOR)
        cur.execute(f"""
            SELECT MATERIAL, STATUS, TEXTO_BREVE_MATERIAL, ZFOR
            FROM   dbo.ODATA_ZFER_HEAD
            WHERE  CENTRO = 'CO01' AND MATERIAL IN ({ph})
        """, materiales)
        head_d = {str(r[0]): {"status": str(r[1]).strip() if r[1] is not None else "",
                          "texto":  str(r[2]).strip() if r[2] is not None else "",
                          "zfor":   str(r[3]).strip() if r[3] is not None else ""}
                  for r in cur.fetchall()}
        conn.close()
        
        resultado = []
        for mat in sorted(materiales):
            d = pivot.get(mat, {})
            h = head_d.get(mat, {})
            color_raw  = d.get("Z_COLOR", "")
            pieza_raw  = d.get("Z_PIECE_TYPE", "")
            resultado.append({
                "material":      mat,
                "texto":         h.get("texto", ""),
                "status":        h.get("status", ""),
                "zfor":          h.get("zfor", ""),
                "vehiculo":      d.get("Z_VEHICLE_MODEL", ""),
                "formula":       d.get("Z_FORMULA_CODE", ""),
                "color_raw":     color_raw,
                "color_nombre":  COLORES.get(color_raw, color_raw),
                "pieza_raw":     pieza_raw,
                "pieza_nombre":  PIEZAS.get(pieza_raw, pieza_raw),
                "version":       d.get("Z_AGP_VERSION", ""),
                "partnumber":    d.get("Z_AGP_PARTNUMBER", ""),
                "shade_band":    d.get("Z_SHADE_BAND", ""),
                "differentials": d.get("Z_BEHAVIOR_DIFFERENTIALS", ""),
                "nivel":         d.get("Z_AGP_LEVEL", ""),
            })
        return resultado
    except Exception as e:
        return [{"_error": str(e)}]


@lru_cache(maxsize=30)
def q_valores_distintos(atnam: str) -> list:
    """Devuelve los 200 valores ATWRT distintos más frecuentes para un ATNAM en CO01."""
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT TOP 200 ATWRT, COUNT(*) AS n
            FROM   dbo.ODATA_ZFER_CLASS_001
            WHERE  CENTRO = 'CO01' AND ATNAM = ?
              AND  ISNULL(ATWRT,'') != ''
            GROUP BY ATWRT
            ORDER BY n DESC
        """, (atnam,))
        rows = cur.fetchall()
        conn.close()
        return [r[0] for r in rows]
    except Exception:
        return []


# ── Rutas Flask ───────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("usuario"):
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        clave = request.form.get("clave", "").strip()
        if _USUARIOS.get(email) == clave:
            session["usuario"] = email
            session.permanent = True
            return redirect(request.args.get("next") or url_for("index"))
        error = "Correo o contraseña incorrectos."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    if request.method == "POST":
        raw = request.form.get("zfer", "").strip()
        if not raw:
            return render_template("index.html", error=None)
        # Si hay comas → multi-ZFER → explorar
        zfers = [z.strip() for z in raw.replace(";", ",").split(",") if z.strip()][:12]
        if len(zfers) > 1:
            return redirect(url_for("explorar") + "?zfers=" + ",".join(zfers))
        return redirect(url_for("detalle_zfer", material=zfers[0]))
    return render_template("index.html", error=None)


@app.route("/explorar")
@login_required
def explorar():
    vehiculo = request.args.get("vehiculo", "").strip()
    formula  = request.args.get("formula",  "").strip()
    pieza    = request.args.get("pieza",    "").strip()
    color    = request.args.get("color",    "").strip()
    version  = request.args.get("version",  "").strip()
    nivel        = request.args.get("nivel",        "").strip()
    cod_vehiculo = request.args.get("cod_vehiculo", "").strip()
    zfers_qs     = request.args.get("zfers",        "").strip()
    
    zfers_lista = [z.strip() for z in zfers_qs.split(",") if z.strip()][:12] if zfers_qs else []

    hay_filtros = any([vehiculo, formula, pieza, color, version, nivel, cod_vehiculo]) or bool(zfers_lista)
    resultados  = []
    error       = None

    if hay_filtros:
        resultados = q_explorar(vehiculo, formula, pieza, color, version, nivel, cod_vehiculo, zfers_lista or None)
        if resultados and "_error" in resultados[0]:
            error      = resultados[0]["_error"]
            resultados = []

    # Autocomplete: solo carga hints cuando el usuario ya busca (evita 2 queries extra en carga inicial)
    vehiculos_hints = q_valores_distintos("Z_VEHICLE_MODEL") if hay_filtros else []
    formulas_hints  = q_valores_distintos("Z_FORMULA_CODE")  if hay_filtros else []

    return render_template("explorar.html",
        vehiculo        = vehiculo,
        formula         = formula,
        pieza           = pieza,
        color           = color,
        version         = version,
        nivel           = nivel,
        cod_vehiculo    = cod_vehiculo,
        zfers_qs        = zfers_qs,
        resultados      = resultados,
        error           = error,
        hay_filtros     = hay_filtros,
        modo_lista      = bool(zfers_lista),
        vehiculos_hints = vehiculos_hints,
        formulas_hints  = formulas_hints,
        COLORES         = COLORES,
        PIEZAS          = PIEZAS,
        FRANJAS         = FRANJAS,
    )


@app.route("/zfer/<material>")
@login_required
def detalle_zfer(material: str):
    material = material.strip()

    # Las 3 queries son independientes → las lanzamos en paralelo
    with ThreadPoolExecutor(max_workers=3) as ex:
        f_head     = ex.submit(q_zfer_head, material)
        f_attrs    = ex.submit(q_atributos, material)
        f_entregas = ex.submit(q_entregas,  material)

    head     = f_head.result()
    attrs    = f_attrs.result()
    entregas = f_entregas.result()

    if head is None:
        return render_template("index.html",
            error=f"ZFER '{material}' no encontrado o STATUS = ZZ (inactivo).")
    if "_error" in head:
        return render_template("index.html",
            error=f"Error de conexión BD: {head['_error']}")

    mercados = q_mercados(entregas)
 
    # Construir lista de atributos para mostrar (en orden definido)
    attrs_display = []
    for atnam, label in ATNAM_LABELS.items():
        val = attrs.get(atnam, "")
        if not val:
            continue
        decoded = val
        if atnam == "Z_COLOR":
            decoded = f"{val} — {COLORES.get(val, val)}"
        elif atnam == "Z_PIECE_TYPE":
            decoded = f"{val} — {PIEZAS.get(val, val)}"
        elif atnam == "Z_SHADE_BAND":
            decoded = f"{val} — {FRANJAS.get(val, val)}"
        attrs_display.append({
            "atnam":   atnam,
            "label":   label,
            "raw":     val,
            "decoded": decoded,
        })

    total_entregas = sum(m["total"] for m in mercados)
    # Top 15 para el gráfico; el resto en la tabla
    mercados_chart = mercados[:15]

    plano_info = _q_plano(material)
    plano = {"doc": plano_info[1], "tiene": True} if plano_info else None

    return render_template("zfer.html",
        material       = material,
        head           = head,
        attrs_display  = attrs_display,
        entregas_n     = len(entregas),
        mercados       = mercados,
        mercados_chart = mercados_chart,
        total_entregas = total_entregas,
        DIFERENCIALES  = DIFERENCIALES,
        SUBPRODUCTOS   = SUBPRODUCTOS,
        plano          = plano,
    )

def _cargar_datos_zfer(material: str, nivel: str = "", subproducto: str = "") -> dict:
    """Carga head + attrs + 3 queries en paralelo para un ZFER. Retorna dict listo para template."""
    # Stage 1: head y attrs en paralelo (evitan 1 round-trip secuencial)
    with ThreadPoolExecutor(max_workers=2) as ex:
        fut_head  = ex.submit(q_zfer_head, material)
        fut_attrs = ex.submit(q_atributos, material)
    head  = fut_head.result()
    attrs = fut_attrs.result()

    if head is None or (isinstance(head, dict) and "_error" in head):
        return {"_error": head}

    formula_code  = attrs.get("Z_FORMULA_CODE",         "")
    piece_type    = attrs.get("Z_PIECE_TYPE",            "")
    color_base    = attrs.get("Z_COLOR",                 "")
    shade_band    = attrs.get("Z_SHADE_BAND",            "00") or "00"
    partnumber    = attrs.get("Z_AGP_PARTNUMBER",        "")
    vehicle_model = attrs.get("Z_VEHICLE_MODEL",         "")
    thickness     = attrs.get("Z_COMMERCIAL_THICKNESS",  "")
    differentials = attrs.get("Z_BEHAVIOR_DIFFERENTIALS","")
    nivel         = nivel or attrs.get("Z_AGP_LEVEL",    "")
    subproducto   = subproducto or attrs.get("Z_SUBPRODUCT", "")
    pn_parsed     = _parsear_partnumber(partnumber)

    # Detectar si el ZFER base tiene acero: diferencial 06 = con acero
    tiene_acero_base = "06" in {d.strip() for d in differentials.split(",") if d.strip()}

    # Stage 2: 3 queries dependientes de attrs, en paralelo
    with ThreadPoolExecutor(max_workers=3) as ex:
        fut_variantes = ex.submit(
            q_variantes_por_pn,
            pn_parsed["vehiculo"], pn_parsed["version"],
            pn_parsed["formula"],  pn_parsed["pieza"]
        ) if pn_parsed else None
        fut_zplas    = ex.submit(q_zplas_compatibles, formula_code, piece_type, shade_band, differentials, tiene_acero_base)
        fut_formulas = ex.submit(q_formulas_por_pieza, piece_type, nivel, subproducto, formula_code)

    variantes    = fut_variantes.result() if fut_variantes else []
    zplas        = fut_zplas.result()
    formulas_alt = fut_formulas.result()

    if variantes    and "_error" in variantes[0]:    variantes    = []
    if zplas        and "_error" in zplas[0]:        zplas        = []
    if formulas_alt and "_error" in formulas_alt[0]: formulas_alt = []

    colores_con_zfer = {v["color_raw"]: v for v in variantes if v.get("color_raw")}
    colores_con_zpla: dict = {}
    for z in zplas:
        colores_con_zpla.setdefault(z["color"], []).append(z)

    matrix = []
    for cod, nombre in COLORES.items():
        if cod not in _COLORES_ACTIVOS:
            continue
        zfer_v    = colores_con_zfer.get(cod)
        zpla_list = colores_con_zpla.get(cod, [])
        estado = "EXISTE" if zfer_v else ("DISPONIBLE" if zpla_list else "SIN_ZPLA")
        matrix.append({
            "color_codigo": cod, "color_nombre": nombre, "estado": estado,
            "zfer":      zfer_v["material"]       if zfer_v    else "",
            "zfer_texto":zfer_v["texto"]          if zfer_v    else "",
            "zfer_pn":   zfer_v["partnumber"]     if zfer_v    else "",
            "zpla":      zpla_list[0]["material"] if zpla_list else "",
            "zpla_count":len(zpla_list),
            "zpla_list": [z["material"] for z in zpla_list],
            "es_base":   cod == color_base,
        })

    pn_pattern_ui = "_".join([pn_parsed["vehiculo"], pn_parsed["version"],
                               pn_parsed["formula"], "**", pn_parsed["pieza"]]) if pn_parsed else ""
    return dict(
        head=head, attrs=attrs, formula_code=formula_code, piece_type=piece_type,
        piece_nombre=PIEZAS.get(piece_type, piece_type), color_base=color_base,
        shade_band=shade_band, partnumber=partnumber, vehicle_model=vehicle_model,
        thickness=thickness, differentials=differentials, nivel=nivel, subproducto=subproducto,
        pn_parsed=pn_parsed, pn_pattern_ui=pn_pattern_ui,
        variantes=variantes, zplas=zplas, formulas_alt=formulas_alt, matrix=matrix,
        n_existe=sum(1 for c in matrix if c["estado"]=="EXISTE"),
        n_disponible=sum(1 for c in matrix if c["estado"]=="DISPONIBLE"),
        n_sin_zpla=sum(1 for c in matrix if c["estado"]=="SIN_ZPLA"),
    )


@app.route("/combinaciones/<material>")
@login_required
def combinaciones(material: str):
    material      = material.strip()
    sim_material  = request.args.get("simetrico", "").strip()

    # Cargar ZFER principal y simétrico en paralelo si aplica
    with ThreadPoolExecutor(max_workers=2) as ex:
        fut_main = ex.submit(_cargar_datos_zfer, material)
        fut_sim  = ex.submit(_cargar_datos_zfer, sim_material) if sim_material else None

    d = fut_main.result()
    if d.get("_error"):
        err = d["_error"]
        msg = f"ZFER '{material}' no encontrado o STATUS = ZZ (inactivo)." if err is None else f"Error BD: {err}"
        return render_template("index.html", error=msg)

    sim = fut_sim.result() if fut_sim else None
    if sim and sim.get("_error"):
        sim = None  # simétrico no disponible: carga normal sin panel lateral

    return render_template("combinaciones.html",
        material       = material,
        head           = d["head"],
        vehicle_model  = d["vehicle_model"],
        formula_code   = d["formula_code"],
        piece_type     = d["piece_type"],
        piece_nombre   = d["piece_nombre"],
        color_base     = d["color_base"],
        shade_band     = d["shade_band"],
        thickness      = d["thickness"],
        differentials  = d["differentials"],
        nivel          = d["nivel"],
        subproducto    = d["subproducto"],
        partnumber     = d["partnumber"],
        pn_parsed      = d["pn_parsed"],
        pn_pattern_ui  = d["pn_pattern_ui"],
        variantes      = d["variantes"],
        zplas          = d["zplas"],
        matrix         = d["matrix"],
        n_existe       = d["n_existe"],
        n_disponible   = d["n_disponible"],
        n_sin_zpla     = d["n_sin_zpla"],
        formulas_alt   = d["formulas_alt"],
        # Panel simétrico
        sim_material   = sim_material if sim else "",
        sim            = sim,
        SUBPRODUCTOS   = SUBPRODUCTOS,
    )


# ── API Bloqueos ───────────────────────────────────────────────────────────────

@app.route("/api/formulas_alt")
@login_required
def api_formulas_alt():
    """Devuelve fórmulas alternativas filtradas dinámicamente."""
    piece_type   = request.args.get("piece_type",   "").strip()
    nivel        = request.args.get("nivel",        "").strip()
    subproducto  = request.args.get("subproducto",  "").strip()
    formula_base = request.args.get("formula_base", "").strip()
    if not piece_type:
        return jsonify([])
    result = q_formulas_por_pieza(piece_type, nivel, subproducto, formula_base)
    if result and "_error" in result[0]:
        return jsonify({"error": result[0]["_error"]}), 500
    return jsonify(result)


@app.route("/api/subproductos")
@login_required
def api_subproductos():
    """Devuelve lista de subproductos {codigo, nombre} para autocomplete."""
    q = request.args.get("q", "").strip().lower()
    items = [{"codigo": k, "nombre": v} for k, v in SUBPRODUCTOS.items()]
    if q:
        items = [i for i in items if q in i["codigo"].lower() or q in i["nombre"].lower()]
    items.sort(key=lambda x: (not x["codigo"].lower().startswith(q), x["codigo"]))
    return jsonify(items[:30])


@app.route("/api/colores_disponibles/<material>")
@login_required
def api_colores_disponibles(material: str):
    """Retorna solo los colores DISPONIBLE (tienen ZPLA, no existen aún) para un ZFER.
    Usado para mostrar inline en la sección de fórmulas del panel simétrico."""
    material = material.strip()
    attrs = q_atributos(material)
    if "_error" in attrs:
        return jsonify({"ok": False, "colores": []})
    formula_code  = attrs.get("Z_FORMULA_CODE", "")
    piece_type    = attrs.get("Z_PIECE_TYPE",   "")
    shade_band    = attrs.get("Z_SHADE_BAND",   "00") or "00"
    differentials = attrs.get("Z_BEHAVIOR_DIFFERENTIALS", "")
    partnumber    = attrs.get("Z_AGP_PARTNUMBER", "")
    pn_parsed     = _parsear_partnumber(partnumber)
    shade_band_val  = attrs.get("Z_SHADE_BAND", "00") or "00"
    tiene_acero_mat = "06" in {d.strip() for d in differentials.split(",") if d.strip()}

    with ThreadPoolExecutor(max_workers=2) as ex:
        fut_variantes = ex.submit(
            q_variantes_por_pn,
            pn_parsed["vehiculo"], pn_parsed["version"],
            pn_parsed["formula"],  pn_parsed["pieza"]
        ) if pn_parsed else None
        fut_zplas = ex.submit(q_zplas_compatibles, formula_code, piece_type, shade_band_val, differentials, tiene_acero_mat)

    variantes = fut_variantes.result() if fut_variantes else []
    zplas     = fut_zplas.result()
    if variantes and "_error" in variantes[0]: variantes = []
    if zplas     and "_error" in zplas[0]:     zplas     = []

    colores_con_zfer = {v["color_raw"]: v for v in variantes if v.get("color_raw")}
    colores_con_zpla: dict = {}
    for z in zplas:
        colores_con_zpla.setdefault(z["color"], []).append(z)

    disponibles = []
    for cod in _COLORES_ACTIVOS:
        if cod in colores_con_zfer:
            continue   # ya existe
        zpla_list = colores_con_zpla.get(cod, [])
        if not zpla_list:
            continue   # sin zpla
        nombre = COLORES.get(cod, cod)
        disponibles.append({
            "color_codigo": cod,
            "color_nombre": nombre,
            "zpla":         zpla_list[0]["material"],
            "zpla_list":    [z["material"] for z in zpla_list],
            "zpla_count":   len(zpla_list),
            "differentials": zpla_list[0].get("differentials", ""),
        })
    disponibles.sort(key=lambda x: x["color_codigo"])
    return jsonify({"ok": True, "colores": disponibles,
                    "franja": shade_band, "partnumber": partnumber,
                    "nivel": attrs.get("Z_AGP_LEVEL",""),
                    "piece_type": piece_type})


@app.route("/api/bloquear", methods=["POST"])
@login_required
def api_bloquear():
    try:
        from db_local import bloquear
        data = request.get_json(force=True)
        ok = bloquear(
            zfer         = data.get("zfer", ""),
            color_codigo = data.get("color", ""),
            formula      = data.get("formula", ""),
            tipo_pieza   = data.get("tipo_pieza", ""),
            acero_variante = data.get("acero", ""),
            motivo       = data.get("motivo", "Sin motivo"),
            bloqueado_por = data.get("usuario", "web"),
        )
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/desbloquear", methods=["POST"])
@login_required
def api_desbloquear():
    try:
        from db_local import desbloquear
        data = request.get_json(force=True)
        ok = desbloquear(data.get("zfer", ""), data.get("color", ""))
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bloqueos/<material>")
def api_bloqueos(material: str):
    try:
        from db_local import bloqueos_para_zfer
        return jsonify(bloqueos_para_zfer(material.strip()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API SAP ────────────────────────────────────────────────────────────────────

@app.route("/api/sap/ejecutar", methods=["POST"])
@login_required
def api_sap_ejecutar():
    """Lanza automatización SAP para múltiples combinaciones, secuencialmente en un hilo."""
    try:
        data = request.get_json(force=True)
        import uuid as _uuid

        # Acepta array "combinaciones" o parámetros individuales (compatibilidad)
        combis_raw = data.get("combinaciones")
        if not combis_raw:
            zfer_base = data.get("zfer", "").strip()
            color_cod = data.get("color", "").strip()
            if not zfer_base or not color_cod:
                return jsonify({"ok": False, "error": "Faltan parámetros"}), 400
            combis_raw = [{
                "zfer":        zfer_base,
                "color":       color_cod,
                "color_nombre": data.get("color_nombre", color_cod),
                "franja":      data.get("franja", "00") or "00",
                "pn_base":     data.get("pn_base", ""),
                "zpla":        data.get("zpla", ""),
            }]

        import datetime as _dt
        batch_id = str(_uuid.uuid4())[:8]
        _sap_jobs[batch_id] = {
            "estado":       "EN_PROCESO",
            "total":        len(combis_raw),
            "procesados":   0,
            "items":        [],
            "zfer_base":    combis_raw[0].get("zfer", "") if combis_raw else "",
            "pn_base":      combis_raw[0].get("pn_base", "") if combis_raw else "",
            "franja":       combis_raw[0].get("franja", "") if combis_raw else "",
            "fecha_inicio": _dt.datetime.now().isoformat(timespec="seconds"),
            "fecha_fin":    None,
            "usuario_sap":  _usuario_actual() or "PROGRAING",
        }

        def _run_batch():
            from sap_auto import procesar_combinacion as _pc
            import datetime as _dt2
            for c in combis_raw:
                zfer  = c.get("zfer", "").strip()
                color = c.get("color", "").strip()
                if not zfer or not color:
                    continue
                t0  = _dt2.datetime.now()

                def _step_cb(paso_num, desc, _color=color):
                    _sap_jobs[batch_id]["_current"] = {
                        "color": _color, "paso": paso_num, "desc": desc
                    }

                res = _pc(
                    zfer, color,
                    c.get("color_nombre", color),
                    c.get("franja", "00") or "00",
                    c.get("pn_base", ""),
                    c.get("zpla", ""),
                    nivel      = c.get("nivel", ""),
                    tipo_pieza = c.get("tipo_pieza", ""),
                    step_callback=_step_cb,
                )
                job = _sap_jobs[batch_id]
                job["items"].append({
                    "color":        color,
                    "color_nombre": c.get("color_nombre", color),
                    "pn_base":      c.get("pn_base", ""),
                    "zpla_entrada": c.get("zpla", ""),
                    "estado":       res.estado,
                    "zfer_nuevo":   res.zfer_nuevo,
                    "zfor_nuevo":   res.zfor_nuevo,
                    "zpla":         res.zpla,
                    "posiciones":   res.posiciones_bom,
                    "error":        res.error,
                    "duracion_seg": res.duracion_seg,
                    "fecha_inicio": t0.isoformat(timespec="seconds"),
                    "fecha_fin":    _dt2.datetime.now().isoformat(timespec="seconds"),
                    "log":          res.log,
                })
                job["procesados"] += 1
            _sap_jobs[batch_id]["estado"]    = "COMPLETADO"
            _sap_jobs[batch_id]["fecha_fin"] = _dt2.datetime.now().isoformat(timespec="seconds")
            _sap_jobs[batch_id]["_current"]  = None

        threading.Thread(target=_run_batch, daemon=True).start()
        return jsonify({"ok": True, "batch_id": batch_id, "total": len(combis_raw)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/sap/ejecutar_formula_sin_acero", methods=["POST"])
@login_required
def api_sap_ejecutar_formula_sin_acero():
    """Lanza automatización SAP para cambio de fórmula (con acero → sin acero)."""
    try:
        data = request.get_json(force=True)
        import uuid as _uuid, datetime as _dt

        combis_raw = data.get("combinaciones", [])
        if not combis_raw:
            return jsonify({"ok": False, "error": "Faltan combinaciones"}), 400

        batch_id = str(_uuid.uuid4())[:8]
        _sap_jobs[batch_id] = {
            "estado":       "EN_PROCESO",
            "total":        len(combis_raw),
            "procesados":   0,
            "items":        [],
            "zfer_base":    combis_raw[0].get("zfer", "") if combis_raw else "",
            "pn_base":      combis_raw[0].get("pn_base", "") if combis_raw else "",
            "franja":       combis_raw[0].get("franja", "") if combis_raw else "",
            "fecha_inicio": _dt.datetime.now().isoformat(timespec="seconds"),
            "fecha_fin":    None,
            "tipo":         "formula_sin_acero",
            "usuario_sap":  _usuario_actual() or "PROGRAING",
        }

        def _run_formula():
            from sap_auto import procesar_combinacion_formula_sin_acero as _pf
            import datetime as _dt2
            for c in combis_raw:
                zfer    = c.get("zfer", "").strip()
                formula = c.get("formula_nueva", "").strip()
                color   = c.get("color", "").strip()
                if not zfer or not formula:
                    continue
                t0 = _dt2.datetime.now()

                def _step_cb(paso_num, desc, _f=formula):
                    _sap_jobs[batch_id]["_current"] = {
                        "formula": _f, "paso": paso_num, "desc": desc
                    }

                res = _pf(
                    zfer, formula, color,
                    c.get("color_nombre", color),
                    c.get("franja", "00") or "00",
                    c.get("pn_base", ""),
                    c.get("zpla", ""),
                    nivel      = c.get("nivel", ""),
                    tipo_pieza = c.get("tipo_pieza", ""),
                    step_callback=_step_cb,
                )
                job = _sap_jobs[batch_id]
                job["items"].append({
                    "formula_nueva": formula,
                    "color":         color,
                    "color_nombre":  c.get("color_nombre", color),
                    "pn_base":       c.get("pn_base", ""),
                    "zpla_entrada":  c.get("zpla", ""),
                    "estado":        res.estado,
                    "zfer_nuevo":    res.zfer_nuevo,
                    "zfor_nuevo":    res.zfor_nuevo,
                    "zpla":          res.zpla,
                    "posiciones":    res.posiciones_bom,
                    "error":         res.error,
                    "duracion_seg":  res.duracion_seg,
                    "fecha_inicio":  t0.isoformat(timespec="seconds"),
                    "fecha_fin":     _dt2.datetime.now().isoformat(timespec="seconds"),
                    "log":           res.log,
                })
                job["procesados"] += 1
            _sap_jobs[batch_id]["estado"]    = "COMPLETADO"
            _sap_jobs[batch_id]["fecha_fin"] = _dt2.datetime.now().isoformat(timespec="seconds")
            _sap_jobs[batch_id]["_current"]  = None

        threading.Thread(target=_run_formula, daemon=True).start()
        return jsonify({"ok": True, "batch_id": batch_id, "total": len(combis_raw)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/sap/verificar_duplicados", methods=["POST"])
@login_required
def api_sap_verificar_duplicados():
    """
    Recibe lista de items {tipo, zfer, color, formula_nueva?} y devuelve
    cuáles ya existen en SAP (mismo vehicle_code + formula + tipo_pieza + color).
    """
    try:
        data  = request.get_json(force=True)
        zfer_base = data.get("zfer_base", "").strip()
        items     = data.get("items", [])
        if not zfer_base or not items:
            return jsonify({"ok": True, "results": []})

        with get_conn() as cn:
            cur = cn.cursor()

            # 1. Obtener atributos del ZFER base (vehicle_code, version, formula, tipo_pieza)
            cur.execute("""
                SELECT MAX(CASE WHEN ATNAM='Z_VEHICLE_CODE' THEN ATWRT END) AS VC,
                       MAX(CASE WHEN ATNAM='Z_AGP_VERSION'  THEN ATWRT END) AS VERSION,
                       MAX(CASE WHEN ATNAM='Z_FORMULA_CODE' THEN ATWRT END) AS FORMULA,
                       MAX(CASE WHEN ATNAM='Z_PIECE_TYPE'   THEN ATWRT END) AS TIPO_PIEZA
                FROM dbo.ODATA_ZFER_CLASS_001
                WHERE MATERIAL=?
                  AND ATNAM IN ('Z_VEHICLE_CODE','Z_AGP_VERSION','Z_FORMULA_CODE','Z_PIECE_TYPE')
            """, zfer_base)
            row = cur.fetchone()
            if not row or not row[0]:
                return jsonify({"ok": True, "results": [], "warn": "ZFER base sin atributos en BD"})

            vehicle_code = str(row[0] or "").strip()
            version      = str(row[1] or "").strip()
            formula_base = str(row[2] or "").strip()
            tipo_pieza   = str(row[3] or "").strip()

            # 2. Batch: construir pares únicos (formula, color) para un solo query
            pares = []
            item_keys = []  # (tipo, color, formula) por item para lookup
            for it in items:
                tipo    = it.get("tipo", "color")
                color   = str(it.get("color", "")).strip()
                formula = str(it.get("formula_nueva", formula_base)).strip() \
                          if tipo == "formula" else formula_base
                item_keys.append((tipo, color, formula))
                pares.append((formula, color))

            # Query única: trae todos los ZFERs activos de este vehicle+version+tipo_pieza
            # filtrando solo las combinaciones (formula, color) que nos interesan.
            # Usamos OR de pares para evitar N queries.
            formulas_unicas = list({p[0] for p in pares})
            colores_unicos  = list({p[1] for p in pares})

            placeholders_f = ",".join("?" * len(formulas_unicas))
            placeholders_c = ",".join("?" * len(colores_unicos))

            # Pre-filtro con semi-joins: solo agrupa materiales que comparten
            # vehicle_code + version + tipo_pieza, reduciendo el escaneo masivamente.
            _SQL_BATCH = f"""
                SELECT b.MATERIAL, b.FORMULA, b.COLOR
                FROM (
                    SELECT MATERIAL,
                           MAX(CASE WHEN ATNAM='Z_FORMULA_CODE' THEN ATWRT END) AS FORMULA,
                           MAX(CASE WHEN ATNAM='Z_COLOR'        THEN ATWRT END) AS COLOR
                    FROM dbo.ODATA_ZFER_CLASS_001
                    WHERE MATERIAL <> ?
                      AND ATNAM IN ('Z_FORMULA_CODE','Z_COLOR')
                      AND MATERIAL IN (
                          SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001
                          WHERE ATNAM='Z_VEHICLE_CODE' AND ATWRT=?
                      )
                      AND MATERIAL IN (
                          SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001
                          WHERE ATNAM='Z_AGP_VERSION' AND ATWRT=?
                      )
                      AND MATERIAL IN (
                          SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001
                          WHERE ATNAM='Z_PIECE_TYPE' AND ATWRT=?
                      )
                    GROUP BY MATERIAL
                ) b
                JOIN dbo.ODATA_ZFER_HEAD h ON h.MATERIAL = b.MATERIAL
                WHERE h.STATUS IS NULL AND h.CENTRO = 'CO01'
                  AND b.FORMULA IN ({placeholders_f})
                  AND b.COLOR   IN ({placeholders_c})
            """
            cur.execute(_SQL_BATCH,
                        zfer_base, vehicle_code, version, tipo_pieza,
                        *formulas_unicas, *colores_unicos)

            # Indexar resultados: (formula, color) → [material, ...]
            found: dict = {}
            for r in cur.fetchall():
                key = (str(r[1] or "").strip(), str(r[2] or "").strip())
                found.setdefault(key, []).append(str(r[0]))

            # Cruzar con cada item original
            results = []
            for it, (tipo, color, formula) in zip(items, item_keys):
                existing = found.get((formula, color), [])
                results.append({
                    **it,
                    "ya_existe":        len(existing) > 0,
                    "zfer_existente":   existing[0] if existing else None,
                    "todos_existentes": existing,
                })

        return jsonify({"ok": True, "results": results,
                        "vehicle_code": vehicle_code,
                        "version":      version,
                        "formula_base": formula_base,
                        "tipo_pieza":   tipo_pieza})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/sap/ejecutar_cola", methods=["POST"])
@login_required
def api_sap_ejecutar_cola():
    """Cola unificada: mezcla items de tipo 'color' y 'formula' en secuencia."""
    try:
        data = request.get_json(force=True)
        import uuid as _uuid, datetime as _dt

        cola = data.get("cola", [])
        if not cola:
            return jsonify({"ok": False, "error": "Cola vacía"}), 400

        batch_id = str(_uuid.uuid4())[:8]
        _sap_jobs[batch_id] = {
            "estado":       "EN_PROCESO",
            "total":        len(cola),
            "procesados":   0,
            "items":        [],
            "zfer_base":    cola[0].get("zfer", "") if cola else "",
            "pn_base":      cola[0].get("pn_base", "") if cola else "",
            "franja":       cola[0].get("franja", "") if cola else "",
            "fecha_inicio": _dt.datetime.now().isoformat(timespec="seconds"),
            "fecha_fin":    None,
            "tipo":         "mixto",
            "usuario_sap":  _usuario_actual() or "PROGRAING",
        }

        def _run_cola():
            from sap_auto import (procesar_combinacion as _pc,
                                  procesar_combinacion_formula_sin_acero as _pf)
            import datetime as _dt2
            for c in cola:
                tipo    = c.get("tipo", "color")
                zfer    = c.get("zfer", "").strip()
                color   = c.get("color", "").strip()
                formula = c.get("formula_nueva", "").strip()
                t0      = _dt2.datetime.now()

                def _step_cb(paso_num, desc, _tipo=tipo, _color=color, _formula=formula):
                    _sap_jobs[batch_id]["_current"] = {
                        "tipo": _tipo, "color": _color, "formula": _formula,
                        "paso": paso_num, "desc": desc,
                    }

                if tipo == "formula":
                    res = _pf(
                        zfer, formula, color,
                        c.get("color_nombre", color),
                        c.get("franja", "00") or "00",
                        c.get("pn_base", ""),
                        c.get("zpla", ""),
                        nivel      = c.get("nivel", ""),
                        tipo_pieza = c.get("tipo_pieza", ""),
                        step_callback=_step_cb,
                    )
                else:
                    res = _pc(
                        zfer, color,
                        c.get("color_nombre", color),
                        c.get("franja", "00") or "00",
                        c.get("pn_base", ""),
                        c.get("zpla", ""),
                        nivel      = c.get("nivel", ""),
                        tipo_pieza = c.get("tipo_pieza", ""),
                        step_callback=_step_cb,
                    )

                job = _sap_jobs[batch_id]
                job["items"].append({
                    "tipo":          tipo,
                    "color":         color,
                    "color_nombre":  c.get("color_nombre", color),
                    "formula_nueva": formula,
                    "pn_base":       c.get("pn_base", ""),
                    "zpla_entrada":  c.get("zpla", ""),
                    "estado":        res.estado,
                    "zfer_nuevo":    res.zfer_nuevo,
                    "zfor_nuevo":    res.zfor_nuevo,
                    "zpla":          res.zpla,
                    "posiciones":    res.posiciones_bom,
                    "error":         res.error,
                    "duracion_seg":  res.duracion_seg,
                    "fecha_inicio":  t0.isoformat(timespec="seconds"),
                    "fecha_fin":     _dt2.datetime.now().isoformat(timespec="seconds"),
                    "log":           res.log,
                })
                job["procesados"] += 1
            _sap_jobs[batch_id]["estado"]    = "COMPLETADO"
            _sap_jobs[batch_id]["fecha_fin"] = _dt2.datetime.now().isoformat(timespec="seconds")
            _sap_jobs[batch_id]["_current"]  = None

        threading.Thread(target=_run_cola, daemon=True).start()
        return jsonify({"ok": True, "batch_id": batch_id, "total": len(cola)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/sap/estado/<batch_id>")
@login_required
def api_sap_estado(batch_id: str):
    """Consulta el estado de un batch SAP."""
    job = _sap_jobs.get(batch_id)
    if not job:
        return jsonify({"error": "batch_id no encontrado"}), 404
    return jsonify({
        "batch_id":     batch_id,
        "estado":       job["estado"],
        "total":        job["total"],
        "procesados":   job["procesados"],
        "items":        job["items"],
        "zfer_base":    job.get("zfer_base", ""),
        "pn_base":      job.get("pn_base", ""),
        "franja":       job.get("franja", ""),
        "fecha_inicio": job.get("fecha_inicio", ""),
        "fecha_fin":    job.get("fecha_fin", ""),
        "usuario_sap":  job.get("usuario_sap", "PROGRAING"),
        "_current":     job.get("_current"),
        "tipo_batch":   job.get("tipo", "color"),
        "log":          [],
    })


@app.route("/api/sap/reporte/<batch_id>")
@login_required
def api_sap_reporte(batch_id: str):
    """Genera y descarga reporte Excel del batch SAP."""
    job = _sap_jobs.get(batch_id)
    if not job:
        return "Batch no encontrado", 404
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io, datetime as _dt

        wb = openpyxl.Workbook()

        # ── Estilos ──────────────────────────────────────────────────────────
        H_FILL   = PatternFill("solid", fgColor="0D1117")
        OK_FILL  = PatternFill("solid", fgColor="1A3A2A")
        ERR_FILL = PatternFill("solid", fgColor="3A1A1A")
        HDR_FILL = PatternFill("solid", fgColor="161B22")
        TH_FILL  = PatternFill("solid", fgColor="21262D")
        thin     = Side(style="thin", color="30363D")
        brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
        fnt_h    = Font(name="Calibri", bold=True, color="E6EDF3", size=11)
        fnt_ok   = Font(name="Calibri", bold=True, color="56D364", size=10)
        fnt_err  = Font(name="Calibri", bold=True, color="FFA198", size=10)
        fnt_muted= Font(name="Calibri", color="8B949E", size=10)
        fnt_norm = Font(name="Calibri", color="E6EDF3", size=10)
        cnt      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lft      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        items  = job.get("items", [])
        n_ok   = sum(1 for i in items if i.get("estado") == "OK")
        n_err  = sum(1 for i in items if i.get("estado") == "ERROR")
        t_segs = sum(i.get("duracion_seg", 0) for i in items)

        def set_col_widths(ws, widths):
            for col, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = w

        def add_header_row(ws, cols, row=1):
            for c, title in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=title)
                cell.font      = fnt_h
                cell.fill      = TH_FILL
                cell.border    = brd
                cell.alignment = cnt

        # ════════════════════════════════════════════════════════════════════
        # HOJA 1 — RESUMEN
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "RESUMEN"
        ws1.sheet_view.showGridLines = False
        ws1.row_dimensions[1].height = 40

        # Título
        ws1.merge_cells("A1:F1")
        t = ws1["A1"]
        t.value     = "🔷  REPORTE DE AUTOMATIZACIÓN SAP — AGP GLASS"
        t.font      = Font(name="Calibri", bold=True, color="58A6FF", size=16)
        t.fill      = H_FILL
        t.alignment = Alignment(horizontal="center", vertical="center")

        # Info batch
        tipo_batch = job.get("tipo", "color")
        n_color    = sum(1 for i in items if (i.get("tipo") or "color") == "color")
        n_formula  = sum(1 for i in items if i.get("tipo") == "formula")
        tipo_lbl   = "Mixto (Color + Fórmula)" if tipo_batch == "mixto" else \
                     ("Fórmula sin acero" if tipo_batch == "formula_sin_acero" else "Cambio de Color")
        meta = [
            ("Batch ID",        batch_id),
            ("Tipo de proceso",  tipo_lbl),
            ("ZFER Base",       job.get("zfer_base", "—")),
            ("PN Base",         job.get("pn_base", "—")),
            ("Franja",          job.get("franja", "—")),
            ("Items Color",     n_color),
            ("Items Fórmula",   n_formula),
            ("Usuario SAP",     job.get("usuario_sap", "PROGRAING")),
            ("Fecha Inicio",    job.get("fecha_inicio", "—")),
            ("Fecha Fin",       job.get("fecha_fin", "—")),
            ("Duración total",  f"{round(t_segs/60, 1)} min ({int(t_segs)}s)"),
        ]
        for r, (k, v) in enumerate(meta, 3):
            lbl  = ws1.cell(row=r, column=1, value=k)
            val  = ws1.cell(row=r, column=2, value=str(v))
            lbl.font = fnt_h;  lbl.fill = TH_FILL; lbl.border = brd; lbl.alignment = lft
            val.font = fnt_norm; val.fill = HDR_FILL; val.border = brd; val.alignment = lft

        # KPIs
        kpis = [
            ("TOTAL",    len(items), "E6EDF3"),
            ("✓ OK",     n_ok,       "56D364"),
            ("✗ ERRORES",n_err,      "FFA198"),
            ("TIEMPO",   f"{round(t_segs/60,1)} min", "E3B341"),
        ]
        for c, (label, val, color) in enumerate(kpis, 1):
            ws1.row_dimensions[12].height = 50
            ws1.row_dimensions[13].height = 30
            lc = ws1.cell(row=12, column=c, value=label)
            lc.font = Font(name="Calibri", bold=True, color=color, size=12)
            lc.fill = TH_FILL; lc.border = brd; lc.alignment = cnt
            vc = ws1.cell(row=13, column=c, value=val)
            vc.font = Font(name="Calibri", bold=True, color=color, size=20)
            vc.fill = HDR_FILL; vc.border = brd; vc.alignment = cnt

        set_col_widths(ws1, [22, 35, 18, 18])

        # ════════════════════════════════════════════════════════════════════
        # HOJA 2 — DETALLE
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("DETALLE")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:I1")
        t2 = ws2["A1"]
        t2.value = "DETALLE POR COMBINACIÓN"
        t2.font  = Font(name="Calibri", bold=True, color="58A6FF", size=13)
        t2.fill  = H_FILL; t2.alignment = cnt
        ws2.row_dimensions[1].height = 28

        tiene_formula = any(i.get("tipo") == "formula" for i in items)
        cols2 = ["#", "Tipo", "Color Código", "Color Nombre", "Fórmula Nueva", "Estado",
                 "ZFER Nuevo", "ZFOR Nuevo", "ZPLA Usado",
                 "Posiciones BOM", "Duración (s)"]
        add_header_row(ws2, cols2, row=2)

        for i, item in enumerate(items, 1):
            r    = i + 2
            es   = item.get("estado", "")
            tipo = item.get("tipo", "color")
            fill = OK_FILL if es == "OK" else ERR_FILL if es == "ERROR" else HDR_FILL
            vals = [
                i,
                "🎨 Color" if tipo == "color" else "🧪 Fórmula",
                item.get("color", ""),
                item.get("color_nombre", ""),
                item.get("formula_nueva", "—") if tipo == "formula" else "—",
                es,
                item.get("zfer_nuevo", ""),
                item.get("zfor_nuevo", ""),
                item.get("zpla", ""),
                ", ".join(
                    p["pos"] if isinstance(p, dict) else str(p)
                    for p in item.get("posiciones", [])
                ),
                item.get("duracion_seg", 0),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.fill   = fill
                cell.border = brd
                cell.alignment = cnt if c in (1, 6, 11) else lft
                if c == 6:
                    cell.font = fnt_ok if es == "OK" else fnt_err
                elif c == 2:
                    cell.font = Font(name="Calibri",
                                     color="79C0FF" if tipo == "color" else "BC8CFF",
                                     bold=True, size=10)
                else:
                    cell.font = fnt_norm

        set_col_widths(ws2, [5, 14, 14, 30, 16, 12, 18, 18, 18, 28, 14])

        # ════════════════════════════════════════════════════════════════════
        # HOJA 3b — DETALLE FÓRMULA (solo si hay items de fórmula)
        # ════════════════════════════════════════════════════════════════════
        items_formula = [it for it in items if it.get("tipo") == "formula"]
        if items_formula:
            wsf = wb.create_sheet("DETALLE_FÓRMULA")
            wsf.sheet_view.showGridLines = False
            wsf.merge_cells("A1:J1")
            tf = wsf["A1"]
            tf.value = "DETALLE — CAMBIOS DE FÓRMULA"
            tf.font  = Font(name="Calibri", bold=True, color="BC8CFF", size=13)
            tf.fill  = H_FILL; tf.alignment = cnt
            wsf.row_dimensions[1].height = 28

            cols_f = ["#", "Fórmula Nueva", "Color Código", "Color Nombre", "Estado",
                      "ZFER Nuevo", "ZFOR Nuevo", "ZPLA Usado", "Posiciones BOM", "Duración (s)"]
            add_header_row(wsf, cols_f, row=2)
            
            for i, item in enumerate(items_formula, 1):
                r  = i + 2
                es = item.get("estado", "")
                fill = OK_FILL if es == "OK" else ERR_FILL if es == "ERROR" else HDR_FILL
                vals = [
                    i,
                    item.get("formula_nueva", ""),
                    item.get("color", ""),
                    item.get("color_nombre", ""),
                    es,
                    item.get("zfer_nuevo", ""),
                    item.get("zfor_nuevo", ""),
                    item.get("zpla", ""),
                    ", ".join(
                        p["pos"] if isinstance(p, dict) else str(p)
                        for p in item.get("posiciones", [])
                    ),
                    item.get("duracion_seg", 0),
                ]
                for c, v in enumerate(vals, 1):
                    cell = wsf.cell(row=r, column=c, value=v)
                    cell.fill   = fill
                    cell.border = brd
                    cell.alignment = cnt if c in (1, 5, 10) else lft
                    cell.font = (fnt_ok if es == "OK" else fnt_err) if c == 5 else fnt_norm
            set_col_widths(wsf, [5, 16, 14, 30, 12, 18, 18, 18, 28, 14])

        # ════════════════════════════════════════════════════════════════════
        # HOJA 3 — ERRORES (solo si hay)
        # ════════════════════════════════════════════════════════════════════
        if n_err > 0:
            ws3 = wb.create_sheet("ERRORES")
            ws3.sheet_view.showGridLines = False
            ws3.merge_cells("A1:D1")
            t3 = ws3["A1"]
            t3.value = f"ERRORES DETALLADOS — {n_err} item(s)"
            t3.font  = Font(name="Calibri", bold=True, color="FFA198", size=13)
            t3.fill  = H_FILL; t3.alignment = cnt
            ws3.row_dimensions[1].height = 28

            add_header_row(ws3, ["Tipo", "Color / Fórmula", "Nombre", "Error", "Log completo"], row=2)
            set_col_widths(ws3, [12, 18, 30, 60, 80])

            fila_err = 3
            for item in items:
                if item.get("estado") != "ERROR":
                    continue
                log_txt  = "\n".join(item.get("log", []))
                tipo_lbl = "Fórmula" if item.get("tipo") == "formula" else "Color"
                id_lbl   = (item.get("formula_nueva") or item.get("color", "")) if item.get("tipo") == "formula" else item.get("color", "")
                for c, v in enumerate([
                    tipo_lbl,
                    id_lbl,
                    item.get("color_nombre", ""),
                    item.get("error", ""),
                    log_txt,
                ], 1):
                    cell = ws3.cell(row=fila_err, column=c, value=v)
                    cell.fill   = ERR_FILL
                    cell.border = brd
                    cell.font   = fnt_err if c == 3 else fnt_norm
                    cell.alignment = Alignment(horizontal="left", vertical="top",
                                               wrap_text=True)
                ws3.row_dimensions[fila_err].height = max(
                    60, len(log_txt.split("\n")) * 14)
                fila_err += 1

        # ════════════════════════════════════════════════════════════════════
        # HOJA 4 — LOG COMPLETO
        # ════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("LOG")
        ws4.sheet_view.showGridLines = False
        ws4.merge_cells("A1:C1")
        t4 = ws4["A1"]
        t4.value = "LOG COMPLETO DE EJECUCIÓN"
        t4.font  = Font(name="Calibri", bold=True, color="8B949E", size=13)
        t4.fill  = H_FILL; t4.alignment = cnt
        ws4.row_dimensions[1].height = 28
        add_header_row(ws4, ["Tipo", "Color / Fórmula", "Estado", "Línea de log"], row=2)
        set_col_widths(ws4, [10, 18, 12, 120])

        fila_log = 3
        for item in items:
            es       = item.get("estado", "")
            fill     = OK_FILL if es == "OK" else ERR_FILL if es == "ERROR" else HDR_FILL
            tipo_lbl = "Fórmula" if item.get("tipo") == "formula" else "Color"
            id_lbl   = (item.get("formula_nueva") or item.get("color", "")) if item.get("tipo") == "formula" else item.get("color", "")
            for linea in item.get("log", []):
                for c, v in enumerate([tipo_lbl, id_lbl, es, linea], 1):
                    cell = ws4.cell(row=fila_log, column=c, value=v)
                    cell.fill   = fill
                    cell.border = brd
                    cell.font   = fnt_muted
                    cell.alignment = lft
                fila_log += 1

        # ── Enviar como descarga ──────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fecha_str = _dt.datetime.now().strftime("%Y%m%d_%H%M")
        filename  = f"SAP_Reporte_{batch_id}_{fecha_str}.xlsx"

        from flask import send_file
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return f"Error generando reporte: {e}", 500


@app.route("/api/vehiculo_lookup")
@login_required
def api_vehiculo_lookup():
    """Dado un cod_vehiculo (prefijo del PARTNUMBER), devuelve el modelo de vehículo."""
    cod = request.args.get("cod", "").strip()
    if not cod or len(cod) < 2:
        return jsonify({"vehiculo": ""})
    def _esc(s):
        return s.replace("!", "!!").replace("%", "!%").replace("_", "!_")
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT TOP 1 c2.ATWRT
            FROM   dbo.ODATA_ZFER_CLASS_001 c1
            JOIN   dbo.ODATA_ZFER_CLASS_001 c2
                   ON  c2.MATERIAL = c1.MATERIAL
                   AND c2.CENTRO   = 'CO01'
                   AND c2.ATNAM    = 'Z_VEHICLE_MODEL'
            WHERE  c1.CENTRO = 'CO01'
              AND  c1.ATNAM  = 'Z_AGP_PARTNUMBER'
              AND  c1.ATWRT  LIKE ? ESCAPE '!'
        """, [f"{_esc(cod)}!_%"])
        row = cur.fetchone()
        conn.close()
        return jsonify({"vehiculo": str(row[0]).strip() if row and row[0] else ""})
    except Exception as e:
        return jsonify({"vehiculo": "", "error": str(e)}), 200


# ── BD AGP Ingeniería (migrada desde localhost\SQLEXPRESS\MODULO_5) ──────────
_DB_LOCAL_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=agpcolombia.database.windows.net;"
    "DATABASE=AGP_Ingenieria;"
    "UID=DevIngenieria;"
    "PWD=HiJE068i0LQVrwA;"
    "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30;"
)

def _get_conn_local():
    return pyodbc.connect(_DB_LOCAL_STR, autocommit=True)


def _guardar_homologacion_formula(item: dict, res, session_user: str = "sistema") -> None:
    """
    Inserta en M5_HomologacionFormula + M5_HomologacionFormula_BOM
    después de una homologación de fórmula exitosa.
    Se llama en el worker SÓLO cuando tipo in FORMULA* y estado == OK.
    No lanza excepción — cualquier error se imprime pero no interrumpe el flujo.
    """
    try:
        zfer_nuevo  = getattr(res, "zfer_nuevo",  "") or ""
        zfor_nuevo  = getattr(res, "zfor_nuevo",  "") or ""
        zpla        = getattr(res, "zpla",         "") or ""
        bom_detalle = getattr(res, "bom_detalle",  [])

        # ── Atributos del ZFER nuevo desde Azure (cacheado) ────────────
        attrs_nuevo = {}
        if zfer_nuevo:
            try:
                attrs_nuevo = q_atributos(zfer_nuevo)
            except Exception:
                pass

        vehiculo_nombre  = attrs_nuevo.get("Z_VEHICLE_MODEL", "")
        version_vehiculo = attrs_nuevo.get("Z_AGP_VERSION",   "")
        pieza            = attrs_nuevo.get("Z_PIECE_TYPE",     "")
        pn_nuevo         = attrs_nuevo.get("Z_AGP_PARTNUMBER","")
        vehiculo_codigo  = (pn_nuevo.split("_")[0] if pn_nuevo and "_" in pn_nuevo else "")

        # ── Ruta y simetría del ZFER base desde M5_RutasZFER ───────────
        ruta_3dm       = None
        tiene_simetria = 0
        zfer_simetrico = None
        try:
            with _get_conn_local() as cn:
                row = cn.cursor().execute(
                    "SELECT ruta, tiene_simetria, zfer_simetrico "
                    "FROM itg.M5_RUTASZFER WHERE zfer = ?",
                    item.get("zfer", "")
                ).fetchone()
                if row:
                    ruta_3dm       = str(row[0] or "") or None
                    tiene_simetria = 1 if row[1] else 0
                    zfer_simetrico = str(row[2] or "") or None
        except Exception:
            pass

        # ── INSERT principal ────────────────────────────────────────────
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                INSERT INTO dbo.M5_HomologacionFormula
                    (zfer_base, formula_base, formula_nueva, acero_dir,
                     color_codigo, color_nombre,
                     zfer_nuevo, zfor_nuevo, zpla,
                     vehiculo_nombre, version_vehiculo, vehiculo_codigo, pieza,
                     ruta_3dm, tiene_simetria, zfer_simetrico,
                     creado_por, estado, batch_id)
                OUTPUT INSERTED.id
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
                item.get("zfer",         ""),
                getattr(res, "formula",  "") or item.get("formula_nueva", ""),
                item.get("formula_nueva",""),
                item.get("acero_dir",    ""),
                item.get("color",        ""),
                item.get("color_nombre", ""),
                zfer_nuevo  or None,
                zfor_nuevo  or None,
                zpla        or None,
                vehiculo_nombre  or None,
                version_vehiculo or None,
                vehiculo_codigo  or None,
                pieza            or None,
                ruta_3dm,
                tiene_simetria,
                zfer_simetrico,
                session_user,
                "OK",
                getattr(res, "batch_id", "") or None,
            )
            hom_id = cur.fetchone()[0]

            # ── INSERT BOM detalle ──────────────────────────────────────
            if bom_detalle and hom_id:
                cur.executemany(
                    "INSERT INTO dbo.M5_HomologacionFormula_BOM (homologacion_id, zfer_nuevo, posnr, clase_destino) VALUES (?,?,?,?)",
                    [(hom_id, zfer_nuevo or None, b.get("posnr",""), b.get("clase_destino","") or None) for b in bom_detalle]
                )

        print(f"[HOMOLOG] Guardado id={hom_id} zfer_nuevo={zfer_nuevo} BOM={len(bom_detalle)} pos")
        return hom_id

    except Exception as e:
        print(f"[HOMOLOG] Error guardando homologación: {e}")
        return None


def _guardar_gestor_auto(item: dict, res, hom_id: int) -> None:
    """
    Inserta en jobs_gestor_auto + bom_zfer_gestor_auto después de una
    homologación de fórmula exitosa (cuando ya existe el registro en
    M5_HomologacionFormula con id=hom_id).
    No lanza excepción — cualquier error solo se imprime.
    Solo inserta si ruta_3dm no está vacía (constraint NOT NULL).
    """
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()

            # Leer el registro recién guardado en M5_HomologacionFormula
            row = cur.execute("""
                SELECT id, vehiculo_nombre, version_vehiculo, vehiculo_codigo,
                       pieza, tiene_simetria, zfer_simetrico,
                       zfer_base, zfor_nuevo, zpla, ruta_3dm, zfer_nuevo
                FROM dbo.M5_HomologacionFormula
                WHERE id = ?
            """, hom_id).fetchone()

            if not row:
                print(f"[GESTOR] No se encontró M5_HomologacionFormula id={hom_id}")
                return

            (id_origen, veh_nombre, veh_version, veh_codigo,
             pieza, tiene_sim, zfer_sim,
             zfer_base, zfor_nuevo, zpla, ruta_3dm, zfer_nuevo) = row

            # ruta_3dm es NOT NULL en la tabla destino — omitir si falta
            if not ruta_3dm or not str(ruta_3dm).strip():
                print(f"[GESTOR] Omitido id={hom_id}: ruta_3dm vacía")
                return

            # Valores con defaults para campos NOT NULL
            veh_nombre   = (veh_nombre   or "").strip() or "SIN NOMBRE"
            veh_version  = (veh_version  or "").strip() or "SIN VERSION"
            veh_codigo   = (veh_codigo   or "").strip() or "0000"
            pieza_3d     = str(pieza or "").strip().zfill(3)[:3] if pieza else "000"
            simetria_val = "SI" if tiene_sim else "NO"
            zfer_sim_val = str(zfer_sim or "").strip() or None

            # Si ya existe id_origen, actualizar en lugar de insertar
            existe = cur.execute(
                "SELECT 1 FROM itg.M5_JOBSGESTORAUTO WHERE id_origen = ?", id_origen
            ).fetchone()

            if existe:
                cur.execute("""
                    UPDATE itg.M5_JOBSGESTORAUTO SET
                        vehiculo_nombre  = ?, version_vehiculo = ?, vehiculo_codigo = ?,
                        pieza = ?, simetria = ?, zfer_simetria = ?,
                        zfer  = ?, zfor = ?, zpla = ?, ruta_3dm = ?
                    WHERE id_origen = ?
                """,
                    veh_nombre, veh_version, veh_codigo,
                    pieza_3d, simetria_val, zfer_sim_val,
                    str(zfer_nuevo or zfer_base), str(zfor_nuevo or "") or None,
                    str(zpla or "") or None, str(ruta_3dm).strip(),
                    id_origen
                )
                print(f"[GESTOR] Actualizado jobs id_origen={id_origen}")
            else:
                cur.execute("""
                    INSERT INTO itg.M5_JOBSGESTORAUTO
                        (id_origen, vehiculo_nombre, version_vehiculo, vehiculo_codigo,
                         pieza, simetria, zfer_simetria, zfer, zfor, zpla, ruta_3dm)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """,
                    id_origen, veh_nombre, veh_version, veh_codigo,
                    pieza_3d, simetria_val, zfer_sim_val,
                    str(zfer_nuevo or zfer_base), str(zfor_nuevo or "") or None,
                    str(zpla or "") or None, str(ruta_3dm).strip()
                )
                print(f"[GESTOR] Insertado jobs id_origen={id_origen}")

            # ── BOM: borrar anteriores y reinsertar (idempotente) ─────────
            zfer_key = str(zfer_nuevo or zfer_base)
            cur.execute("DELETE FROM itg.M5_BOMGESTORAUTO WHERE zfer = ?", zfer_key)

            bom_detalle = getattr(res, "bom_detalle", []) or []
            if bom_detalle:
                cur.executemany(
                    "INSERT INTO itg.M5_BOMGESTORAUTO (zfer, posicion, clase, descripcion) VALUES (?,?,?,?)",
                    [(zfer_key,
                      str(b.get("posnr", "")).strip(),
                      str(b.get("clase_destino", "")).strip(),
                      None)   # descripcion siempre NULL — no disponible
                     for b in bom_detalle if b.get("posnr")]
                )
                print(f"[GESTOR] BOM {len(bom_detalle)} posiciones para {zfer_key}")

    except Exception as e:
        print(f"[GESTOR] Error guardando gestor_auto: {e}")


def _migracion_bd_local():
    """Aplica migraciones de columnas faltantes en la BD local (idempotente)."""
    migraciones = [
        ("itg.M5_COLA", "cambiar_hr",  "ALTER TABLE itg.M5_COLA ADD cambiar_hr BIT NOT NULL DEFAULT 0"),
        ("itg.M5_COLA", "zhal",        "ALTER TABLE itg.M5_COLA ADD zhal NVARCHAR(20) NULL"),
        ("itg.M5_COLA", "acero_dir",   "ALTER TABLE itg.M5_COLA ADD acero_dir NVARCHAR(10) NULL"),
        ("itg.M5_COLA", "subproducto",   "ALTER TABLE itg.M5_COLA ADD subproducto NVARCHAR(20) NULL"),
        ("itg.M5_COLA", "plano_manual",  "ALTER TABLE itg.M5_COLA ADD plano_manual NVARCHAR(100) NULL"),
    ]
    try:
        cn  = _get_conn_local()
        cur = cn.cursor()
        for tabla, col, sql in migraciones:
            cur.execute(f"""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME='{tabla.split('.')[-1]}' AND COLUMN_NAME='{col}'
                )
                EXEC('{sql}')
            """)
        cn.close()
    except Exception as e:
        print(f"[MIGRACIÓN] error: {e}")

_migracion_bd_local()


@app.route("/api/ruta/<zfer>", methods=["GET"])
@login_required
def api_ruta_get(zfer: str):
    """Devuelve la URL y datos de simetría guardados para un ZFER."""
    zfer = zfer.strip()
    try:
        cn  = _get_conn_local()
        cur = cn.cursor()
        cur.execute(
            "SELECT ruta, descripcion, modificado_el, tiene_simetria, zfer_simetrico, pieza_contraria "
            "FROM itg.M5_RUTASZFER WHERE zfer = ?", zfer
        )
        row = cur.fetchone()
        cn.close()
        if row:
            return jsonify({
                "ok": True, "zfer": zfer,
                "ruta":           str(row[0] or ""),
                "descripcion":    str(row[1] or ""),
                "modificado_el":  str(row[2])[:19] if row[2] else "",
                "tiene_simetria": bool(row[3]),
                "zfer_simetrico": str(row[4] or ""),
                "pieza_contraria":str(row[5] or ""),
            })
        return jsonify({"ok": True, "zfer": zfer, "ruta": "", "descripcion": "",
                        "modificado_el": "", "tiene_simetria": False,
                        "zfer_simetrico": "", "pieza_contraria": ""})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200


@app.route("/api/ruta/<zfer>", methods=["POST"])
@login_required
def api_ruta_set(zfer: str):
    """Guarda (upsert) la URL y simetría para un ZFER."""
    zfer = zfer.strip()
    body = request.get_json(force=True) or {}
    ruta          = str(body.get("ruta", "")).strip()[:500]
    desc          = str(body.get("descripcion", "")).strip()[:200]
    tiene_sim     = 1 if body.get("tiene_simetria") else 0
    zfer_sim      = str(body.get("zfer_simetrico", "")).strip()[:20]
    pieza_contra  = str(body.get("pieza_contraria", "")).strip()[:10]
    try:
        cn  = _get_conn_local()
        cur = cn.cursor()
        cur.execute("""
            MERGE itg.M5_RUTASZFER AS t
            USING (SELECT ? AS zfer, ? AS ruta, ? AS descripcion,
                          ? AS tiene_simetria, ? AS zfer_simetrico, ? AS pieza_contraria) AS s
              ON t.zfer = s.zfer
            WHEN MATCHED THEN
                UPDATE SET ruta=s.ruta, descripcion=s.descripcion,
                           tiene_simetria=s.tiene_simetria, zfer_simetrico=s.zfer_simetrico,
                           pieza_contraria=s.pieza_contraria, modificado_el=GETDATE()
            WHEN NOT MATCHED THEN
                INSERT (zfer, ruta, descripcion, tiene_simetria, zfer_simetrico, pieza_contraria)
                VALUES (s.zfer, s.ruta, s.descripcion, s.tiene_simetria, s.zfer_simetrico, s.pieza_contraria);
        """, zfer, ruta, desc, tiene_sim, zfer_sim, pieza_contra)
        cn.close()
        return jsonify({"ok": True, "zfer": zfer})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200


@app.route("/api/buscar_zhal/<zfer>")
@login_required
def api_buscar_zhal(zfer: str):
    """
    Busca el código ZHAL para el flujo sin acero → con acero.
    Lógica: toma Z_VEHICLE_MODEL + Z_AGP_VERSION + Z_PIECE_TYPE del ZFER base,
    busca otro ZFER con Z_BEHAVIOR_DIFFERENTIALS=06 y esos mismos atributos,
    luego retorna el COMPONENTE (IDNRK) en posición 0106 o 0116 de su BOM.
    """
    zfer = zfer.strip()
    try:
        attrs = q_atributos(zfer)
        vehiculo  = attrs.get("Z_VEHICLE_MODEL", "")
        version   = attrs.get("Z_AGP_VERSION",   "")
        pieza     = attrs.get("Z_PIECE_TYPE",     "")

        if not (vehiculo and version and pieza):
            return jsonify({"ok": True, "zhal": None,
                            "msg": f"Atributos insuficientes (vehiculo={vehiculo} version={version} pieza={pieza})"})

        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT TOP 1 b.IDNRK
            FROM dbo.ODATA_ZFER_BOM b
            WHERE b.CENTRO = 'CO01'
              AND b.POSNR  IN ('0106','0116')
              AND b.MATNR IN (
                SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001
                WHERE CENTRO='CO01' AND ATNAM='Z_VEHICLE_MODEL' AND ATWRT=?
                INTERSECT
                SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001
                WHERE CENTRO='CO01' AND ATNAM='Z_AGP_VERSION' AND ATWRT=?
                INTERSECT
                SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001
                WHERE CENTRO='CO01' AND ATNAM='Z_PIECE_TYPE' AND ATWRT=?
                INTERSECT
                SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001
                WHERE CENTRO='CO01' AND ATNAM='Z_BEHAVIOR_DIFFERENTIALS' AND ATWRT='06'
              )
              AND b.MATNR <> ?
            ORDER BY TRY_CAST(b.MATNR AS BIGINT) DESC
        """, vehiculo, version, pieza, zfer)
        row = cur.fetchone()
        conn.close()

        if row and row[0]:
            return jsonify({"ok": True, "zhal": str(row[0]).strip(), "msg": ""})
        return jsonify({"ok": True, "zhal": None, "msg": "No se encontró ZHAL en BD"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200


@app.route("/api/simetria/<zfer>")
@login_required
def api_simetria_buscar(zfer: str):
    """
    Busca el ZFER simétrico (par LH/RH).
    Estrategia 1 (rápida): LIKE en PARTNUMBER — 1 sola query, usa q_atributos cacheado.
    Estrategia 2 (fallback): INTERSECT con 4 atributos clave si no hay partnumber.
    """
    zfer = zfer.strip()
    try:
        # q_atributos tiene lru_cache — si el usuario ya cargó el ZFER, es gratis (0ms)
        attrs = q_atributos(zfer)
        if not attrs or "_error" in attrs:
            return jsonify({"ok": True, "aplica": False, "encontrados": []})

        piece_code      = attrs.get("Z_PIECE_TYPE", "").split(",")[0].strip().zfill(3)
        pieza_contraria = _PARES_SIMETRIA.get(piece_code)
        if not pieza_contraria:
            return jsonify({"ok": True, "aplica": False,
                            "motivo": f"Pieza '{piece_code}' sin simétrico definido",
                            "encontrados": []})

        _NO_SIM = {"ok": True, "aplica": True,
                   "pieza_contraria": pieza_contraria, "encontrados": []}

        # ── INTERSECT: todos los atributos que deben coincidir exactamente ────
        # Solo cambia Z_PIECE_TYPE (LH↔RH, etc.) — todo lo demás idéntico
        criterios = {
            "Z_VEHICLE_MODEL":          attrs.get("Z_VEHICLE_MODEL",          ""),
            "Z_SUBPRODUCT":             attrs.get("Z_SUBPRODUCT",             ""),
            "Z_FORMULA_CODE":           attrs.get("Z_FORMULA_CODE",           ""),
            "Z_COLOR":                  attrs.get("Z_COLOR",                  ""),
            "Z_SHADE_BAND":             attrs.get("Z_SHADE_BAND",             ""),
            "Z_AGP_LEVEL":              attrs.get("Z_AGP_LEVEL",              ""),
            "Z_BEHAVIOR_DIFFERENTIALS": attrs.get("Z_BEHAVIOR_DIFFERENTIALS", ""),
            "Z_COMMERCIAL_THICKNESS":   attrs.get("Z_COMMERCIAL_THICKNESS",   ""),
            "Z_AGP_VERSION":            attrs.get("Z_AGP_VERSION",            ""),
        }
        
        intersects, params_i = [], []
        for atnam, val in criterios.items():
            if not val:
                continue
            # Z_COMMERCIAL_THICKNESS se almacena desde ATFLV (numérico), no ATWRT
            if atnam == "Z_COMMERCIAL_THICKNESS":
                intersects.append(
                    f"SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001 "
                    f"WHERE CENTRO='CO01' AND ATNAM='{atnam}' AND CAST(ATFLV AS VARCHAR(50))=?"
                )
            else:
                intersects.append(
                    f"SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001 "
                    f"WHERE CENTRO='CO01' AND ATNAM='{atnam}' AND ATWRT=?"
                )
            params_i.append(val)

        # Debe tener la pieza contraria (LH→RH, etc.)
        intersects.append(
            "SELECT MATERIAL FROM dbo.ODATA_ZFER_CLASS_001 "
            "WHERE CENTRO='CO01' AND ATNAM='Z_PIECE_TYPE' AND ATWRT LIKE ?"
        )
        params_i.append(f"%{pieza_contraria}%")

        if len(intersects) < 2:
            return jsonify(_NO_SIM)

        intersect_sql = "\nINTERSECT\n".join(intersects)
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute(f"""
            SELECT TOP 3 m.MATERIAL, h.TEXTO_BREVE_MATERIAL
            FROM ({intersect_sql}) m
            JOIN dbo.ODATA_ZFER_HEAD h ON h.MATERIAL=m.MATERIAL AND h.CENTRO='CO01'
            WHERE m.MATERIAL <> ?
              AND UPPER(ISNULL(h.STATUS,'')) != 'ZZ'
            ORDER BY TRY_CAST(m.MATERIAL AS BIGINT) DESC
        """, *params_i, zfer)
        filas = cur.fetchall()
        conn.close()

        resultados = [{"zfer": str(r[0]), "desc": str(r[1] or ""),
                       "pieza_contraria": pieza_contraria} for r in filas]
        return jsonify({"ok": True, "aplica": True,
                        "pieza_contraria": pieza_contraria,
                        "encontrados": resultados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200


# ══════════════════════════════════════════════════════════════════════════════
# ── COLA DE HOMOLOGACIONES SAP ────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

import time as _time
from datetime import datetime as _dt

def _cola_proximo_bloque() -> dict | None:
    """Retorna el bloque PENDIENTE más próximo con timer activo. None si no hay."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                SELECT TOP 1 id, bloque_num, hora_prog, timer_activo,
                       (SELECT COUNT(*) FROM itg.M5_COLA WHERE bloque_id=b.id AND estado='PENDIENTE') AS n
                FROM itg.M5_BLOQUES b
                WHERE estado='PENDIENTE' AND timer_activo=1
                ORDER BY hora_prog ASC
            """)
            r = cur.fetchone()
            if not r: return None
            return {"id": r[0], "bloque_num": r[1],
                    "hora_prog": r[2].strftime("%d/%m/%Y %H:%M") if r[2] else "",
                    "pendientes": r[4]}
    except Exception:
        return None


def _cola_archivar_y_limpiar(bloque_id: int, ejecutado_por: str = "sistema"):
    """Al completar un bloque: mueve ejecutados → M5_LogEjecuciones, limpia M5_Cola. Deja el bloque COMPLETADO visible."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                INSERT INTO itg.M5_LOGEJECUCIONES
                    (bloque_id, zfer_base, tipo, color, color_nombre, zpla, franja,
                     pn_base, nivel, tipo_pieza, formula_nueva, acero_dir,
                     zfer_nuevo, estado, error_msg, ejecutado_el, ejecutado_por)
                SELECT bloque_id, zfer_base, tipo, color, color_nombre, zpla, franja,
                       pn_base, nivel, tipo_pieza, formula_nueva, acero_dir,
                       zfer_nuevo, estado, error_msg,
                       ISNULL(ejecutado_el, GETDATE()), ?
                FROM itg.M5_COLA
                WHERE bloque_id = ? AND estado IN ('OK','ERROR')
            """, ejecutado_por, bloque_id)
            archivados = cur.rowcount
            # Limpiar cola temporal (todos los items del bloque)
            cur.execute("DELETE FROM itg.M5_COLA WHERE bloque_id=?", bloque_id)
            # NO borrar el bloque — queda COMPLETADO para que el usuario vea el reporte
            print(f"[COLA] bloque {bloque_id}: {archivados} archivados → M5_LogEjecuciones, bloque queda COMPLETADO")
    except Exception as e:
        print(f"[COLA] error archivando bloque {bloque_id}: {e}")


def _cola_ejecutar_bloque(bloque_id: int, ejecutado_por: str = "sistema"):
    """Saca los items PENDIENTE del bloque y los envía a SAP."""
    import importlib
    ok_n, err_n = 0, 0   # inicializar ANTES de cualquier try para evitar UnboundLocalError
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("UPDATE itg.M5_BLOQUES SET estado='EJECUTANDO' WHERE id=?", bloque_id)
            cur.execute("""
                SELECT id, zfer_base, tipo, color, color_nombre, zpla, franja,
                       pn_base, nivel, tipo_pieza, formula_nueva, acero_dir, cambiar_hr, zhal,
                       ISNULL(subproducto,'')
                FROM itg.M5_COLA
                WHERE bloque_id=? AND estado='PENDIENTE'
            """, bloque_id)
            rows = cur.fetchall()

        if not rows:
            with _get_conn_local() as cn:
                cn.cursor().execute("UPDATE itg.M5_BLOQUES SET estado='COMPLETADO' WHERE id=?", bloque_id)
            return

        cola = [{"tipo": r[2], "zfer": r[1], "color": r[3], "color_nombre": r[4],
                 "zpla": r[5], "franja": r[6] or "00", "pn_base": r[7] or "",
                 "nivel": r[8] or "", "tipo_pieza": r[9] or "", "formula_nueva": r[10] or "",
                 "acero_dir": r[11] or "", "cambiar_hr": bool(r[12]), "zhal": r[13] or "",
                 "subproducto": r[14] or "",
                 "_cola_id": r[0]} for r in rows]

        try:
            sap = importlib.import_module("sap_auto")

            # ── Verificar / abrir SAP antes de empezar ────────────────────
            _AutoSAP = getattr(sap, "AutomatizadorSAP", None)
            if _AutoSAP:
                _sap_inst = _AutoSAP()
                if not _sap_inst.asegurar_sap_abierto():
                    print(f"[COLA] bloque {bloque_id}: no se pudo abrir SAP — abortando")
                    with _get_conn_local() as cn:
                        cn.cursor().execute(
                            "UPDATE itg.M5_BLOQUES SET estado='ERROR' WHERE id=?", bloque_id
                        )
                    return

            # Usar funciones libres del módulo (no instancia — ver sap_auto.py)
            _proc_color             = getattr(sap, "procesar_combinacion",                          None)
            _proc_formula_sin       = getattr(sap, "procesar_combinacion_formula_sin_acero",      None)
            _proc_formula_con       = getattr(sap, "procesar_combinacion_formula_con_acero",      None)
            _proc_formula_mismo     = getattr(sap, "procesar_combinacion_formula_mismo_acero",    None)

            def _tiene_diferencial_06(zfer: str) -> bool:
                """Consulta BD SAP: retorna True si el ZFER tiene '06' en Z_BEHAVIOR_DIFFERENTIALS."""
                try:
                    with get_conn() as cn:
                        cur = cn.cursor()
                        cur.execute("""
                            SELECT TOP 1 ATWRT FROM dbo.ODATA_ZFER_CLASS_001
                            WHERE MATERIAL = ? AND ATNAM = 'Z_BEHAVIOR_DIFFERENTIALS' AND ATWRT = '06'
                        """, zfer)
                        return cur.fetchone() is not None
                except Exception as e:
                    print(f"[COLA] _tiene_diferencial_06({zfer}): {e}")
                    return False  # ante duda, no asumir con acero

            for item in cola:
                # ── Cada item tiene su propio try/except — un error nunca detiene los demás ──
                try:
                    tipo = item["tipo"].upper()
                    if tipo == "FORMULA":
                        # Usar dirección explícita del usuario; fallback: auto-detectar
                        acero_dir = item.get("acero_dir", "")
                        if acero_dir == "con_sin":
                            usar_sin = True
                        elif acero_dir == "sin_con":
                            usar_sin = False
                        else:
                            usar_sin = _tiene_diferencial_06(item["zfer"])
                        print(f"[COLA] {item['zfer']} acero_dir={acero_dir or 'auto'} → {'con→sin' if usar_sin else 'sin→con'}")
                        _pm = item.get("plano_manual","") or ""
                        if usar_sin and _proc_formula_sin:
                            res = _proc_formula_sin(
                                item["zfer"], item.get("formula_nueva",""),
                                item["color"], item.get("color_nombre",""),
                                item.get("franja","00"), item.get("pn_base",""),
                                item.get("zpla",""), item.get("nivel",""), item.get("tipo_pieza",""),
                                subproducto=item.get("subproducto",""),
                                plano_manual=_pm
                            )
                        elif not usar_sin and _proc_formula_con:
                            res = _proc_formula_con(
                                item["zfer"], item.get("formula_nueva",""),
                                item["color"], item.get("color_nombre",""),
                                item.get("franja","00"), item.get("pn_base",""),
                                item.get("zpla",""), item.get("nivel",""), item.get("tipo_pieza",""),
                                zhal=item.get("zhal",""),
                                subproducto=item.get("subproducto",""),
                                plano_manual=_pm
                            )
                        else:
                            raise RuntimeError("No se encontró función de procesamiento de fórmula en sap_auto")
                    elif tipo in ("FORMULA_SIN_SIN", "FORMULA_CON_CON"):
                        print(f"[COLA] {item['zfer']} tipo={tipo} → mismo acero")
                        if _proc_formula_mismo:
                            res = _proc_formula_mismo(
                                item["zfer"], item.get("formula_nueva",""),
                                item["color"], item.get("color_nombre",""),
                                item.get("franja","00"), item.get("pn_base",""),
                                item.get("zpla",""), item.get("nivel",""), item.get("tipo_pieza",""),
                                cambio_hr=item.get("cambiar_hr", True),
                                subproducto=item.get("subproducto",""),
                                plano_manual=item.get("plano_manual","") or ""
                            )
                        else:
                            raise RuntimeError("No se encontró función procesar_combinacion_formula_mismo_acero en sap_auto")
                    elif tipo == "FORMULA_CON_ACERO":
                        print(f"[COLA] {item['zfer']} tipo=FORMULA_CON_ACERO → sin→con acero")
                        if _proc_formula_con:
                            res = _proc_formula_con(
                                item["zfer"], item.get("formula_nueva",""),
                                item["color"], item.get("color_nombre",""),
                                item.get("franja","00"), item.get("pn_base",""),
                                item.get("zpla",""), item.get("nivel",""), item.get("tipo_pieza",""),
                                zhal=item.get("zhal",""),
                                subproducto=item.get("subproducto",""),
                                plano_manual=item.get("plano_manual","") or ""
                            )
                        else:
                            raise RuntimeError("No se encontró función de procesamiento en sap_auto")
                    elif _proc_color:
                        res = _proc_color(
                            item["zfer"], item["color"], item.get("color_nombre",""),
                            item.get("franja","00"), item.get("pn_base",""),
                            item.get("zpla",""), item.get("nivel",""), item.get("tipo_pieza","")
                        )
                    else:
                        raise RuntimeError("No se encontró función de procesamiento en sap_auto")
                    estado_item = "OK" if getattr(res, "estado", "") == "OK" else "ERROR"
                    zfer_nuevo  = getattr(res, "zfer_nuevo", "") or ""
                    error_msg   = getattr(res, "error",     "") or ""

                    # ── Cambio de Hoja de Ruta (CA02) ──────────────────────────
                    # FORMULA_SIN_SIN y FORMULA_CON_CON ya hacen CA02 internamente en procesar_formula_mismo_acero
                    # FORMULA y FORMULA_CON_ACERO: el worker lo hace aquí
                    # Color: solo si el usuario marcó cambiar_hr=True
                    if estado_item == "OK" and zfer_nuevo:
                        # SIN_SIN y CON_CON ya hacen CA02 internamente — nunca repetir aquí
                        if tipo in ("FORMULA_SIN_SIN", "FORMULA_CON_CON"):
                            hacer_hr = False
                        else:
                            hacer_hr = (tipo in ("FORMULA", "FORMULA_CON_ACERO")) or item.get("cambiar_hr", False)
                        if hacer_hr:
                            print(f"[COLA] {item['zfer']} → buscando HR candidata para {zfer_nuevo}…")
                            hr_id, hr_desc, hr_err = _hr_buscar_candidata(item["zfer"], zfer_nuevo)
                            if hr_err:
                                print(f"[COLA] HR candidata no encontrada: {hr_err}")
                                error_msg = (error_msg + f" | HR: {hr_err}").strip(" |")
                            else:
                                print(f"[COLA] HR candidata: {hr_id} ({hr_desc})")
                                try:
                                    hr_res = sap.cambiar_hoja_ruta(zfer_nuevo, hr_id)
                                    if hr_res["ok"]:
                                        print(f"[COLA] HR cambiada: {zfer_nuevo} → HR {hr_id}")
                                        # ZINGP0004 al final si HR cambió exitosamente
                                        try:
                                            import importlib as _il
                                            _fn_zing = getattr(_il.import_module("sap_mantenimiento"), "zinpg0004_actualizar", None)
                                            if _fn_zing:
                                                print(f"[COLA] ZINGP0004 para {zfer_nuevo}...")
                                                _zing_res = _fn_zing([zfer_nuevo])
                                                if _zing_res.get("ok"):
                                                    print(f"[COLA] ZINGP0004 OK: {zfer_nuevo}")
                                                else:
                                                    print(f"[COLA] ZINGP0004 WARN: {_zing_res.get('error','')}")
                                        except Exception as _zing_ex:
                                            print(f"[COLA] ZINGP0004 error (no bloquea): {_zing_ex}")
                                    else:
                                        error_msg = (error_msg + f" | CA02: {hr_res['error']}").strip(" |")
                                except Exception as hr_ex:
                                    print(f"[COLA] error CA02: {hr_ex}")
                                    error_msg = (error_msg + f" | CA02: {str(hr_ex)[:200]}").strip(" |")

                except Exception as ex:
                    estado_item, zfer_nuevo, error_msg = "ERROR", "", str(ex)[:500]
                    print(f"[COLA] item {item['_cola_id']} error: {ex}")

                # Adjuntar advertencias SAP al error_msg si el item fue OK pero tuvo warnings
                if estado_item == "OK" and res and getattr(res, "advertencias", []):
                    adv_txt = " | ".join(res.advertencias)
                    error_msg = f"[ADV] {adv_txt}"
                # Guardar resultado — también en try para no cortar el loop si falla la BD
                try:
                    with _get_conn_local() as cn:
                        cn.cursor().execute("""
                            UPDATE itg.M5_COLA
                            SET estado=?, ejecutado_el=GETDATE(), zfer_nuevo=?, error_msg=?
                            WHERE id=?
                        """, estado_item, zfer_nuevo or None, error_msg or None, item["_cola_id"])
                except Exception as db_ex:
                    print(f"[COLA] error guardando item {item['_cola_id']}: {db_ex}")

                # Registrar homologación de fórmula en tabla dedicada
                if estado_item == "OK" and tipo in ("FORMULA", "FORMULA_CON_ACERO", "FORMULA_SIN_SIN", "FORMULA_CON_CON") and res:
                    _hom_id = _guardar_homologacion_formula(item, res, session_user=ejecutado_por)
                    if _hom_id:
                        _guardar_gestor_auto(item, res, _hom_id)

                if estado_item == "OK": ok_n += 1
                else:                  err_n += 1

        except Exception as sap_ex:
            # Solo llega aquí si falló la carga del módulo sap_auto en sí
            print(f"[COLA] error cargando sap_auto en bloque {bloque_id}: {sap_ex}")
            try:
                with _get_conn_local() as cn:
                    cn.cursor().execute("""
                        UPDATE itg.M5_COLA SET estado='ERROR', error_msg=?
                        WHERE bloque_id=? AND estado='PENDIENTE'
                    """, str(sap_ex)[:500], bloque_id)
                err_n = len(cola)
            except Exception:
                pass

        _scheduler_disparados.discard(bloque_id)
        with _get_conn_local() as cn:
            cur2 = cn.cursor()
            cur2.execute("""
                UPDATE itg.M5_BLOQUES
                SET estado='COMPLETADO', ejecutado_el=GETDATE(), ok_count=?, error_count=?
                WHERE id=?
            """, ok_n, err_n, bloque_id)
            # Si no quedó ningún bloque PENDIENTE, crear uno nuevo para mañana 7am
            cur2.execute("SELECT COUNT(*) FROM itg.M5_BLOQUES WHERE estado='PENDIENTE'")
            if cur2.fetchone()[0] == 0:
                from datetime import timedelta as _td
                manana7 = (_dt.now() + _td(days=1)).replace(hour=7, minute=0, second=0, microsecond=0)
                cur2.execute("SELECT ISNULL(MAX(bloque_num),0)+1 FROM itg.M5_BLOQUES")
                nuevo_num = cur2.fetchone()[0]
                cur2.execute(
                    "INSERT INTO itg.M5_BLOQUES (bloque_num, hora_prog, timer_activo) VALUES (?,?,1)",
                    nuevo_num, manana7
                )
        # Mover ejecutados de M5_Cola → M5_LogEjecuciones y limpiar cola temporal
        _cola_archivar_y_limpiar(bloque_id, ejecutado_por)
        print(f"[COLA] bloque {bloque_id} completado: OK={ok_n} ERR={err_n}")

    except Exception as e:
        _scheduler_disparados.discard(bloque_id)
        print(f"[COLA] error ejecutando bloque {bloque_id}: {e}")


def _cola_limpiar_al_inicio():
    """Al arrancar: borra COMPLETADOS vacíos y resetea EJECUTANDO a PENDIENTE (por reinicios inesperados)."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            # Borrar bloques COMPLETADOS con más de 7 días (historial en LogEjecuciones ya los tiene)
            cur.execute("""
                DELETE FROM itg.M5_BLOQUES
                WHERE estado = 'COMPLETADO'
                  AND ejecutado_el < DATEADD(day, -7, GETDATE())
            """)
            if cur.rowcount:
                print(f"[COLA] limpieza: {cur.rowcount} bloques completados viejos eliminados")
            # Resetear bloques EJECUTANDO a PENDIENTE (quedaron pegados por crash/restart)
            cur.execute("SELECT id FROM itg.M5_BLOQUES WHERE estado='EJECUTANDO'")
            bloques_pegados = [r[0] for r in cur.fetchall()]
            if bloques_pegados:
                cur.execute("UPDATE itg.M5_BLOQUES SET estado='PENDIENTE' WHERE estado='EJECUTANDO'")
                ph = ",".join("?" * len(bloques_pegados))
                # Resetear TODOS los items (PENDIENTE y ERROR) de esos bloques
                cur.execute(f"""
                    UPDATE itg.M5_COLA
                    SET estado='PENDIENTE', error_msg=NULL, ejecutado_el=NULL
                    WHERE bloque_id IN ({ph})
                      AND estado IN ('PENDIENTE','ERROR')
                """, *bloques_pegados)
                print(f"[COLA] reset: {len(bloques_pegados)} bloque(s) pegados reseteados a PENDIENTE")
    except Exception as e:
        print(f"[COLA] error limpieza inicial: {e}")


_scheduler_disparados: set = set()   # IDs ya lanzados en este proceso

def _cola_scheduler():
    """Hilo de fondo: cada 20s revisa bloques vencidos."""
    _cola_limpiar_al_inicio()  # limpiar residuos y resetear pegados al arrancar
    # Revisión inmediata al arrancar (por si la hora ya pasó)
    _cola_scheduler_tick()
    while True:
        _time.sleep(20)
        _cola_scheduler_tick()

def _cola_scheduler_tick():
    """Un ciclo de revisión: dispara bloques cuya hora_prog ya llegó."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                SELECT id, hora_prog FROM itg.M5_BLOQUES
                WHERE estado='PENDIENTE' AND timer_activo=1 AND hora_prog IS NOT NULL
            """)
            ahora = _dt.now()
            vencidos = [r[0] for r in cur.fetchall() if r[1] <= ahora]
        for bid in vencidos:
            if bid not in _scheduler_disparados:
                _scheduler_disparados.add(bid)
                print(f"[COLA] scheduler: disparando bloque {bid}")
                threading.Thread(target=_cola_ejecutar_bloque, args=(bid,), daemon=True).start()
    except Exception as e:
        print(f"[COLA] scheduler error: {e}")

# Arrancar hilo scheduler al iniciar la app
_t_cola = threading.Thread(target=_cola_scheduler, daemon=True)
_t_cola.start()


@app.route("/api/cola/estado")
@login_required
def api_cola_estado():
    """Estado actual de la cola: próximo bloque + total pendientes."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            # Próximo bloque pendiente con timer activo
            cur.execute("""
                SELECT TOP 1 b.id, b.bloque_num, b.hora_prog, b.timer_activo,
                       COUNT(c.id) AS n_items
                FROM itg.M5_BLOQUES b
                LEFT JOIN itg.M5_COLA c ON c.bloque_id=b.id AND c.estado='PENDIENTE'
                WHERE b.estado='PENDIENTE' AND b.timer_activo=1
                GROUP BY b.id, b.bloque_num, b.hora_prog, b.timer_activo
                ORDER BY b.hora_prog ASC
            """)
            bloque = cur.fetchone()
            # Total en cola (todos los bloques)
            cur.execute("SELECT COUNT(*) FROM itg.M5_COLA WHERE estado='PENDIENTE'")
            total = cur.fetchone()[0]
            # Timer activo o no (si todos los bloques tienen timer=0)
            cur.execute("SELECT COUNT(*) FROM itg.M5_BLOQUES WHERE timer_activo=1 AND estado='PENDIENTE'")
            timer_on = cur.fetchone()[0] > 0

        if bloque:
            hora_str = bloque[2].strftime("%d/%m/%Y %H:%M") if bloque[2] else ""
            ya_paso  = bloque[2] < _dt.now() if bloque[2] else False
            return jsonify({
                "ok": True, "timer_activo": bool(bloque[3]),
                "bloque_id": bloque[0], "bloque_num": bloque[1],
                "hora_prog": hora_str, "ya_paso": ya_paso,
                "pendientes": bloque[4], "total_pendientes": total,
            })
        return jsonify({"ok": True, "timer_activo": timer_on,
                        "bloque_num": None, "hora_prog": None,
                        "pendientes": 0, "total_pendientes": total})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "total_pendientes": 0})


@app.route("/api/cola/agregar", methods=["POST"])
@login_required
def api_cola_agregar():
    """Agrega items al próximo bloque disponible. Si timer OFF → ejecuta inmediato."""
    body  = request.get_json() or {}
    items = body.get("items", [])
    if not items:
        return jsonify({"ok": False, "error": "No hay ítems"}), 400

    usuario = _usuario_actual()
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            # Buscar bloque PENDIENTE (nunca EJECUTANDO) — cualquier hora, timer on/off
            cur.execute("""
                SELECT TOP 1 id, bloque_num, hora_prog
                FROM itg.M5_BLOQUES
                WHERE estado='PENDIENTE'
                ORDER BY hora_prog ASC
            """)
            bloque = cur.fetchone()

            if not bloque:
                # No hay bloque PENDIENTE → crear uno para mañana 7am
                from datetime import timedelta
                manana7 = (_dt.now() + timedelta(days=1)).replace(hour=7, minute=0, second=0, microsecond=0)
                cur.execute("SELECT ISNULL(MAX(bloque_num),0)+1 FROM itg.M5_BLOQUES")
                nuevo_num = cur.fetchone()[0]
                cur.execute("""
                    INSERT INTO itg.M5_BLOQUES (bloque_num, hora_prog, timer_activo)
                    OUTPUT INSERTED.id, INSERTED.bloque_num, INSERTED.hora_prog
                    VALUES (?, ?, 1)
                """, nuevo_num, manana7)
                row = cur.fetchone()
                bloque_id, bloque_num, hora_prog = row[0], row[1], row[2]
            else:
                bloque_id, bloque_num, hora_prog = bloque[0], bloque[1], bloque[2]

            # Insertar items
            for it in items:
                desc = f"{it.get('tipo','COLOR').upper()} · {it.get('zfer','')} · color {it.get('color','')} · {it.get('formula_nueva','')}"
                tipo_raw   = str(it.get("tipo","color")).upper()
                # Mapear tipo según acero_dir para fórmulas
                acero_dir_it = str(it.get("acero_dir","")).lower().strip()
                if tipo_raw == "FORMULA":
                    if acero_dir_it == "con_sin":
                        tipo_item = "FORMULA"
                    elif acero_dir_it == "sin_con":
                        tipo_item = "FORMULA_CON_ACERO"
                    elif acero_dir_it == "sin_sin":
                        tipo_item = "FORMULA_SIN_SIN"
                    elif acero_dir_it == "con_con":
                        tipo_item = "FORMULA_CON_CON"
                    else:
                        tipo_item = "FORMULA"  # fallback
                else:
                    tipo_item = tipo_raw
                # cambiar_hr: fórmulas siempre True, colores según elección del usuario
                cambiar_hr = True if tipo_item.startswith("FORMULA") else bool(it.get("cambiar_hr", False))
                zhal_val = str(it.get("zhal",""))[:20] or None if tipo_item.startswith("FORMULA") else None
                plano_manual_val = str(it.get("plano_manual","")).strip()[:100] or None
                cur.execute("""
                    INSERT INTO itg.M5_COLA
                    (bloque_id, zfer_base, tipo, color, color_nombre, zpla, franja,
                     pn_base, nivel, tipo_pieza, formula_nueva, descripcion, acero_dir, cambiar_hr, zhal, subproducto, plano_manual)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, bloque_id,
                    str(it.get("zfer",""))[:20], tipo_item[:20],
                    str(it.get("color",""))[:10], str(it.get("color_nombre",""))[:100],
                    str(it.get("zpla",""))[:20], str(it.get("franja","00"))[:5],
                    str(it.get("pn_base",""))[:50], str(it.get("nivel",""))[:10],
                    str(it.get("tipo_pieza",""))[:10], str(it.get("formula_nueva",""))[:30],
                    desc[:200], str(it.get("acero_dir",""))[:10] or None,
                    1 if cambiar_hr else 0, zhal_val,
                    str(it.get("subproducto",""))[:20] or None, plano_manual_val)

        hora_str = hora_prog.strftime("%d/%m/%Y %H:%M") if hora_prog else ""
        return jsonify({"ok": True, "bloque_id": bloque_id, "bloque_num": bloque_num,
                        "hora_prog": hora_str, "n_agregados": len(items)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200


@app.route("/api/cola/bloques")
@login_required
def api_cola_bloques():
    """Lista todos los bloques con sus items."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                SELECT b.id, b.bloque_num, b.hora_prog, b.timer_activo, b.estado,
                       b.creado_el, b.ejecutado_el, b.ok_count, b.error_count,
                       COUNT(c.id) AS total_items,
                       SUM(CASE WHEN c.estado='PENDIENTE'  THEN 1 ELSE 0 END) AS pend,
                       SUM(CASE WHEN c.estado='OK'         THEN 1 ELSE 0 END) AS ok_n,
                       SUM(CASE WHEN c.estado='ERROR'      THEN 1 ELSE 0 END) AS err_n
                FROM itg.M5_BLOQUES b
                LEFT JOIN itg.M5_COLA c ON c.bloque_id = b.id
                GROUP BY b.id, b.bloque_num, b.hora_prog, b.timer_activo, b.estado,
                         b.creado_el, b.ejecutado_el, b.ok_count, b.error_count
                ORDER BY b.hora_prog DESC
            """)
            bloques_rows = cur.fetchall()
            # Fetch items per bloque
            bloque_ids = [r[0] for r in bloques_rows]
            items_map: dict = {}
            if bloque_ids:
                placeholders = ",".join("?" * len(bloque_ids))
                cur.execute(f"""
                    SELECT bloque_id, zfer_base, tipo, color, color_nombre, zpla, formula_nueva,
                           estado, zfer_nuevo, error_msg
                    FROM itg.M5_COLA WHERE bloque_id IN ({placeholders})
                    ORDER BY id
                """, *bloque_ids)
                for row in cur.fetchall():
                    bid = row[0]
                    items_map.setdefault(bid, []).append({
                        "zfer_base": row[1], "tipo": row[2], "color": row[3],
                        "color_nombre": row[4], "zpla": row[5], "formula_nueva": row[6],
                        "estado": row[7], "zfer_nuevo": row[8], "error_msg": row[9],
                    })
            bloques = []
            for r in bloques_rows:
                bloques.append({
                    "id": r[0], "bloque_num": r[1],
                    "hora_prog": r[2].isoformat() if r[2] else "",
                    "ejecutado_el": r[6].isoformat() if r[6] else None,
                    "timer_activo": bool(r[3]), "estado": r[4],
                    "ok_count": r[7], "error_count": r[8],
                    "total_items": r[9] or 0, "pendientes": r[10] or 0,
                    "items": items_map.get(r[0], []),
                })
            return jsonify({"ok": True, "bloques": bloques})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 200

@app.route("/api/cola/bloque/<int:bloque_id>/ejecutar", methods=["POST"])
@login_required
def api_cola_ejecutar_bloque(bloque_id: int):
    """Dispara la ejecución manual de un bloque. Valida que tenga ítems pendientes."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                SELECT b.estado, COUNT(c.id) AS n_pend
                FROM itg.M5_BLOQUES b
                LEFT JOIN itg.M5_COLA c ON c.bloque_id=b.id AND c.estado='PENDIENTE'
                WHERE b.id=?
                GROUP BY b.estado
            """, bloque_id)
            row = cur.fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Bloque no encontrado"})
        estado_bloque, n_pend = row[0], row[1] or 0
        if estado_bloque == "EJECUTANDO":
            return jsonify({"ok": False, "error": "El bloque ya está ejecutándose"})
        if n_pend == 0:
            return jsonify({"ok": False, "error": "El bloque no tiene ítems pendientes"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    usuario = _usuario_actual() or "sistema"
    threading.Thread(target=_cola_ejecutar_bloque, args=(bloque_id, usuario), daemon=True).start()
    return jsonify({"ok": True, "mensaje": f"Bloque {bloque_id} iniciado ({n_pend} ítems)"})


@app.route("/api/cola/bloque/<int:bloque_id>/timer", methods=["POST"])
@login_required
def api_cola_toggle_timer(bloque_id: int):
    """Activa/desactiva el timer de un bloque."""
    body = request.get_json() or {}
    activo = 1 if body.get("activo") else 0
    try:
        with _get_conn_local() as cn:
            cn.cursor().execute("UPDATE itg.M5_BLOQUES SET timer_activo=? WHERE id=?", activo, bloque_id)
        return jsonify({"ok": True, "timer_activo": bool(activo)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/cola/bloque/<int:bloque_id>/hora", methods=["POST"])
@login_required
def api_cola_cambiar_hora(bloque_id: int):
    """Cambia la hora programada de un bloque."""
    body = request.get_json() or {}
    hora_str = (body.get("hora") or body.get("hora_prog", "")).strip().rstrip("Z")
    try:
        # Normalizar: quitar zona horaria y milisegundos, dejar solo YYYY-MM-DDTHH:MM
        hora_str_norm = hora_str[:16]  # "2025-01-15T07:00"
        hora = _dt.strptime(hora_str_norm, "%Y-%m-%dT%H:%M")
        with _get_conn_local() as cn:
            cn.cursor().execute("UPDATE itg.M5_BLOQUES SET hora_prog=? WHERE id=? AND estado='PENDIENTE'",
                                hora, bloque_id)
        return jsonify({"ok": True, "hora_prog": hora.strftime("%d/%m/%Y %H:%M")})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/cola/bloque/<int:bloque_id>/reset", methods=["POST"])
@login_required
def api_cola_reset_bloque(bloque_id: int):
    """Resetea un bloque EJECUTANDO a PENDIENTE (para bloques pegados por error/crash)."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                UPDATE itg.M5_BLOQUES SET estado='PENDIENTE'
                WHERE id=? AND estado IN ('EJECUTANDO','ERROR','COMPLETADO')
            """, bloque_id)
            if cur.rowcount == 0:
                return jsonify({"ok": False, "error": "Bloque no encontrado o ya está en estado PENDIENTE"})
            # Resetear TODOS los ítems (incluso EJECUTANDO) a PENDIENTE para reintento completo
            cur.execute("""
                UPDATE itg.M5_COLA SET estado='PENDIENTE', error_msg=NULL, ejecutado_el=NULL
                WHERE bloque_id=? AND estado IN ('PENDIENTE','ERROR','EJECUTANDO')
            """, bloque_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/cola/bloque/<int:bloque_id>/borrar", methods=["POST"])
@login_required
def api_cola_borrar_bloque(bloque_id: int):
    """Borra un bloque y todos sus ítems permanentemente."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("DELETE FROM itg.M5_COLA WHERE bloque_id=?", bloque_id)
            cur.execute("DELETE FROM itg.M5_BLOQUES WHERE id=?", bloque_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/cola/bloque/<int:bloque_id>/vaciar", methods=["POST"])
@login_required
def api_cola_vaciar_bloque(bloque_id: int):
    """Elimina los ítems PENDIENTE de un bloque (no los ya ejecutados)."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("DELETE FROM itg.M5_COLA WHERE bloque_id=? AND estado='PENDIENTE'", bloque_id)
            deleted = cur.rowcount
            # Si el bloque queda vacío y sigue PENDIENTE, eliminarlo también
            cur.execute("""
                DELETE FROM itg.M5_BLOQUES WHERE id=? AND estado='PENDIENTE'
                AND NOT EXISTS (SELECT 1 FROM itg.M5_COLA WHERE bloque_id=?)
            """, bloque_id, bloque_id)
        return jsonify({"ok": True, "deleted": deleted})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/indicadores")
@login_required
def indicadores():
    return render_template("indicadores.html")


@app.route("/api/indicadores/data")
@login_required
def api_indicadores_data():
    """KPIs y series temporales desde M5_LogEjecuciones."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()

            # ── Totales globales ────────────────────────────────────────────
            cur.execute("""
                SELECT
                    SUM(CASE WHEN estado='OK'    THEN 1 ELSE 0 END) AS total_ok,
                    SUM(CASE WHEN estado='ERROR' THEN 1 ELSE 0 END) AS total_error,
                    SUM(CASE WHEN tipo='COLOR' AND estado='OK'    THEN 1 ELSE 0 END) AS colores_ok,
                    SUM(CASE WHEN tipo='COLOR' AND estado='ERROR' THEN 1 ELSE 0 END) AS colores_error,
                    SUM(CASE WHEN tipo LIKE 'FORMULA%' AND estado='OK'    THEN 1 ELSE 0 END) AS formulas_ok,
                    SUM(CASE WHEN tipo LIKE 'FORMULA%' AND estado='ERROR' THEN 1 ELSE 0 END) AS formulas_error
                FROM itg.M5_LOGEJECUCIONES
            """)
            row = cur.fetchone()
            total_ok       = row[0] or 0
            total_error    = row[1] or 0
            colores_ok     = row[2] or 0
            colores_error  = row[3] or 0
            formulas_ok    = row[4] or 0
            formulas_error = row[5] or 0

            # ── Por día — últimos 30 días ───────────────────────────────────
            cur.execute("""
                SELECT
                    CAST(ejecutado_el AS DATE) AS fecha,
                    SUM(CASE WHEN tipo='COLOR'        AND estado='OK'    THEN 1 ELSE 0 END) AS colores_ok,
                    SUM(CASE WHEN tipo='COLOR'        AND estado='ERROR' THEN 1 ELSE 0 END) AS colores_error,
                    SUM(CASE WHEN tipo LIKE 'FORMULA%' AND estado='OK'   THEN 1 ELSE 0 END) AS formulas_ok,
                    SUM(CASE WHEN tipo LIKE 'FORMULA%' AND estado='ERROR' THEN 1 ELSE 0 END) AS formulas_error
                FROM itg.M5_LOGEJECUCIONES
                WHERE ejecutado_el >= DATEADD(day, -30, GETDATE())
                GROUP BY CAST(ejecutado_el AS DATE)
                ORDER BY fecha
            """)
            por_dia = [
                {
                    "fecha": str(r[0]),
                    "colores_ok": r[1] or 0,
                    "colores_error": r[2] or 0,
                    "formulas_ok": r[3] or 0,
                    "formulas_error": r[4] or 0,
                }
                for r in cur.fetchall()
            ]

            # ── Por semana — últimas 12 semanas ────────────────────────────
            cur.execute("""
                SELECT
                    DATEPART(year,  ejecutado_el)*100 + DATEPART(week, ejecutado_el) AS semana,
                    MIN(CAST(ejecutado_el AS DATE)) AS fecha_inicio,
                    SUM(CASE WHEN tipo='COLOR'        AND estado='OK'    THEN 1 ELSE 0 END) AS colores_ok,
                    SUM(CASE WHEN tipo='COLOR'        AND estado='ERROR' THEN 1 ELSE 0 END) AS colores_error,
                    SUM(CASE WHEN tipo LIKE 'FORMULA%' AND estado='OK'   THEN 1 ELSE 0 END) AS formulas_ok,
                    SUM(CASE WHEN tipo LIKE 'FORMULA%' AND estado='ERROR' THEN 1 ELSE 0 END) AS formulas_error
                FROM itg.M5_LOGEJECUCIONES
                WHERE ejecutado_el >= DATEADD(week, -12, GETDATE())
                GROUP BY DATEPART(year, ejecutado_el)*100 + DATEPART(week, ejecutado_el)
                ORDER BY semana
            """)
            por_semana = [
                {
                    "semana": r[0],
                    "fecha_inicio": str(r[1]),
                    "colores_ok": r[2] or 0,
                    "colores_error": r[3] or 0,
                    "formulas_ok": r[4] or 0,
                    "formulas_error": r[5] or 0,
                }
                for r in cur.fetchall()
            ]

            # ── Top 10 ZFERs base por ejecuciones OK ───────────────────────
            cur.execute("""
                SELECT TOP 10
                    zfer_base,
                    COUNT(*) AS total
                FROM itg.M5_LOGEJECUCIONES
                WHERE estado='OK' AND zfer_base IS NOT NULL AND zfer_base <> ''
                GROUP BY zfer_base
                ORDER BY total DESC
            """)
            top_zfer = [{"zfer_base": r[0], "total": r[1]} for r in cur.fetchall()]

            # ── Últimas 10 ejecuciones ──────────────────────────────────────
            cur.execute("""
                SELECT TOP 10
                    zfer_nuevo, tipo, estado, ejecutado_el, formula_nueva, color_nombre, zfer_base
                FROM itg.M5_LOGEJECUCIONES
                ORDER BY ejecutado_el DESC
            """)
            recientes = [
                {
                    "zfer_nuevo":    r[0] or "",
                    "tipo":          r[1] or "",
                    "estado":        r[2] or "",
                    "ejecutado_el":  r[3].strftime("%d/%m/%Y %H:%M") if r[3] else "",
                    "formula_nueva": r[4] or "",
                    "color_nombre":  r[5] or "",
                    "zfer_base":     r[6] or "",
                }
                for r in cur.fetchall()
            ]

        return jsonify({
            "ok": True,
            "total_ok": total_ok,
            "total_error": total_error,
            "colores_ok": colores_ok,
            "colores_error": colores_error,
            "formulas_ok": formulas_ok,
            "formulas_error": formulas_error,
            "por_dia": por_dia,
            "por_semana": por_semana,
            "top_zfer": top_zfer,
            "recientes": recientes,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/cola/historial")
@login_required
def api_cola_historial():
    """Últimas ejecuciones del log permanente (M5_LogEjecuciones)."""
    limit = min(int(request.args.get("limit", 100)), 500)
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute(f"""
                SELECT TOP {limit}
                    id, bloque_id, zfer_base, tipo, color, color_nombre, zpla,
                    formula_nueva, tipo_pieza, zfer_nuevo, estado, error_msg, ejecutado_el
                FROM itg.M5_LOGEJECUCIONES
                ORDER BY ejecutado_el DESC
            """)
            rows = cur.fetchall()
        log = [{"id": r[0], "bloque_id": r[1], "zfer_base": r[2], "tipo": r[3],
                "color": r[4], "color_nombre": r[5], "zpla": r[6],
                "formula_nueva": r[7], "tipo_pieza": r[8], "zfer_nuevo": r[9],
                "estado": r[10], "error_msg": r[11],
                "ejecutado_el": r[12].strftime("%d/%m/%Y %H:%M") if r[12] else ""}
               for r in rows]
        return jsonify({"ok": True, "log": log})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "log": []})


@app.route("/api/cola/bloque/<int:bloque_id>/reporte")
@login_required
def api_cola_bloque_reporte(bloque_id: int):
    """Items ejecutados de un bloque (desde M5_LogEjecuciones)."""
    try:
        with _get_conn_local() as cn:
            cur = cn.cursor()
            # Info del bloque
            cur.execute("""
                SELECT bloque_num, hora_prog, ejecutado_el, ok_count, error_count
                FROM itg.M5_BLOQUES WHERE id=?
            """, bloque_id)
            br = cur.fetchone()
            cur.execute("""
                SELECT zfer_base, tipo, color, color_nombre, zpla, franja,
                       formula_nueva, tipo_pieza, zfer_nuevo, estado, error_msg, ejecutado_el
                FROM itg.M5_LOGEJECUCIONES WHERE bloque_id=?
                ORDER BY ejecutado_el
            """, bloque_id)
            rows = cur.fetchall()
        items = [{"zfer_base": r[0], "tipo": r[1], "color": r[2], "color_nombre": r[3],
                  "zpla": r[4], "franja": r[5] or "00", "formula_nueva": r[6], "tipo_pieza": r[7],
                  "zfer_nuevo": r[8], "estado": r[9], "error_msg": r[10],
                  "ejecutado_el": r[11].strftime("%d/%m/%Y %H:%M:%S") if r[11] else ""}
                 for r in rows]
        ok_n  = sum(1 for i in items if i["estado"] == "OK")
        err_n = sum(1 for i in items if i["estado"] == "ERROR")
        bloque_info = {
            "num": br[0] if br else bloque_id,
            "hora_prog": br[1].strftime("%d/%m/%Y %H:%M") if br and br[1] else "",
            "ejecutado_el": br[2].strftime("%d/%m/%Y %H:%M") if br and br[2] else "",
        } if br else {}
        return jsonify({"ok": True, "bloque_id": bloque_id, "bloque": bloque_info,
                        "items": items, "ok_count": ok_n, "error_count": err_n})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "items": []})


@app.route("/api/cola/bloque/<int:bloque_id>/reporte/excel")
@login_required
def api_cola_bloque_excel(bloque_id: int):
    """Descarga Excel corporativo del reporte de un bloque ejecutado."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        with _get_conn_local() as cn:
            cur = cn.cursor()
            cur.execute("""
                SELECT bloque_num, hora_prog, ejecutado_el, ok_count, error_count
                FROM itg.M5_BLOQUES WHERE id=?
            """, bloque_id)
            br = cur.fetchone()
            # Detectar columnas opcionales (pueden no existir si aún no se corrió la migración)
            cur.execute("""
                SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME='M5_LogEjecuciones'
                AND COLUMN_NAME IN ('pn_base','acero_dir','ejecutado_por','nivel')
            """)
            existing_cols = {r[0] for r in cur.fetchall()}
            def _col(name, alias=None, default="''"):
                a = alias or name
                return f"ISNULL({name},'')" if name in existing_cols else f"{default} AS {a}"
            sel_extra = ", ".join([
                _col("pn_base"),
                _col("acero_dir"),
                _col("ejecutado_por"),
                _col("nivel"),
            ])
            cur.execute(f"""
                SELECT zfer_base, tipo, color, color_nombre, zpla, franja,
                       formula_nueva, tipo_pieza, zfer_nuevo, estado, error_msg,
                       ejecutado_el, {sel_extra}
                FROM itg.M5_LOGEJECUCIONES WHERE bloque_id=?
                ORDER BY tipo, ejecutado_el
            """, bloque_id)
            rows = cur.fetchall()

        # ── Paleta corporativa AGP ──────────────────────────────────────────
        C_NAVY    = "0D1B2A"   # header principal
        C_TEAL    = "00848A"   # acento AGP
        C_OK      = "0A3D1F"   # fondo verde OK
        C_OK_TXT  = "56D364"   # texto verde
        C_ERR     = "3D0A0A"   # fondo rojo ERROR
        C_ERR_TXT = "FF7B7B"   # texto rojo
        C_ADV     = "2E2200"   # fondo amarillo ADVERTENCIA
        C_ADV_TXT = "E3B341"   # texto amarillo
        C_HDR_TXT = "FFFFFF"
        C_ROW_A   = "0D1117"   # fila par
        C_ROW_B   = "161B22"   # fila impar
        C_LABEL   = "8B949E"   # etiquetas resumen

        def _fill(c): return PatternFill("solid", fgColor=c)
        def _font(c, bold=False, sz=10): return Font(color=c, bold=bold, size=sz, name="Calibri")
        thin = Side(style="thin", color="21262D")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr  = Alignment(horizontal="center", vertical="center", wrap_text=False)
        lft  = Alignment(horizontal="left",   vertical="center", wrap_text=False)

        def _estado_fill_font(estado, error_msg):
            em = (error_msg or "").strip()
            if estado == "OK" and em.startswith("[ADV]"):
                return _fill(C_ADV), _font(C_ADV_TXT)
            if estado == "OK":
                return _fill(C_OK),  _font(C_OK_TXT)
            return _fill(C_ERR), _font(C_ERR_TXT)

        def _ts(dt):
            return dt.strftime("%d/%m/%Y %H:%M:%S") if dt else ""

        def _duracion(r):
            # r no tiene duracion_seg directo en este query — dejamos vacío
            return ""

        # Clasificar rows
        ok_rows  = [r for r in rows if r[9] == "OK" and not (r[10] or "").startswith("[ADV]")]
        adv_rows = [r for r in rows if r[9] == "OK" and (r[10] or "").startswith("[ADV]")]
        err_rows = [r for r in rows if r[9] == "ERROR"]
        col_rows = [r for r in rows if r[1] and r[1].upper() == "COLOR"]
        for_rows = [r for r in rows if r[1] and r[1].upper().startswith("FORMULA")]

        ok_n  = len(ok_rows) + len(adv_rows)
        err_n = len(err_rows)
        adv_n = len(adv_rows)
        total = len(rows)
        hora_prog_str = _ts(br[1]) if br else ""
        ejecutado_str = _ts(br[2]) if br else ""
        operador      = rows[0][14] if rows and rows[0][14] else "—"

        COLS = ["#", "ZFER Base", "PN Base", "Tipo", "Dirección Acero",
                "Color (cód.)", "Color Nombre", "Fórmula Nueva", "Tipo Pieza",
                "ZPLA", "Franja", "Nivel", "ZFER Nuevo", "Estado",
                "Hora Ejecución", "Operador", "Advertencias / Error"]

        def _header_row(ws, title_color=C_NAVY):
            for col_idx, col_name in enumerate(COLS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = _fill(title_color)
                cell.font = _font(C_HDR_TXT, bold=True, sz=10)
                cell.alignment = ctr
                cell.border = brd
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        def _write_row(ws, row_num, r, alt=False):
            ts   = _ts(r[11])
            em   = (r[10] or "").strip()
            adv  = em[5:].strip() if em.startswith("[ADV]") else ""
            err  = em if not em.startswith("[ADV]") else ""
            msg  = adv if adv else err
            base_fill, txt_font = _estado_fill_font(r[9], r[10])
            row_fill = base_fill if r[9] == "ERROR" or (r[9]=="OK" and em.startswith("[ADV]")) else _fill(C_ROW_B if alt else C_ROW_A)

            vals = [row_num - 1, r[0] or "", r[12] or "", r[1] or "",
                    r[13] or "", r[2] or "", r[3] or "", r[6] or "",
                    r[7] or "", r[4] or "", r[5] or "00", r[15] or "",
                    r[8] or "", r[9] or "", ts, r[14] or "", msg]

            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = brd
                cell.alignment = lft
                if r[9] == "ERROR":
                    cell.fill = _fill(C_ERR)
                    cell.font = _font(C_ERR_TXT, sz=9)
                elif em.startswith("[ADV]"):
                    cell.fill = _fill(C_ADV)
                    cell.font = _font(C_ADV_TXT, sz=9)
                else:
                    cell.fill = row_fill
                    cell.font = _font("C9D1D9", sz=9)
                # Columna Estado con color especial
                if col_idx == 14:
                    cell.font = txt_font
                    cell.font = Font(color=txt_font.color, bold=True, size=9, name="Calibri")
                    cell.alignment = ctr

        def _write_items(ws, subset, color_header=C_NAVY):
            _header_row(ws, color_header)
            for i, r in enumerate(subset):
                _write_row(ws, i + 2, r, alt=(i % 2 == 1))
            # Ajuste de anchos fijos (mejor que autofit)
            widths = [4, 14, 22, 18, 16, 10, 22, 14, 12, 14, 8, 8, 14, 10, 20, 18, 45]
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

        wb = Workbook()
        
        # ── Hoja 1: RESUMEN ──────────────────────────────────────────────────
        ws_r = wb.active
        ws_r.title = "RESUMEN"
        ws_r.sheet_view.showGridLines = False

        # Título
        ws_r.merge_cells("A1:D1")
        t = ws_r["A1"]
        t.value = "AGP Glass — Reporte de Ejecución SAP"
        t.fill  = _fill(C_NAVY)
        t.font  = _font(C_HDR_TXT, bold=True, sz=14)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws_r.row_dimensions[1].height = 32

        ws_r.merge_cells("A2:D2")
        sub = ws_r["A2"]
        sub.value = f"Bloque #{br[0] if br else bloque_id}  ·  {ejecutado_str}  ·  Operador: {operador}"
        sub.fill  = _fill(C_TEAL)
        sub.font  = _font(C_HDR_TXT, sz=10)
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws_r.row_dimensions[2].height = 20

        ws_r.append([])  # fila vacía

        kpis = [
            ("Total ítems procesados", total, "C9D1D9"),
            ("✅  Exitosos (OK)",        ok_n,  C_OK_TXT),
            ("⚠️  Con advertencias",     adv_n, C_ADV_TXT),
            ("❌  Con error",            err_n, C_ERR_TXT),
            ("",                         "",    "C9D1D9"),
            ("Tasa de éxito",   f"{round(ok_n/total*100,1)}%" if total else "—", C_OK_TXT),
            ("",                         "",    "C9D1D9"),
            ("Cambios de Color",   len(col_rows), "58A6FF"),
            ("Cambios de Fórmula", len(for_rows), "D2A8FF"),
            ("",                         "",    "C9D1D9"),
            ("Hora programada",    hora_prog_str, "C9D1D9"),
            ("Hora de ejecución",  ejecutado_str, "C9D1D9"),
        ]
        for lbl, val, color in kpis:
            ws_r.append([lbl, val])
            row = ws_r.max_row
            if lbl:
                ws_r.cell(row, 1).font = _font(C_LABEL, bold=True, sz=10)
                ws_r.cell(row, 2).font = _font(color, bold=True, sz=12)
                ws_r.cell(row, 1).fill = _fill(C_ROW_A)
                ws_r.cell(row, 2).fill = _fill(C_ROW_A)
            ws_r.row_dimensions[row].height = 18

        ws_r.column_dimensions["A"].width = 28
        ws_r.column_dimensions["B"].width = 18

        # ── Hoja 2: Todos ────────────────────────────────────────────────────
        ws_t = wb.create_sheet("📋 Todos")
        ws_t.sheet_view.showGridLines = False
        _write_items(ws_t, rows, C_NAVY)

        # ── Hoja 3: OK ───────────────────────────────────────────────────────
        ws_ok = wb.create_sheet("✅ OK")
        ws_ok.sheet_view.showGridLines = False
        _write_items(ws_ok, ok_rows, "0A3D1F")

        # ── Hoja 4: Advertencias ─────────────────────────────────────────────
        if adv_rows:
            ws_a = wb.create_sheet("⚠️ Advertencias")
            ws_a.sheet_view.showGridLines = False
            _write_items(ws_a, adv_rows, "2E2200")

        # ── Hoja 5: Errores ───────────────────────────────────────────────────
        if err_rows:
            ws_e = wb.create_sheet("❌ Errores")
            ws_e.sheet_view.showGridLines = False
            _write_items(ws_e, err_rows, "3D0A0A")

        # ── Hoja 6: Fórmulas ─────────────────────────────────────────────────
        if for_rows:
            ws_f = wb.create_sheet("🔬 Fórmulas")
            ws_f.sheet_view.showGridLines = False
            _write_items(ws_f, for_rows, "1A0E2D")

        # ── Hoja 7: Colores ───────────────────────────────────────────────────
        if col_rows:
            ws_c = wb.create_sheet("🎨 Colores")
            ws_c.sheet_view.showGridLines = False
            _write_items(ws_c, col_rows, "0D1E2E")

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        bloque_num = br[0] if br else bloque_id
        return send_file(buf, as_attachment=True,
                         download_name=f"AGP_Reporte_Bloque_{bloque_num}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/cola")
@login_required
def cola_page():
    return render_template("cola.html")


# ══════════════════════════════════════════════════════════════════════════════
# HOJAS DE RUTA
# ══════════════════════════════════════════════════════════════════════════════

# Mapas fijos (tabla del cuadro de claves modelo)
_HR_CLAVE_FORMULA = {100:"01VEXT",200:"02VP01",300:"03VP02",400:"04VP03",
                     500:"05VP04",600:"06VP05",700:"07VP06",800:"08VP07"}
_HR_CLAVE_BASE    = {99:"32VPMO"}
_HR_CLAVE_PROT    = {199:"33VPR01",299:"34VPR02"}
_HR_CLAVE_TAPAS   = {3600:"36VTPA",3700:"37VSTP"}


def _hr_construir_criterios(attrs_base: dict, area, bom_posiciones: list,
                             metrologia_base, prueba_agua_base) -> dict:
    """Construye el dict de criterios para buscar en ODATA_HR_CONSULTA."""
    bom = set(bom_posiciones)

    # ── Atributos del ZFER base ──────────────────────────────────────────────
    level_raw = str(attrs_base.get("Z_AGP_LEVEL","") or "").strip().lstrip("0") or "0"
    try:    nivel = "BAJO" if int(level_raw) <= 3 else "ALTO"
    except: nivel = level_raw or None

    geo = str(attrs_base.get("Z_GEOMETRY_TYPE","") or "").strip()
    geometria = "CURVO" if geo == "02" else ("PLANO" if geo == "01" else None)

    try:
        a = float(area or 0)
        tamano = "PEQUEÑO" if a <= 0.6 else ("MEDIANO" if a <= 0.99 else "GRANDE")
    except:
        tamano = None

    # ── BOM → claves ─────────────────────────────────────────────────────────
    formula_claves = [_HR_CLAVE_FORMULA[p] for p in sorted(_HR_CLAVE_FORMULA) if p in bom]
    base_claves    = [_HR_CLAVE_BASE[p]    for p in sorted(_HR_CLAVE_BASE)    if p in bom]
    prot_claves    = [_HR_CLAVE_PROT[p]    for p in sorted(_HR_CLAVE_PROT)    if p in bom]
    tapas_claves   = [_HR_CLAVE_TAPAS[p]   for p in sorted(_HR_CLAVE_TAPAS)   if p in bom]

    # ── SERIGRAFIA ────────────────────────────────────────────────────────────
    has_9301      = 9301 in bom
    ext_9302_9312 = [p for p in bom if 9302 <= p <= 9312]
    ext_9452_9462 = [p for p in bom if 9452 <= p <= 9462]
    seri_claves = None
    if has_9301:
        seri_claves = ["01VEXT"]
        if ext_9302_9312:
            if formula_claves:
                seri_claves.append(formula_claves[-1])
            if ext_9452_9462 and len(formula_claves) >= 2:
                seri_claves.append(formula_claves[-2])
        elif ext_9452_9462:
            if formula_claves:
                seri_claves.append(formula_claves[-1])

    # VITRIFICADO = SERIGRAFIA
    vit_claves = list(seri_claves) if seri_claves else None

    # ── MECANIZADO: misma lógica que SERIGRAFIA ───────────────────────────────
    mec_claves = list(seri_claves) if seri_claves else None

    # ── EMPALME (solo CURVO) ──────────────────────────────────────────────────
    # Si es CURVO: suma claves de formula+base+protectors+tapas
    # Si es PLANO: NULL
    # Si es CURVO pero BOM sin posiciones conocidas: marca "not_null" (IS NOT NULL)
    es_curvo = geometria == "CURVO"
    empalme_claves = (formula_claves + base_claves + prot_claves + tapas_claves) \
                     if es_curvo else None

    # ── CURVADO (solo CURVO) ──────────────────────────────────────────────────
    curv = 1 if es_curvo else None

    # ── CURV_ACERO ────────────────────────────────────────────────────────────
    curv_acero = 1 if (106 in bom or 116 in bom) else None

    def _nn(lst):
        return (len(lst), ",".join(lst)) if lst else (None, None)

    f_n, f_t   = _nn(formula_claves)
    b_n, b_t   = _nn(base_claves)
    p_n, p_t   = _nn(prot_claves)
    ta_n, ta_t = _nn(tapas_claves)
    s_n, s_t   = _nn(seri_claves)
    v_n, v_t   = _nn(vit_claves)
    m_n, m_t   = _nn(mec_claves)

    if empalme_claves is None:
        # pieza plana → EMPALME IS NULL
        e_n, e_t = None, None
        empalme_not_null = False
    elif empalme_claves:
        # pieza curva con claves conocidas → filtrar exacto
        e_n, e_t = len(empalme_claves), ",".join(empalme_claves)
        empalme_not_null = False
    else:
        # pieza curva pero BOM sin posiciones mapeadas → solo exigir IS NOT NULL
        e_n, e_t = None, None
        empalme_not_null = True

    # ── Validador antenas pasta plata en paquete (posiciones 9452–9456) ────────
    alertas = []
    antenas_pp = [p for p in bom if 9452 <= p <= 9456]
    if antenas_pp:
        # Última posición de vidrio de fórmula presente en el BOM
        formula_pos_en_bom = sorted(p for p in _HR_CLAVE_FORMULA if p in bom)
        if formula_pos_en_bom:
            ultima_pos_formula = formula_pos_en_bom[-1]       # ej: 400
            digito_formula = str(ultima_pos_formula)[0]        # ej: "4"
            for pos_antena in antenas_pp:
                digito_antena = str(pos_antena)[-1]            # ej: 9454 → "4"
                if digito_antena != digito_formula:
                    alertas.append(
                        f"⚠ Posición antena {pos_antena} no coincide con última posición "
                        f"de fórmula {ultima_pos_formula} "
                        f"(dígito fórmula='{digito_formula}', dígito antena='{digito_antena}')"
                    )
        else:
            alertas.append(
                f"⚠ Posiciones de antena {antenas_pp} encontradas pero no hay "
                f"posiciones de fórmula (100-800) en el BOM para validar"
            )

    return {
        "nivel": nivel, "geometria": geometria, "tamano": tamano,
        "formula": f_n, "txt_formula": f_t,
        "base": b_n,    "txt_base": b_t,
        "protectors": p_n, "txt_protectors": p_t,
        "tapas": ta_n,  "txt_tapas": ta_t,
        "serigrafia": s_n, "txt_serigrafia": s_t,
        "vitrificado": v_n, "txt_vitrificado": v_t,
        "mecanizado": m_n,  "txt_mecanizado": m_t,
        "empalme": e_n, "txt_empalme": e_t,
        "empalme_not_null": empalme_not_null,
        "ent_horno_cur": curv, "curvado": curv, "sal_horno_cur": curv,
        "curv_acero": curv_acero,
        "metrologia": metrologia_base,
        "prueba_agua": prueba_agua_base,
        "alertas": alertas,
    }


def _hr_buscar_candidata(zfer_base: str, zfer_nuevo: str) -> tuple:
    """
    Dado zfer_base (atributos BD) y zfer_nuevo (BOM SAP), busca la HR candidata:
    la de mayor MATERIALES que no supere 450.
    Retorna (id_hruta: str | None, descripcion: str, error: str | None)
    """
    try:
        import importlib
        sap = importlib.import_module("sap_auto")
        bom_result = sap.leer_bom_material(zfer_nuevo)
        if not bom_result.get("ok"):
            return None, "", f"SAP BOM error: {bom_result.get('error','')}"
        bom_posiciones = bom_result["posiciones"]
    except Exception as e:
        return None, "", f"Error leyendo BOM: {e}"

    try:
        attrs = q_atributos(zfer_base)
        if "_error" in attrs:
            return None, "", f"Error atributos: {attrs['_error']}"
        head  = q_zfer_head(zfer_base)
        area  = None
        if head and "_error" not in head:
            try: area = float(head.get("AREA") or 0) or None
            except Exception: pass

        metrologia_base = prueba_agua_base = None
        try:
            with get_conn() as cn:
                cur = cn.cursor()
                cur.execute("""
                    SELECT TOP 1 C.METROLOGIA, C.PRUEBA_AGUA
                    FROM dbo.ODATA_HR_CONSULTA C
                    JOIN dbo.HR_MATERIALS M ON C.ID_HRUTA = M.ID_HRUTA
                    WHERE M.MATERIAL = ? AND C.TIPO_HR = 'PRODUCCION'
                """, zfer_base)
                row = cur.fetchone()
                if row:
                    metrologia_base  = row[0]
                    prueba_agua_base = row[1]
        except Exception:
            pass

        criterios = _hr_construir_criterios(attrs, area, bom_posiciones, metrologia_base, prueba_agua_base)
        resultados, _, _, _ = _hr_buscar(criterios)

        elegibles = [r for r in resultados if r.get("MATERIALES") is not None and r["MATERIALES"] <= 300]
        if not elegibles:
            return None, "", "Sin HRs candidatas con materiales ≤ 300"
        candidata = max(elegibles, key=lambda r: r["MATERIALES"])
        return str(candidata["ID_HRUTA"]), candidata.get("DESCRIPCION",""), None

    except Exception as e:
        return None, "", f"Error buscando candidata: {e}"


def _hr_buscar(crit: dict) -> list:
    """Ejecuta el query sobre ODATA_HR_CONSULTA y retorna filas + SQL construido."""
    conditions, params = ["C.TIPO_HR = 'PRODUCCION'", "C.MATERIALES IS NOT NULL"], []

    def _campo_txt(col_n, col_t, count, txt, excluir_null=False):
        if count is None:
            conditions.append(f"C.{col_n} IS NULL")
        else:
            claves = [c.strip() for c in (txt or "").split(",") if c.strip()]
            # excluir_null=True → solo trae filas con valor exacto (no acepta NULL)
            # excluir_null=False → acepta NULL o valor exacto (comportamiento original)
            inner = f"C.{col_n} = ? AND " + " AND ".join(f"C.{col_t} LIKE ?" for _ in claves)
            cond  = f"({inner})" if excluir_null else f"(C.{col_n} IS NULL OR ({inner}))"
            conditions.append(cond)
            params.append(count)
            params.extend(f"%{c}%" for c in claves)

    def _campo_int(col, val):
        if val is None:
            conditions.append(f"C.{col} IS NULL")
        else:
            conditions.append(f"C.{col} = ?")
            params.append(val)

    if crit["tamano"]:   conditions.append("C.TAMANO = ?");   params.append(crit["tamano"])
    if crit["nivel"]:    conditions.append("C.NIVEL = ?");    params.append(crit["nivel"])
    if crit["geometria"]:conditions.append("C.GEOMETRIA = ?");params.append(crit["geometria"])

    _campo_txt("FORMULA",    "TXT_FORMULA",    crit["formula"],    crit["txt_formula"])
    _campo_txt("BASE",       "TXT_BASE",       crit["base"],       crit["txt_base"])
    _campo_txt("PROTECTORS", "TXT_PROTECTORS", crit["protectors"], crit["txt_protectors"])
    _campo_txt("TAPAS",      "TXT_TAPAS",      crit["tapas"],      crit["txt_tapas"])
    _campo_txt("SERIGRAFIA", "TXT_SERIGRAFIA", crit["serigrafia"], crit["txt_serigrafia"], excluir_null=True)
    _campo_txt("VITRIFICADO","TXT_VITRIFICADO",crit["vitrificado"],crit["txt_vitrificado"], excluir_null=True)
    _campo_txt("MECANIZADO", "TXT_MECANIZADO", crit["mecanizado"], crit["txt_mecanizado"],  excluir_null=True)
    # EMPALME: si es curvo pero BOM sin claves → solo IS NOT NULL (no filtrar exacto)
    if crit.get("empalme_not_null"):
        conditions.append("C.EMPALME IS NOT NULL")
    else:
        _campo_txt("EMPALME", "TXT_EMPALME", crit["empalme"], crit["txt_empalme"])

    _campo_int("ENT_HORNO_CUR", crit["ent_horno_cur"])
    _campo_int("CURVADO",       crit["curvado"])
    _campo_int("SAL_HORNO_CUR", crit["sal_horno_cur"])
    _campo_int("CURV_ACERO",    crit["curv_acero"])
    _campo_int("METROLOGIA",    crit["metrologia"])
    _campo_int("PRUEBA_AGUA",   crit["prueba_agua"])
    _campo_int("PRELAMINADO",   None)   # siempre NULL por ahora

    sql = f"""
        SELECT C.ID_HRUTA, C.DESCRIPCION, C.SUB_RUTA,
               C.TAMANO, C.NIVEL, C.GEOMETRIA, C.MATERIALES,
               C.FORMULA, C.TXT_FORMULA, C.BASE, C.TXT_BASE,
               C.PROTECTORS, C.TXT_PROTECTORS, C.TAPAS, C.TXT_TAPAS,
               C.SERIGRAFIA, C.TXT_SERIGRAFIA, C.VITRIFICADO, C.TXT_VITRIFICADO,
               C.MECANIZADO, C.TXT_MECANIZADO, C.EMPALME, C.TXT_EMPALME,
               C.ENT_HORNO_CUR, C.CURVADO, C.SAL_HORNO_CUR,
               C.CURV_ACERO, C.METROLOGIA, C.PRUEBA_AGUA,
               M.TOTAL_MATERIALES
        FROM dbo.ODATA_HR_CONSULTA C
        LEFT JOIN (
            SELECT ID_HRUTA, COUNT(DISTINCT MATERIAL) AS TOTAL_MATERIALES
            FROM dbo.HR_MATERIALS GROUP BY ID_HRUTA
        ) M ON C.ID_HRUTA = M.ID_HRUTA
        WHERE {" AND ".join(conditions)}
        ORDER BY C.ID_HRUTA
    """.strip()

    # SQL con valores inline para mostrar en UI
    sql_display = sql
    for p in params:
        val = f"'{p}'" if isinstance(p, str) else str(p)
        sql_display = sql_display.replace("?", val, 1)

    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(sql, params)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    # Contar HRs con todos los campos idénticos (excepto ID_HRUTA, MATERIALES y TOTAL_MATERIALES)
    _excluir = {"ID_HRUTA", "TOTAL_MATERIALES", "MATERIALES"}
    _cmp_cols = [c for c in cols if c not in _excluir]
    from collections import Counter
    _grupos = Counter(tuple(r.get(c) for c in _cmp_cols) for r in rows)
    n_identicas = sum(cnt for cnt in _grupos.values() if cnt > 1)

    return rows, sql_display, params, n_identicas


@app.route("/hojas_ruta")
@login_required
def hojas_ruta_page():
    return render_template("hoja_ruta.html")


@app.route("/api/hojas_ruta/buscar", methods=["POST"])
@login_required
def api_hojas_ruta_buscar():
    """
    Body: {zfer_base, zfer_nuevo}
    1. Consulta BD con zfer_base → atributos + área + metrología/prueba_agua de HR base
    2. Llama SAP ZPPR0008 con zfer_nuevo → posiciones BOM
    3. Construye criterios → busca en ODATA_HR_CONSULTA
    """
    body       = request.get_json(force=True) or {}
    zfer_base  = str(body.get("zfer_base","")).strip()
    zfer_nuevo = str(body.get("zfer_nuevo","")).strip()
    if not zfer_base or not zfer_nuevo:
        return jsonify({"ok": False, "error": "Se requieren zfer_base y zfer_nuevo"})

    try:
        # ── 1. BD: atributos del ZFER BASE (nivel, geometría, tamaño, etc.) ───
        attrs = q_atributos(zfer_base)
        if "_error" in attrs:
            return jsonify({"ok": False, "error": f"Error consultando atributos: {attrs['_error']}"})

        head = q_zfer_head(zfer_base)
        area = None
        if head and "_error" not in head:
            try:
                area = float(head.get("AREA") or 0) or None
            except Exception:
                pass

        attrs_base = attrs
        head_base  = head

        # ── 2. BD: metrología y prueba de agua del ZFER base (de su HR) ─────
        metrologia_base = None
        prueba_agua_base = None
        try:
            with get_conn() as cn:
                cur = cn.cursor()
                cur.execute("""
                    SELECT TOP 1 C.METROLOGIA, C.PRUEBA_AGUA
                    FROM dbo.ODATA_HR_CONSULTA C
                    JOIN dbo.HR_MATERIALS M ON C.ID_HRUTA = M.ID_HRUTA
                    WHERE M.MATERIAL = ? AND C.TIPO_HR = 'PRODUCCION'
                """, zfer_base)
                row = cur.fetchone()
                if row:
                    metrologia_base   = row[0]
                    prueba_agua_base  = row[1]
        except Exception as e:
            print(f"[HR] metrología/prueba base: {e}")

        # ── 3. SAP: BOM del ZFER nuevo ──────────────────────────────────────
        try:
            import importlib
            sap = importlib.import_module("sap_auto")
            bom_result = sap.leer_bom_material(zfer_nuevo)
        except Exception as e:
            return jsonify({"ok": False, "error": f"Error accediendo SAP: {e}"})

        if not bom_result.get("ok"):
            return jsonify({"ok": False, "error": bom_result.get("error","SAP error")})

        bom_posiciones = bom_result["posiciones"]
        bom_filas      = bom_result["filas"]

        # ── 4. Construir criterios y buscar HR ───────────────────────────────
        criterios = _hr_construir_criterios(
            attrs, area, bom_posiciones, metrologia_base, prueba_agua_base
        )
        resultados, sql_usado, _, n_identicas = _hr_buscar(criterios)

        # Info para mostrar en UI
        zfer_base_info = {
            "material": zfer_base,
            "texto": head_base.get("TEXTO_BREVE_MATERIAL","") if head_base and "_error" not in head_base else "",
            "formula": attrs_base.get("Z_FORMULA_CODE","") if "_error" not in attrs_base else "",
            "differentials": attrs_base.get("Z_BEHAVIOR_DIFFERENTIALS","") if "_error" not in attrs_base else "",
            # cabecera del ZFER nuevo (usada para criterios)
            "area": area,
            "nivel": attrs.get("Z_AGP_LEVEL",""),
            "geometria": attrs.get("Z_GEOMETRY_TYPE",""),
        }

        return jsonify({
            "ok": True,
            "zfer_base_info": zfer_base_info,
            "bom_posiciones": bom_posiciones,
            "bom_filas": bom_filas,
            "criterios": criterios,
            "resultados": [
                {"id_hruta": r["ID_HRUTA"], "descripcion": r["DESCRIPCION"],
                 "sub_ruta": r["SUB_RUTA"], "materiales": r.get("MATERIALES"),
                 "tamano": r["TAMANO"], "nivel": r["NIVEL"], "geometria": r["GEOMETRIA"],
                 "formula": r["FORMULA"], "txt_formula": r["TXT_FORMULA"],
                 "base": r["BASE"], "txt_base": r["TXT_BASE"],
                 "protectors": r["PROTECTORS"], "txt_protectors": r["TXT_PROTECTORS"],
                 "tapas": r["TAPAS"], "txt_tapas": r["TXT_TAPAS"],
                 "serigrafia": r["SERIGRAFIA"], "txt_serigrafia": r["TXT_SERIGRAFIA"],
                 "mecanizado": r["MECANIZADO"], "txt_mecanizado": r["TXT_MECANIZADO"],
                 "vitrificado": r["VITRIFICADO"], "txt_vitrificado": r["TXT_VITRIFICADO"],
                 "empalme": r["EMPALME"], "txt_empalme": r["TXT_EMPALME"],
                 "curv_acero": r["CURV_ACERO"], "metrologia": r["METROLOGIA"],
                 "prueba_agua": r["PRUEBA_AGUA"],
                 "ent_horno_cur": r["ENT_HORNO_CUR"], "curvado": r["CURVADO"],
                 "sal_horno_cur": r["SAL_HORNO_CUR"]}
                for r in resultados
            ],
            "sql": sql_usado,
            "n_resultados": len(resultados),
            "n_identicas": n_identicas,
            "alertas": criterios.get("alertas", []),
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── Mantenimiento HR ──────────────────────────────────────────────────────────

_MHR_JSON     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ultimo_mantenimiento_hr.json")
_MHR_TEMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp_mhr")
_MHR_QAS_LOG  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp_mhr", "qas_desasignados.json")
os.makedirs(_MHR_TEMP_DIR, exist_ok=True)


def _qas_leer_desasignados() -> set:
    """Lee el set de ZFERs ya desasignados en pruebas QAS."""
    import json as _j
    try:
        if os.path.exists(_MHR_QAS_LOG):
            return set(_j.load(open(_MHR_QAS_LOG, encoding='utf-8')))
    except Exception:
        pass
    return set()


def _qas_guardar_desasignados(zfers_nuevos: list):
    """Agrega ZFERs al log QAS acumulativo."""
    import json as _j
    actual = _qas_leer_desasignados()
    actual.update(zfers_nuevos)
    with open(_MHR_QAS_LOG, 'w', encoding='utf-8') as f:
        _j.dump(sorted(actual), f)
    print(f"[QAS-LOG] {len(actual)} ZFERs desasignados acumulados")


def _mhr_generar_excel_disco(hr: dict, limite: int = None) -> str:
    """Genera el Excel de desasignación para una HR y lo guarda en disco. Retorna la ruta."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    id_hruta = hr["id_hruta"]
    ruta = os.path.join(_MHR_TEMP_DIR, f"desasignar_{id_hruta}.xlsx")

    # Eliminar archivo previo si existe (evita Permission denied si SAP lo tenía abierto)
    try:
        if os.path.exists(ruta):
            os.remove(ruta)
    except OSError as e:
        raise RuntimeError(
            f"No se puede sobrescribir el archivo Excel — ciérralo en SAP o Excel antes de continuar. ({e})"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(id_hruta)[:31]

    hdr_fill = PatternFill("solid", fgColor="1a4731")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    fill_err  = PatternFill("solid", fgColor="ffc7ce")

    for col, txt in enumerate(["Grupo de hoja de ruta", "Contador HR", "Material"], 1):
        c = ws.cell(row=1, column=col, value=txt)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    # Filtrar QAS primero, luego aplicar límite — así toma los siguientes disponibles
    ya_desasignados = _qas_leer_desasignados()
    todos_disponibles = [z for z in hr["zfers_fuera"] if z not in ya_desasignados]
    filtrados = len(hr["zfers_fuera"]) - len(todos_disponibles)
    if filtrados:
        print(f"[QAS-FILTER] {filtrados} ZFERs omitidos (ya desasignados en pruebas)")
    zfers_list = todos_disponibles[:limite] if limite else todos_disponibles
    if not zfers_list:
        raise RuntimeError("Todos los ZFERs de esta HR ya fueron desasignados en pruebas. Limpia el log QAS para repetir.")
    # Buscar ZFORs — deduplicar antes de la query
    zfor_map = hr.get("zfor_map") or {}
    zfers_uniq = list(dict.fromkeys(zfers_list))
    if zfers_uniq:
        try:
            with get_conn() as cn:
                ph = ",".join(["?"] * len(zfers_uniq))
                cur = cn.cursor()
                cur.execute(
                    f"SELECT MATERIAL, MAX(MAT_CONFIG) AS MAT_CONFIG "
                    f"FROM dbo.ODATA_ZFER_BOM "
                    f"WHERE MATERIAL IN ({ph}) AND MAT_CONFIG IS NOT NULL AND MAT_CONFIG != '' "
                    f"GROUP BY MATERIAL",
                    zfers_uniq
                )
                for row in cur.fetchall():
                    mat = str(row[0]).strip()
                    cfg = str(row[1]).strip() if row[1] else ""
                    if cfg:
                        zfor_map[mat] = cfg
            print(f"[EXCEL-ZFOR] {len(zfor_map)} ZFORs encontrados para {len(zfers_uniq)} ZFERs únicos")
        except Exception as e_z:
            print(f"[EXCEL-ZFOR] ERROR: {e_z}")
    row_i = 2
    for zfer in zfers_list:
        ws.cell(row=row_i, column=1, value=str(id_hruta)).fill = fill_err
        ws.cell(row=row_i, column=2, value="01").fill = fill_err
        ws.cell(row=row_i, column=3, value=zfer).fill = fill_err
        row_i += 1
        zfor = zfor_map.get(zfer, "")
        if zfor:
            ws.cell(row=row_i, column=1, value=str(id_hruta)).fill = fill_err
            ws.cell(row=row_i, column=2, value="01").fill = fill_err
            ws.cell(row=row_i, column=3, value=zfor).fill = fill_err
            row_i += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    wb.save(ruta)
    return ruta


@app.route("/mantenimiento_hr")
@login_required
def mantenimiento_hr():
    return render_template("mantenimiento_hr.html")


@app.route("/api/mantenimiento_hr/consultar")
@login_required
def api_mantenimiento_hr_consultar():
    import json as _json
    from datetime import datetime as _dt
    from collections import defaultdict as _dd
    import time as _time

    def _chunk_in(cn, sql_prefix, sql_suffix, items, chunk=500):
        """Ejecuta query con IN en chunks de 500 para evitar límite de 2100 params de SQL Server."""
        rows = []
        for i in range(0, len(items), chunk):
            batch = items[i:i+chunk]
            ph = ",".join(["?"] * len(batch))
            cur = cn.cursor()
            cur.execute(f"{sql_prefix} ({ph}){sql_suffix}", batch)
            rows.extend(cur.fetchall())
        return rows

    try:
        t0 = _time.time()

        # 1+2. Un solo JOIN: HRs ≥300 materiales + sus ZFERs de HR_MATERIALS
        print("[MHR] Consultando HRs + materiales (JOIN)...")
        with get_conn() as cn:
            cur = cn.cursor()
            cur.execute("""
                SELECT C.ID_HRUTA, C.DESCRIPCION, C.MATERIALES, M.MATERIAL
                FROM dbo.ODATA_HR_CONSULTA C
                JOIN dbo.HR_MATERIALS M ON C.ID_HRUTA = M.ID_HRUTA
                WHERE C.TIPO_HR = 'PRODUCCION' AND C.MATERIALES >= 300
                ORDER BY C.MATERIALES DESC
            """)
            hrs_meta = {}   # id_hruta → {descripcion, materiales}
            zfers_by_hr = _dd(set)   # set evita duplicados (HR_MATERIALS tiene 1 fila por posición BOM)
            for row in cur.fetchall():
                hr_id = str(row[0]).strip()
                if hr_id not in hrs_meta:
                    hrs_meta[hr_id] = {"id_hruta": hr_id,
                                       "descripcion": str(row[1] or ""),
                                       "materiales": int(row[2] or 0)}
                zfers_by_hr[hr_id].add(str(row[3]).strip())
            # Convertir sets a listas para serialización JSON
            zfers_by_hr = {k: list(v) for k, v in zfers_by_hr.items()}

        print(f"[MHR] HRs: {len(hrs_meta)} | ZFERs totales: {sum(len(v) for v in zfers_by_hr.values())} | {_time.time()-t0:.1f}s")

        if not hrs_meta:
            result = {"ok": True, "fecha_consulta": _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                      "total_hrs": 0, "hojas_ruta": []}
            with open(_MHR_JSON, "w", encoding="utf-8") as f:
                _json.dump(result, f, ensure_ascii=False)
            return jsonify(result)

        all_zfers = list({z for zlist in zfers_by_hr.values() for z in zlist})

        # 3. CalendarioAGP — traer TODOS los ZFERcodes de una vez (sin IN enorme)
        #    Más rápido que mandar 5000+ params: una sola query sin filtro, comparar en Python
        print(f"[MHR] Consultando CalendarioAGP ({len(all_zfers)} ZFERs a cruzar)...")
        t1 = _time.time()
        try:
            cal_cn = pyodbc.connect(_CONN_CALENDARIO, timeout=30)
            try:
                cal_cur = cal_cn.cursor()
                cal_cur.execute("SELECT DISTINCT ZFERcode FROM dbo.TCAL_CALENDARIO_COLOMBIA WHERE ZFERcode IS NOT NULL")
                rows_cal = cal_cur.fetchall()
                en_calendario_set = {str(r[0]).strip() for r in rows_cal if r[0]}
            finally:
                cal_cn.close()
        except Exception as cal_err:
            return jsonify({"ok": False, "error": f"Error conectando a CalendarioAGP: {cal_err}"})
        print(f"[MHR] Calendario: {len(en_calendario_set)} ZFERcodes cargados | {_time.time()-t1:.1f}s")
        if not en_calendario_set:
            return jsonify({"ok": False, "error": "ALERTA: CalendarioAGP devolvio 0 registros — tabla TCAL_CALENDARIO_COLOMBIA vacia o inaccesible. Cruce cancelado para evitar resultados incorrectos."})

        # 4. Clasificar en Python (O(1) por lookup en set)
        hrs_orden = list(dict.fromkeys(  # preservar orden original por MATERIALES DESC
            r for zlist in [list(zfers_by_hr.keys())] for r in zlist
        ))
        hojas = []
        for hr_id, meta in hrs_meta.items():
            zfers = zfers_by_hr[hr_id]
            fuera = [z for z in zfers if z not in en_calendario_set]
            hojas.append({**meta,
                "total_zfer":      len(zfers),
                "en_calendario":   len(zfers) - len(fuera),
                "fuera_calendario": len(fuera),
                "zfers_fuera":     fuera,
            })
        hojas.sort(key=lambda h: h["materiales"], reverse=True)

        # 5. ZFOR — solo para ZFERs fuera (deduplicados). Con GROUP BY es rápido (~1-2s)
        all_fuera_uniq = list(dict.fromkeys(z for h in hojas for z in h["zfers_fuera"]))
        zfor_map = {}
        if all_fuera_uniq:
            try:
                t_zfor = _time.time()
                with get_conn() as cn:
                    cur = cn.cursor()
                    ph_z = ",".join(["?"] * len(all_fuera_uniq))
                    cur.execute(
                        f"SELECT MATERIAL, MAX(MAT_CONFIG) FROM dbo.ODATA_ZFER_BOM "
                        f"WHERE MATERIAL IN ({ph_z}) AND MAT_CONFIG IS NOT NULL AND MAT_CONFIG != '' "
                        f"GROUP BY MATERIAL",
                        all_fuera_uniq
                    )
                    for row in cur.fetchall():
                        mat = str(row[0]).strip(); cfg = str(row[1]).strip() if row[1] else ""
                        if cfg: zfor_map[mat] = cfg
                print(f"[MHR] ZFORs: {len(zfor_map)}/{len(all_fuera_uniq)} ZFERs fuera | {_time.time()-t_zfor:.1f}s")
            except Exception as e_zfor:
                print(f"[MHR] WARN ZFOR: {e_zfor}")

        for h in hojas:
            h["zfor_map"] = {z: zfor_map.get(z, "") for z in h["zfers_fuera"]}
            h["sin_zfor"] = [z for z in h["zfers_fuera"] if not zfor_map.get(z, "")]

        print(f"[MHR] Consulta completa en {_time.time()-t0:.1f}s")

        result = {"ok": True,
                  "fecha_consulta": _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                  "total_hrs": len(hojas), "hojas_ruta": hojas}
        with open(_MHR_JSON, "w", encoding="utf-8") as f:
            _json.dump(result, f, ensure_ascii=False)
        return jsonify(result)

    except Exception as e:
        import traceback
        print(f"[MHR] ERROR: {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/exportar/<id_hruta>")
@login_required
def api_mantenimiento_hr_exportar(id_hruta: str):
    """Descarga Excel de ZFERs fuera de calendario para una HR específica.
    2 columnas: ZFER | Hoja de Ruta (descripción). Solo los fuera de calendario."""
    import json as _json
    from datetime import datetime as _dt
    try:
        data = None
        if os.path.exists(_MHR_JSON):
            with open(_MHR_JSON, "r", encoding="utf-8") as f:
                data = _json.load(f)
        if not data or not data.get("ok"):
            return jsonify({"ok": False, "error": "No hay datos. Ejecuta la consulta primero."}), 400

        hr = next((h for h in data["hojas_ruta"] if str(h["id_hruta"]) == str(id_hruta)), None)
        if not hr:
            return jsonify({"ok": False, "error": f"HR {id_hruta} no encontrada en el último resultado."}), 404

        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(id_hruta)[:31]

        hdr_fill = PatternFill("solid", fgColor="1a4731")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        fill_err  = PatternFill("solid", fgColor="ffc7ce")

        # Cabeceras: Grupo HR | Contador | Material
        for col, txt in enumerate(["Grupo de hoja de ruta", "Contador HR", "Material"], 1):
            c = ws.cell(row=1, column=col, value=txt)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

        # Buscar ZFORs frescos (el cache no los tiene, se buscan al generar Excel)
        zfers_para_excel = list(dict.fromkeys(hr.get("zfers_fuera", [])))  # deduplicados
        zfor_map = {}
        if zfers_para_excel:
            try:
                ph_e = ",".join(["?"] * len(zfers_para_excel))
                with get_conn() as cn:
                    cur = cn.cursor()
                    cur.execute(
                        f"SELECT MATERIAL, MAX(MAT_CONFIG) FROM dbo.ODATA_ZFER_BOM "
                        f"WHERE MATERIAL IN ({ph_e}) AND MAT_CONFIG IS NOT NULL AND MAT_CONFIG != '' "
                        f"GROUP BY MATERIAL",
                        zfers_para_excel
                    )
                    for row in cur.fetchall():
                        mat = str(row[0]).strip(); cfg = str(row[1]).strip() if row[1] else ""
                        if cfg: zfor_map[mat] = cfg
                print(f"[EXPORTAR-ZFOR] {len(zfor_map)} ZFORs / {len(zfers_para_excel)} ZFERs únicos")
            except Exception as e_ze:
                print(f"[EXPORTAR-ZFOR] ERROR: {e_ze}")
        row_i = 2
        for zfer in zfers_para_excel:   # deduplicados
            # Fila ZFER
            ws.cell(row=row_i, column=1, value=str(id_hruta)).fill = fill_err
            ws.cell(row=row_i, column=2, value="01").fill = fill_err
            ws.cell(row=row_i, column=3, value=zfer).fill = fill_err
            row_i += 1
            # Fila ZFOR (si existe)
            zfor = zfor_map.get(zfer, "")
            if zfor:
                ws.cell(row=row_i, column=1, value=str(id_hruta)).fill = fill_err
                ws.cell(row=row_i, column=2, value="01").fill = fill_err
                ws.cell(row=row_i, column=3, value=zfor).fill = fill_err
                row_i += 1

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"mhr_{id_hruta}_{_dt.now().strftime('%Y%m%d')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/desasignar/<id_hruta>", methods=["POST"])
@login_required
def api_mantenimiento_hr_desasignar(id_hruta: str):
    """Genera el Excel de desasignación y lo sube a SAP vía ZPPP0084."""
    import json as _json
    try:
        if not os.path.exists(_MHR_JSON):
            return jsonify({"ok": False, "error": "No hay datos. Ejecuta la consulta primero."})
        with open(_MHR_JSON, "r", encoding="utf-8") as f:
            data = _json.load(f)
        hr = next((h for h in data.get("hojas_ruta", []) if str(h["id_hruta"]) == str(id_hruta)), None)
        if not hr:
            return jsonify({"ok": False, "error": f"HR {id_hruta} no encontrada en el último resultado."})
        if not hr.get("zfers_fuera"):
            return jsonify({"ok": False, "error": "Esta HR no tiene ZFERs fuera de calendario."})

        # Límite opcional para pruebas (query param ?limite=10)
        limite = request.args.get("limite", type=int)  # None = todos

        # Generar Excel en disco (con límite si aplica)
        excel_path = _mhr_generar_excel_disco(hr, limite=limite)
        n_enviados = limite if limite and limite < len(hr["zfers_fuera"]) else len(hr["zfers_fuera"])
        print(f"[MHR] Excel generado: {excel_path} ({n_enviados} ZFERs)")

        # Ejecutar ZPPP0084 en SAP desde sap_mantenimiento.py
        import importlib
        sap_mant = importlib.import_module("sap_mantenimiento")
        fn = getattr(sap_mant, "zppp0084_desasignar", None)
        if not fn:
            return jsonify({"ok": False, "error": "sap_mantenimiento.py no encontrado o función faltante."})

        resultado = fn(excel_path)
        # Guardar en log QAS los ZFERs desasignados (solo si SAP OK)
        if resultado.get("ok"):
            zfers_enviados = hr["zfers_fuera"][:limite] if limite else hr["zfers_fuera"]
            ya = _qas_leer_desasignados()
            zfers_reales = [z for z in zfers_enviados if z not in ya]
            if zfers_reales:
                _qas_guardar_desasignados(zfers_reales)
        return jsonify({**resultado, "excel_path": excel_path,
                        "id_hruta": id_hruta, "n_enviados": n_enviados})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/qas_log")
@login_required
def api_mhr_qas_log():
    """Info del log QAS: cuántos ZFERs acumulados."""
    zfers = sorted(_qas_leer_desasignados())
    return jsonify({"total": len(zfers), "zfers": zfers})


@app.route("/api/mantenimiento_hr/qas_limpiar", methods=["POST"])
@login_required
def api_mhr_qas_limpiar():
    """Limpia el log QAS de ZFERs desasignados en pruebas."""
    try:
        if os.path.exists(_MHR_QAS_LOG):
            os.remove(_MHR_QAS_LOG)
        return jsonify({"ok": True, "mensaje": "Log QAS limpiado."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/diag_zfor/<zfer>")
@login_required
def api_mhr_diag_zfor(zfer: str):
    """Diagnóstico: busca el ZFOR de un ZFER específico en ODATA_ZFER_BOM."""
    try:
        with get_conn() as cn:
            cur = cn.cursor()
            cur.execute(
                "SELECT TOP 5 MATERIAL, MAT_CONFIG, CENTRO, TIPO_MATERIAL "
                "FROM dbo.ODATA_ZFER_BOM WHERE MATERIAL = ?", zfer
            )
            rows = [{"MAT_CONFIG": str(r[1] or ""), "CENTRO": str(r[2] or ""),
                     "TIPO": str(r[3] or "")} for r in cur.fetchall()]
            cur.execute(
                "SELECT MAX(MAT_CONFIG) FROM dbo.ODATA_ZFER_BOM "
                "WHERE MATERIAL = ? AND MAT_CONFIG IS NOT NULL AND MAT_CONFIG != ''", zfer
            )
            max_row = cur.fetchone()
            max_cfg = str(max_row[0] or "") if max_row else ""
        return jsonify({"ok": True, "zfer": zfer, "total_rows": len(rows),
                        "sample": rows, "max_mat_config": max_cfg})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/plan_asignacion/<id_hruta>")
@login_required
def api_mhr_plan_asignacion(id_hruta: str):
    """
    Calcula el plan de asignación para los ZFERs fuera de la HR indicada.
    Para cada ZFER busca la HR candidata (solo BD, sin SAP) y agrupa en batches
    respetando el límite de 300 materiales por HR.
    """
    import json as _json
    try:
        if not os.path.exists(_MHR_JSON):
            return jsonify({"ok": False, "error": "No hay datos. Ejecuta la consulta primero."})
        with open(_MHR_JSON, "r", encoding="utf-8") as f:
            data = _json.load(f)
        hr_origen = next((h for h in data.get("hojas_ruta", []) if str(h["id_hruta"]) == str(id_hruta)), None)
        if not hr_origen:
            return jsonify({"ok": False, "error": f"HR {id_hruta} no encontrada."})

        # Aplicar limite si viene (para flujo completo con 1/10 ZFERs)
        # Filtrar QAS primero para tomar los realmente pendientes
        _limite_plan = request.args.get("limite", type=int)
        ya_desasig = _qas_leer_desasignados()
        zfers_fuera_all = hr_origen.get("zfers_fuera", [])
        zfers_disponibles = [z for z in zfers_fuera_all if z not in ya_desasig]
        zfers_fuera = zfers_disponibles[:_limite_plan] if _limite_plan else zfers_disponibles

        if not zfers_fuera:
            return jsonify({"ok": False, "error": "No hay ZFERs disponibles para asignar (todos en log QAS o ninguno fuera de calendario)."})

        # ZFOR: buscar en ODATA_ZFER_BOM — deduplicar ZFERs antes de la query
        zfor_map = hr_origen.get("zfor_map") or {}
        zfers_unicos = list(dict.fromkeys(zfers_fuera))  # orden preservado, sin duplicados
        if zfers_unicos:
            try:
                with get_conn() as cn:
                    ph = ",".join(["?"] * len(zfers_unicos))
                    cur = cn.cursor()
                    cur.execute(
                        f"SELECT MATERIAL, MAX(MAT_CONFIG) AS MAT_CONFIG "
                        f"FROM dbo.ODATA_ZFER_BOM "
                        f"WHERE MATERIAL IN ({ph}) AND MAT_CONFIG IS NOT NULL AND MAT_CONFIG != '' "
                        f"GROUP BY MATERIAL",
                        zfers_unicos
                    )
                    rows = cur.fetchall()
                    print(f"[PLAN-ZFOR] query returned {len(rows)} rows para {len(zfers_unicos)} ZFERs únicos")
                    for row in rows:
                        mat = str(row[0]).strip()
                        cfg = str(row[1]).strip() if row[1] else ""
                        if cfg:
                            zfor_map[mat] = cfg
            except Exception as e_z:
                print(f"[PLAN] ERROR ZFOR: {e_z}")

        # ── Obtener atributos de todos los ZFERs en batch ────────────────────
        attrs_map = {}   # {zfer: {nivel, geometria, tamano, bom_posiciones}}
        if zfers_fuera:
            with get_conn() as cn:
                cur = cn.cursor()
                ph = ",".join(["?"] * len(zfers_fuera))

                # Atributos básicos
                cur.execute(f"""
                    SELECT MATERIAL, AREA FROM dbo.ODATA_ZFER_HEAD
                    WHERE MATERIAL IN ({ph}) AND CENTRO = 'CO01'
                """, zfers_fuera)
                area_map = {str(r[0]).strip(): float(r[1] or 0) for r in cur.fetchall()}

                # Características de clasificación
                cur.execute(f"""
                    SELECT MATERIAL, ATNAM, ATWRT FROM dbo.ODATA_ZFER_CLASS_001
                    WHERE MATERIAL IN ({ph}) AND CENTRO = 'CO01'
                    AND ATNAM IN ('Z_AGP_LEVEL','Z_GEOMETRY_TYPE')
                """, zfers_fuera)
                class_map = {}
                for r in cur.fetchall():
                    mat = str(r[0]).strip()
                    if mat not in class_map:
                        class_map[mat] = {}
                    class_map[mat][str(r[1]).strip()] = str(r[2]).strip()

                # Posiciones BOM desde ODATA_ZFER_BOM (evita llamada SAP)
                cur.execute(f"""
                    SELECT MATERIAL, POSICION FROM dbo.ODATA_ZFER_BOM
                    WHERE MATERIAL IN ({ph})
                """, zfers_fuera)
                bom_map = {}
                for r in cur.fetchall():
                    mat = str(r[0]).strip()
                    try:
                        pos = int(str(r[1]).strip())
                    except Exception:
                        continue
                    bom_map.setdefault(mat, set()).add(pos)

            for zfer in zfers_fuera:
                area  = area_map.get(zfer, 0)
                cls   = class_map.get(zfer, {})
                posiciones = list(bom_map.get(zfer, set()))
                # Guardar valores RAW para _hr_construir_criterios (espera "03", "02", etc.)
                attrs_map[zfer] = {
                    "Z_AGP_LEVEL":     cls.get("Z_AGP_LEVEL", ""),
                    "Z_GEOMETRY_TYPE": cls.get("Z_GEOMETRY_TYPE", ""),
                    "area":            area,
                    "bom_posiciones":  posiciones,
                    "sin_bom":         len(posiciones) == 0,
                }

        # ── Obtener capacidades actuales de todas las HRs de producción ──────
        hr_capacidades = {}   # {id_hruta: materiales_actuales}
        with get_conn() as cn:
            cur = cn.cursor()
            cur.execute("""
                SELECT ID_HRUTA, MATERIALES FROM dbo.ODATA_HR_CONSULTA
                WHERE TIPO_HR = 'PRODUCCION' AND ID_HRUTA IS NOT NULL
            """)
            for r in cur.fetchall():
                hr_capacidades[str(r[0]).strip()] = int(r[1] or 0)

        LIMITE = 300

        # ── Calcular plan: asignar ZFERs a HRs respetando límite ─────────────
        pendientes    = list(zfers_fuera)
        batches       = []
        sin_hr        = []
        hr_asignados  = {}   # {id_hruta: n_zfers_ya_asignados_en_este_plan}

        while pendientes:
            zfer = pendientes[0]
            attrs = attrs_map.get(zfer, {})

            # Buscar HR candidata — pasar valores RAW que _hr_construir_criterios necesita
            criterios = _hr_construir_criterios(
                attrs_base={
                    "Z_AGP_LEVEL":     attrs.get("Z_AGP_LEVEL", ""),
                    "Z_GEOMETRY_TYPE": attrs.get("Z_GEOMETRY_TYPE", ""),
                },
                area=attrs.get("area"),
                bom_posiciones=attrs.get("bom_posiciones", []),
                metrologia_base=None, prueba_agua_base=None
            )
            resultados, _, _, _ = _hr_buscar(criterios)

            # Filtrar HRs con capacidad disponible (excluir la HR origen)
            candidatas = []
            for r in resultados:
                hid = str(r.get("ID_HRUTA", "")).strip()
                if hid == str(id_hruta):
                    continue   # no reasignar a la misma HR que se desasignó
                actual    = hr_capacidades.get(hid, 0)
                en_plan   = hr_asignados.get(hid, 0)
                disponible = LIMITE - actual - en_plan
                if disponible > 0:
                    candidatas.append((hid, r.get("DESCRIPCION",""), disponible))

            if not candidatas:
                sin_hr.append(zfer)
                pendientes.pop(0)
                continue

            # Elegir la HR con más materiales actuales (la más llena que aún cabe)
            candidatas.sort(key=lambda x: hr_capacidades.get(x[0], 0), reverse=True)
            hr_elegida, hr_desc, disponible = candidatas[0]

            # Tomar los primeros `disponible` ZFERs pendientes para esta HR
            lote = pendientes[:disponible]
            pendientes = pendientes[disponible:]

            hr_asignados[hr_elegida] = hr_asignados.get(hr_elegida, 0) + len(lote)
            batches.append({
                "hr_destino":  hr_elegida,
                "hr_desc":     hr_desc,
                "materiales_actuales": hr_capacidades.get(hr_elegida, 0),
                "n_zfers":     len(lote),
                "zfers":       lote,
                "zfor_map":    {z: zfor_map.get(z, "") for z in lote},
            })

        sin_bom = [z for z in zfers_fuera if attrs_map.get(z, {}).get("sin_bom")]

        return jsonify({
            "ok": True,
            "id_hruta_origen": id_hruta,
            "total_zfers":     len(zfers_fuera),
            "asignables":      sum(b["n_zfers"] for b in batches),
            "sin_hr":          sin_hr,
            "sin_bom":         sin_bom,
            "batches":         batches,
        })
    except Exception as e:
        import traceback
        print(f"[MHR-PLAN] {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/descargar_plan_asignacion/<id_hruta>", methods=["POST"])
@login_required
def api_mhr_descargar_plan_asignacion(id_hruta: str):
    """Genera el Excel del plan de asignación y lo descarga SIN enviar a SAP."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from io import BytesIO

    body    = request.get_json() or {}
    batches = body.get("batches", [])
    if not batches:
        return jsonify({"ok": False, "error": "Plan vacío."})

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asignacion"

        hdr_fill  = PatternFill("solid", fgColor="1a4731")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        fills     = [PatternFill("solid", fgColor=c) for c in
                     ("c6efce","dae8fc","ffe6cc","e1d5e7","fff2cc","f8cecc","d5e8d4","dae3f3")]

        for col, txt in enumerate(["Grupo de hoja de ruta", "Contador HR", "Material"], 1):
            c = ws.cell(row=1, column=col, value=txt)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

        row_i = 2
        for bi, batch in enumerate(batches):
            fill   = fills[bi % len(fills)]
            hr_id  = str(batch["hr_destino"])
            zfers  = batch["zfers"]
            zfor_m = batch.get("zfor_map", {})
            for zfer in zfers:
                ws.cell(row=row_i, column=1, value=hr_id).fill = fill
                ws.cell(row=row_i, column=2, value="01").fill = fill
                ws.cell(row=row_i, column=3, value=zfer).fill = fill
                row_i += 1
                zfor = zfor_m.get(zfer, "")
                if zfor:
                    ws.cell(row=row_i, column=1, value=hr_id).fill = fill
                    ws.cell(row=row_i, column=2, value="01").fill = fill
                    ws.cell(row=row_i, column=3, value=zfor).fill = fill
                    row_i += 1

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        from datetime import datetime as _dt
        fname = f"plan_asignacion_{id_hruta}_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/ejecutar_asignacion", methods=["POST"])
@login_required
def api_mhr_ejecutar_asignacion():
    """
    Recibe el plan completo (todos los batches) y genera UN solo Excel
    con todos los ZFERs — el campo 'Grupo HR' cambia según el batch.
    Llama ZPPP0084 Asignar una sola vez con ese Excel.
    """
    body = request.get_json() or {}
    batches = body.get("batches", [])
    id_hruta_origen = str(body.get("id_hruta_origen", "plan"))

    if not batches:
        return jsonify({"ok": False, "error": "Plan vacío."})

    try:
        import importlib, openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        sap_mant = importlib.import_module("sap_mantenimiento")
        fn = getattr(sap_mant, "zppp0084_asignar", None)
        if not fn:
            return jsonify({"ok": False, "error": "zppp0084_asignar no encontrada."})

        # ── Generar Excel único con todos los batches ────────────────────────
        total_zfers = sum(b["n_zfers"] for b in batches)
        excel_path  = os.path.join(_MHR_TEMP_DIR, f"asignar_plan_{id_hruta_origen}.xlsx")

        try:
            if os.path.exists(excel_path):
                os.remove(excel_path)
        except OSError as e:
            return jsonify({"ok": False, "error": f"No se puede sobrescribir Excel: {e}"})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asignacion"
        hdr_fill = PatternFill("solid", fgColor="1a4731")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        fill_ok  = PatternFill("solid", fgColor="c6efce")

        for col, txt in enumerate(["Grupo de hoja de ruta", "Contador HR", "Material"], 1):
            c = ws.cell(row=1, column=col, value=txt)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

        row_i = 2
        for batch in batches:
            hr_id    = str(batch["hr_destino"])
            zfers    = batch["zfers"]
            zfor_map = batch.get("zfor_map", {})
            for zfer in zfers:
                ws.cell(row=row_i, column=1, value=hr_id).fill = fill_ok
                ws.cell(row=row_i, column=2, value="01").fill = fill_ok
                ws.cell(row=row_i, column=3, value=zfer).fill = fill_ok
                row_i += 1
                zfor = zfor_map.get(zfer, "")
                if zfor:
                    ws.cell(row=row_i, column=1, value=hr_id).fill = fill_ok
                    ws.cell(row=row_i, column=2, value="01").fill = fill_ok
                    ws.cell(row=row_i, column=3, value=zfor).fill = fill_ok
                    row_i += 1

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        wb.save(excel_path)
        print(f"[MHR-ASIGNAR] Excel generado: {excel_path} ({row_i-2} filas, {total_zfers} ZFERs)")

        # ── Una sola llamada a ZPPP0084 ──────────────────────────────────────
        resultado = fn(excel_path)
        return jsonify({**resultado,
                        "n_zfers": total_zfers,
                        "n_batches": len(batches),
                        "excel_path": excel_path})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/ejecutar_c223", methods=["POST"])
@login_required
def api_mhr_ejecutar_c223():
    """Ejecuta C223 para un batch: lista de ZFERs + HR destino."""
    body    = request.get_json() or {}
    zfers   = body.get("zfers", [])
    hr_id   = str(body.get("hr_id", ""))
    if not zfers or not hr_id:
        return jsonify({"ok": False, "error": "Faltan zfers o hr_id."})
    try:
        import importlib
        fn = getattr(importlib.import_module("sap_mantenimiento"), "c223_mantenimiento", None)
        if not fn:
            return jsonify({"ok": False, "error": "c223_mantenimiento no encontrada."})
        return jsonify(fn(zfers, hr_id))
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/ejecutar_zingp0004", methods=["POST"])
@login_required
def api_mhr_ejecutar_zingp0004():
    """Ejecuta ZINGP0004. Body opcional: {zfers: [...]} — si vacío ejecuta para todos."""
    body  = request.get_json() or {}
    zfers = body.get("zfers") or None
    try:
        import importlib
        fn = getattr(importlib.import_module("sap_mantenimiento"), "zinpg0004_actualizar", None)
        if not fn:
            return jsonify({"ok": False, "error": "zinpg0004_actualizar no encontrada."})
        return jsonify(fn(zfers))
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/reporte_flujo", methods=["POST"])
@login_required
def api_mhr_reporte_flujo():
    """Genera Excel con el reporte del flujo completo de mantenimiento."""
    import openpyxl, json as _json
    from openpyxl.styles import PatternFill, Font, Alignment
    from io import BytesIO
    from datetime import datetime as _dt
    body = request.get_json() or {}
    filas = body.get("filas", [])   # [{zfer, hr_origen, hr_destino, desasignar, asignar, c223, zingp, estado}]
    if not filas:
        return jsonify({"ok": False, "error": "Sin datos para reporte."})
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Mantenimiento HR"

        hdrs = ["ZFER", "ZFOR", "HR Origen", "HR Destino", "Desasignar", "Asignar", "C223", "ZINGP0004", "Estado"]
        hdr_fill = PatternFill("solid", fgColor="1a4731")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        fill_ok  = PatternFill("solid", fgColor="c6efce")
        fill_err = PatternFill("solid", fgColor="ffc7ce")
        fill_warn= PatternFill("solid", fgColor="fff2cc")

        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

        for ri, f in enumerate(filas, 2):
            estado = f.get("estado", "")
            fill   = fill_ok if estado == "OK" else (fill_err if estado == "ERROR" else fill_warn)
            vals   = [f.get("zfer",""), f.get("zfor",""), f.get("hr_origen",""), f.get("hr_destino",""),
                      f.get("desasignar",""), f.get("asignar",""),
                      f.get("c223",""), f.get("zingp",""), estado]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=ri, column=ci, value=v).fill = fill

        widths = [18,18,14,14,12,12,12,12,10]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1,ci).column_letter].width = w

        buf = BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"reporte_mant_hr_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/mantenimiento_hr/buscar_zfer")
@login_required
def api_mantenimiento_hr_buscar_zfer():
    """Busca un ZFER en el último resultado: indica si está fuera o en calendario y en qué HR."""
    import json as _json
    zfer = request.args.get("zfer", "").strip()
    if not zfer:
        return jsonify({"ok": False, "error": "Falta parámetro zfer"})
    try:
        if not os.path.exists(_MHR_JSON):
            return jsonify({"ok": False, "error": "No hay datos. Ejecuta la consulta primero."})
        with open(_MHR_JSON, "r", encoding="utf-8") as f:
            data = _json.load(f)
        if not data.get("ok"):
            return jsonify({"ok": False, "error": "Datos inválidos. Vuelve a consultar."})

        resultados = []
        for hr in data["hojas_ruta"]:
            fuera = set(hr.get("zfers_fuera", []))
            en_cal = hr.get("en_calendario", 0)
            total  = hr.get("total_zfer", 0)
            if zfer in fuera:
                resultados.append({"id_hruta": hr["id_hruta"], "descripcion": hr["descripcion"],
                                    "estado": "fuera"})
            elif total - en_cal < total:  # puede estar en calendario en esta HR
                # No tenemos lista de los que SÍ están, pero si no está en fuera y total > 0
                # significa que puede estar. Marcamos como "en_calendario" tentativamente
                # Solo lo reportamos si el ZFER efectivamente pertenece a esta HR
                # (no tenemos la lista completa de en_calendario, solo la de fuera)
                pass  # no podemos confirmar sin lista completa

        if resultados:
            return jsonify({"ok": True, "zfer": zfer, "encontrado": True,
                            "estado": "fuera", "hrs": resultados,
                            "fecha_consulta": data.get("fecha_consulta","")})
        else:
            return jsonify({"ok": True, "zfer": zfer, "encontrado": False,
                            "estado": "en_calendario_o_no_asignado",
                            "mensaje": "No aparece como FUERA en ninguna HR analizada.",
                            "fecha_consulta": data.get("fecha_consulta","")})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.after_request
def _set_content_length(response):
    # Prevent browser tab from showing infinite spinner on dynamic HTML pages
    if response.content_type and "text/html" in response.content_type:
        if not response.is_streamed:
            response.headers["Content-Length"] = len(response.get_data())
    return response


if __name__ == "__main__":
    print("\n  AGP Intelligence — MODULO 5")
    print("  Abre tu navegador en: http://localhost:5000\n")
    app.run(debug=True, host="0.0.0.0", port=5000, use_reloader=False)
